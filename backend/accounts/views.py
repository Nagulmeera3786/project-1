from rest_framework import status, generics, permissions
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.conf import settings
from django.utils import timezone
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db import transaction
from django.db import close_old_connections
from django.db.models import Q, Count, F
from django.core.mail import get_connection, EmailMessage
from django.core.validators import validate_email
from decimal import Decimal, InvalidOperation
from django.http import HttpResponseRedirect, Http404
from django.views import View
import requests
import time
import re
import difflib
import io
import smtplib
import hmac
import hashlib
import random
import secrets
import string
import uuid
import socket
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone as dt_timezone, timedelta
from zoneinfo import ZoneInfo, available_timezones
from urllib.parse import urlparse

try:
    import dns.resolver
except Exception:  # pragma: no cover - optional import in dev
    dns = None
from .serializers import (
    SignupSerializer, OTPVerifySerializer, LoginSerializer,
    ForgotPasswordSerializer, ResetPasswordSerializer,
    SMSMessageSerializer, SMSSendSerializer, SMSCredentialSerializer,
    UserSMSEligibilitySerializer, SMSMessageStatusSerializer,
    SMSContactGroupSerializer, SMSContactGroupCreateSerializer,
    SMSShortURLSerializer,
    NotificationRecipientPreviewSerializer,
    AdminNotificationSendSerializer,
    AdminNotificationHistorySerializer,
    UserNotificationSerializer,
    UserWalletSerializer,
    WalletRechargePaymentSerializer,
    WalletRechargeCreateOrderSerializer,
    WalletRechargeVerifySerializer,
    PlatformSettingSerializer,
    UserAPIKeySerializer,
    AdminUserAPIKeySerializer,
    EmailValidationHistorySerializer,
    SenderIdRequestSerializer,
    SenderIdRequestAdminSerializer,
    EmployeeSignupSerializer,
    EmployeeVerifySerializer,
    EmployeeLoginSerializer,
    EmployeeSerializer,
)
from .utils import generate_otp, send_otp_via_email, otp_is_valid, calculate_sms_segments
from .serializers import EmployeeSignupSerializer, EmployeeDualOTPVerifySerializer, EmployeeLoginSerializer
from django.contrib.auth import get_user_model
from rest_framework_simplejwt.tokens import RefreshToken
from .models import Employee
from .models import (
    SMSMessage,
    SMSCredential,
    SMSContactGroup,
    SMSContact,
    SMSShortURL,
    FreeTrialVerifiedNumber,
    InternalNotification,
    InternalNotificationRecipient,
    UserWallet,
    WalletRechargePayment,
    PlatformSetting,
    UserAPIKey,
    EmailValidationHistory,
    SenderIdRequest,
    Employee,
)

User = get_user_model()

_COMMON_FREE_EMAIL_DOMAINS = {
    'gmail.com', 'googlemail.com', 'yahoo.com', 'yahoo.co.in', 'yahoo.co.uk',
    'outlook.com', 'hotmail.com', 'live.com', 'msn.com', 'icloud.com',
    'me.com', 'aol.com', 'proton.me', 'protonmail.com', 'zoho.com',
    'gmx.com', 'gmx.net', 'mail.com', 'yandex.com', 'yandex.ru',
    'rediffmail.com', 'mailinator.com', 'tempmail.com', 'guerrillamail.com',
}

_EMAIL_VALIDATION_WORKERS = {}
_EMAIL_VALIDATION_WORKERS_LOCK = threading.Lock()
_MX_CACHE = {}
_MX_CACHE_LOCK = threading.Lock()
_HISTORY_SIGNAL_CACHE = {}
_HISTORY_SIGNAL_CACHE_LOCK = threading.Lock()

_POPULAR_MAIL_DOMAINS = {
    'gmail.com', 'googlemail.com', 'yahoo.com', 'outlook.com', 'hotmail.com', 'live.com',
    'icloud.com', 'zoho.com', 'proton.me', 'protonmail.com', 'gmx.com', 'aol.com',
}

_POPULAR_DOMAIN_TYPOS = {
    'gmil.com': 'gmail.com',
    'gamil.com': 'gmail.com',
    'gmai.com': 'gmail.com',
    'gmail.co': 'gmail.com',
    'hmail.com': 'hotmail.com',
    'hotnail.com': 'hotmail.com',
    'outlok.com': 'outlook.com',
    'otlook.com': 'outlook.com',
    'yaho.com': 'yahoo.com',
    'yhoo.com': 'yahoo.com',
    'zho.com': 'zoho.com',
    'zohoo.com': 'zoho.com',
    'protonmai.com': 'protonmail.com',
    'protonmaill.com': 'protonmail.com',
}


def _is_celery_enabled():
    return bool(getattr(settings, 'EMAIL_VALIDATION_USE_CELERY', False))


def _get_celery_task_state(history):
    summary = _get_history_summary(history)
    task_id = str(summary.get('celery_task_id') or '').strip()
    if not task_id:
        return ''

    try:
        from celery.result import AsyncResult

        result = AsyncResult(task_id)
        return str(result.state or '').strip().upper()
    except Exception:
        return ''


def _is_celery_task_active(history):
    state = _get_celery_task_state(history)
    return state in {'PENDING', 'RECEIVED', 'STARTED', 'RETRY'}


def _revoke_celery_task(history):
    summary = _get_history_summary(history)
    task_id = str(summary.get('celery_task_id') or '').strip()
    if not task_id:
        return

    try:
        from project.celery import app as celery_app

        celery_app.control.revoke(task_id, terminate=False)
    except Exception:
        pass


def _is_primary_admin_email(email):
    return email.strip().lower() == getattr(settings, 'PRIMARY_ADMIN_EMAIL', '').strip().lower()


def _has_primary_admin_access(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    return bool(
        _is_primary_admin_email(getattr(user, 'email', '') or '')
        and getattr(user, 'is_active', False)
        and getattr(user, 'is_staff', False)
        and getattr(user, 'is_superuser', False)
    )


def _is_active_employee(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    return Employee.objects.filter(user=user, status=Employee.STATUS_ACTIVE).exists()


def _has_admin_access(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    return bool(
        getattr(user, 'is_active', False)
        and (
            _has_primary_admin_access(user)
            or getattr(user, 'is_staff', False)
            or getattr(user, 'is_superuser', False)
        )
    )


def _has_support_read_access(user):
    return _has_admin_access(user) or _is_active_employee(user)


def _primary_admin_guard(request):
    if not getattr(settings, 'PRIMARY_ADMIN_ENFORCEMENT', not getattr(settings, 'DEBUG', False)):
        return None
    if _has_admin_access(request.user):
        return None
    return Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)


def _get_primary_admin_mediator():
    primary_email = str(getattr(settings, 'PRIMARY_ADMIN_EMAIL', '') or '').strip()
    if not primary_email:
        return None

    return (
        User.objects
        .filter(
            email__iexact=primary_email,
            is_active=True,
            is_staff=True,
            is_superuser=True,
        )
        .order_by('-id')
        .first()
    )


def _promote_primary_admin(user):
    if _is_primary_admin_email(user.email):
        changed = False
        if not user.is_staff:
            user.is_staff = True
            changed = True
        if not user.is_superuser:
            user.is_superuser = True
            changed = True
        if not user.is_active:
            user.is_active = True
            changed = True
        if not user.is_sms_enabled:
            user.is_sms_enabled = True
            changed = True
        if changed:
            user.save()


def _find_user_by_email(email):
    normalized_email = (email or '').strip().lower()
    if not normalized_email:
        return None
    return User.objects.filter(email__iexact=normalized_email).order_by('-is_active', '-id').first()


def _otp_email_diagnostics(email_sent):
    configured_provider = str(getattr(settings, 'EMAIL_PROVIDER', '') or '').strip().lower()
    backend = str(getattr(settings, 'EMAIL_BACKEND', '') or '').strip()
    host = str(getattr(settings, 'EMAIL_HOST', '') or '').strip()

    if configured_provider:
        provider = configured_provider
    elif 'sendgrid' in host.lower():
        provider = 'sendgrid'
    elif 'mailgun' in host.lower():
        provider = 'mailgun'
    elif 'gmail' in host.lower() or 'google' in host.lower():
        provider = 'gmail'
    elif host:
        provider = 'custom-smtp'
    else:
        provider = 'unknown'

    diagnostics = {
        'otp_generated': True,
        'email_delivery': {
            'sent': bool(email_sent),
            'provider': provider,
            'backend': backend,
            'host': host,
        },
    }

    if not email_sent:
        diagnostics['error_code'] = 'OTP_EMAIL_DELIVERY_FAILED'
        diagnostics['next_step'] = 'Check Render logs and provider credentials/API key.'

    return diagnostics


def _get_sms_provider_config():
    cred = SMSCredential.objects.filter(is_active=True).order_by('-updated_at', '-id').first()
    if cred:
        db_user = (cred.user or '').strip()
        db_password = (cred.password or '').strip()
        if db_user and db_password:
            db_sender_ids = [str(item).strip() for item in (cred.sender_ids or []) if str(item).strip()]
            return {
                'credential': cred,
                'user': db_user,
                'password': db_password,
                'sender_ids': db_sender_ids,
                'source': 'database',
            }

    env_user = getattr(settings, 'SMS_PROVIDER_USER', '').strip()
    env_password = getattr(settings, 'SMS_PROVIDER_PASSWORD', '').strip()
    sender_ids = [str(item).strip() for item in getattr(settings, 'SMS_DEFAULT_SENDER_IDS', []) if str(item).strip()]
    if env_user and env_password:
        return {
            'credential': None,
            'user': env_user,
            'password': env_password,
            'sender_ids': sender_ids,
            'source': 'env',
        }

    return None


def _get_admin_managed_sms_provider_config():
    cred = SMSCredential.objects.filter(is_active=True).order_by('-updated_at', '-id').first()
    if cred:
        db_user = (cred.user or '').strip()
        db_password = (cred.password or '').strip()
        db_sender_ids = [str(item).strip() for item in (cred.sender_ids or []) if str(item).strip()]
        db_free_trial_sender_id = str(getattr(cred, 'free_trial_default_sender_id', '') or '').strip()
        if db_user and db_password:
            return {
                'credential': cred,
                'user': db_user,
                'password': db_password,
                'sender_ids': db_sender_ids,
                'free_trial_default_sender_id': db_free_trial_sender_id,
                'source': 'database',
            }

    env_user = getattr(settings, 'SMS_PROVIDER_USER', '').strip()
    env_password = getattr(settings, 'SMS_PROVIDER_PASSWORD', '').strip()
    env_sender_ids = [str(item).strip() for item in getattr(settings, 'SMS_DEFAULT_SENDER_IDS', []) if str(item).strip()]
    env_free_trial_sender_id = str(getattr(settings, 'SMS_FREE_TRIAL_DEFAULT_SENDER_ID', '') or '').strip()
    if env_user and env_password:
        return {
            'credential': None,
            'user': env_user,
            'password': env_password,
            'sender_ids': env_sender_ids,
            'free_trial_default_sender_id': env_free_trial_sender_id,
            'source': 'env',
        }

    return None


def _get_free_trial_mediator_user():
    primary_admin_email = getattr(settings, 'PRIMARY_ADMIN_EMAIL', '').strip().lower()
    if primary_admin_email:
        primary_admin = User.objects.filter(
            email__iexact=primary_admin_email,
            is_staff=True,
            is_active=True,
        ).first()
        if primary_admin:
            return primary_admin

    return User.objects.filter(
        email__iexact=getattr(settings, 'PRIMARY_ADMIN_EMAIL', '').strip().lower(),
        is_staff=True,
        is_superuser=True,
        is_active=True,
    ).order_by('id').first()


def _resolve_free_trial_sender_id(user, provider_config):
    available_sender_ids = [
        str(item).strip()
        for item in (provider_config.get('sender_ids') or [])
        if str(item).strip()
    ]
    if not available_sender_ids:
        raise ValueError('Free trial sender configuration is unavailable')

    trial_sender_id = str(provider_config.get('free_trial_default_sender_id') or '').strip()
    user_sender_id = str(getattr(user, 'free_trial_sender_id', '') or '').strip() if user else ''

    if trial_sender_id and trial_sender_id in available_sender_ids:
        resolved_sender_id = trial_sender_id
    elif user_sender_id and user_sender_id in available_sender_ids:
        resolved_sender_id = user_sender_id
    else:
        # Auto-fallback lets normal users use free-trial flow even if admin skipped selecting a dedicated default.
        resolved_sender_id = available_sender_ids[0]

    if user and not _has_primary_admin_access(user) and user_sender_id != resolved_sender_id:
        try:
            user.free_trial_sender_id = resolved_sender_id
            user.save(update_fields=['free_trial_sender_id'])
        except Exception:
            pass

    return resolved_sender_id


def _normalize_sender_id(sender_id_type, sender_id):
    normalized_type = (sender_id_type or 'alphanumeric').strip().lower()
    if normalized_type not in ['numeric', 'alphanumeric']:
        raise ValueError('Invalid sender ID type. Allowed values: numeric, alphanumeric')

    raw_sender_id = (sender_id or '').strip()
    if not raw_sender_id:
        return normalized_type, ''

    if normalized_type == 'numeric':
        if not raw_sender_id.isdigit():
            raise ValueError('Numeric sender ID must contain only digits')
        if len(raw_sender_id) < 10 or len(raw_sender_id) > 15:
            raise ValueError('Numeric sender ID length must be between 10 and 15 digits')
        return normalized_type, raw_sender_id

    normalized_sender_id = raw_sender_id.upper()
    if not re.fullmatch(r'[A-Z0-9]+', normalized_sender_id):
        raise ValueError('Alphanumeric sender ID can contain only letters and numbers')
    if len(normalized_sender_id) < 3 or len(normalized_sender_id) > 11:
        raise ValueError('Alphanumeric sender ID length must be between 3 and 11 characters')
    return normalized_type, normalized_sender_id


def _sender_id_exists(sender_id, exclude_user_id=None):
    if not sender_id:
        return False
    queryset = User.objects.filter(sender_id__iexact=sender_id)
    if exclude_user_id:
        queryset = queryset.exclude(id=exclude_user_id)
    return queryset.exists()


def _build_sender_id_suggestions(sender_id, sender_id_type, exclude_user_id=None, limit=4):
    normalized_type = (sender_id_type or 'alphanumeric').strip().lower()
    existing_sender_ids = set(
        User.objects.exclude(sender_id__isnull=True)
        .exclude(sender_id='')
        .exclude(id=exclude_user_id if exclude_user_id else -1)
        .values_list('sender_id', flat=True)
    )
    existing_sender_ids = {str(item).upper() for item in existing_sender_ids}

    suggestions = []

    if normalized_type == 'numeric':
        base_digits = ''.join(ch for ch in (sender_id or '') if ch.isdigit())
        while len(base_digits) < 10:
            base_digits += str(random.randint(0, 9))
        base_prefix = base_digits[:10]

        for index in range(1, 50):
            candidate = f"{base_prefix}{index:02d}"[:15]
            if candidate.upper() not in existing_sender_ids and candidate not in suggestions:
                suggestions.append(candidate)
            if len(suggestions) >= limit:
                break
    else:
        clean_base = ''.join(ch for ch in (sender_id or '').upper() if ch.isalnum())
        if not clean_base:
            clean_base = 'SENDER'
        base_prefix = clean_base[:8]

        for index in range(1, 70):
            suffix = f"{index:02d}"
            candidate = f"{base_prefix}{suffix}"[:11]
            if len(candidate) < 3:
                extra = ''.join(random.choice(string.ascii_uppercase) for _ in range(3 - len(candidate)))
                candidate = f"{candidate}{extra}"
            if candidate.upper() not in existing_sender_ids and candidate not in suggestions:
                suggestions.append(candidate)
            if len(suggestions) >= limit:
                break

    return suggestions


def _normalize_phone_number(raw_value):
    digits = ''.join(ch for ch in str(raw_value or '') if ch.isdigit())
    return digits if len(digits) >= 10 else ''


def _render_personalized_template(template_text, row_values):
    rendered = str(template_text or '')
    for index, value in enumerate(row_values, start=1):
        rendered = rendered.replace(f'#{index}#', str(value if value is not None else '').strip())
    return rendered.strip()


def _extract_phone_from_row(row_values):
    for value in row_values:
        normalized = _normalize_phone_number(value)
        if normalized:
            return normalized
    return ''


def _generate_short_code(length=7):
    alphabet = string.ascii_letters + string.digits
    return ''.join(random.choice(alphabet) for _ in range(length))


def _format_utc_offset(offset_delta):
    if offset_delta is None:
        return 'UTC+00:00', '+0.00', 0

    total_minutes = int(offset_delta.total_seconds() // 60)
    sign = '+' if total_minutes >= 0 else '-'
    abs_minutes = abs(total_minutes)
    hours, minutes = divmod(abs_minutes, 60)
    human = f'UTC{sign}{hours:02d}:{minutes:02d}'
    compact = f'{sign}{hours}.{minutes:02d}'
    return human, compact, total_minutes


FREE_TRIAL_MESSAGE_LIMIT = 3
FREE_TRIAL_OTP_EXPIRY_MINUTES = 10
_PROVIDER_BALANCE_CACHE = {'value': None, 'expires_at': None}
_EMAIL_PROVIDER_CREDITS_CACHE = {'value': None, 'expires_at': None}


def _extract_first_numeric_value(text):
    try:
        match = re.search(r'-?\d+(?:\.\d+)?', str(text or ''))
        if not match:
            return None
        return float(match.group(0))
    except Exception:
        return None


def _get_balance_candidate_urls(provider_config):
    configured_balance_url = str(getattr(settings, 'SMS_PROVIDER_BALANCE_URL', '') or '').strip()
    if configured_balance_url:
        return [configured_balance_url]

    send_url = str(getattr(settings, 'SMS_PROVIDER_URL', '') or '').strip().lower()
    if 'infobip.com' in send_url:
        parsed_url = urlparse(send_url)
        if parsed_url.scheme and parsed_url.netloc:
            return [f'{parsed_url.scheme}://{parsed_url.netloc}/account/1/balance']

    if 'mshastra.com' in send_url:
        return [
            'https://mshastra.com/bsms/buser/balance.aspx',
            'https://mshastra.com/bsms/buser/get_balance.aspx',
            'https://mshastra.com/bsms/buser/user_balance.aspx',
        ]

    return []


def _is_indian_number(number):
    digits = _normalize_phone_number(number)
    if not digits:
        return False
    if len(digits) == 10 and digits.startswith(('6', '7', '8', '9')):
        return True
    if len(digits) == 12 and digits.startswith('91'):
        return True
    return False


def _map_verifalia_quality(classification, status_text=''):
    normalized = str(classification or '').strip().lower()
    normalized_status = str(status_text or '').strip().lower()

    if normalized in ['deliverable', 'safe']:
        return {
            'quality': 'deliverable',
            'is_deliverable': True,
            'is_risky': False,
        }

    if normalized in ['undeliverable', 'invalid']:
        return {
            'quality': 'invalid',
            'is_deliverable': False,
            'is_risky': False,
        }

    if normalized in ['risky', 'unknown'] or normalized_status in ['inprogress', 'waiting', 'queued']:
        return {
            'quality': 'risky',
            'is_deliverable': False,
            'is_risky': True,
        }

    return {
        'quality': 'unknown',
        'is_deliverable': False,
        'is_risky': True,
    }


def _get_email_validation_max_file_size_bytes():
    try:
        configured_mb = int(getattr(settings, 'EMAIL_VALIDATION_MAX_FILE_SIZE_MB', 25) or 25)
    except (TypeError, ValueError):
        configured_mb = 25
    return max(1, configured_mb) * 1024 * 1024


def _get_email_validation_max_request_count():
    try:
        configured_limit = int(getattr(settings, 'EMAIL_VALIDATION_MAX_EMAILS_PER_REQUEST', 5000) or 5000)
    except (TypeError, ValueError):
        configured_limit = 5000
    return max(1, configured_limit)


def _get_email_validation_batch_size():
    try:
        configured_batch_size = int(getattr(settings, 'EMAIL_VALIDATION_BATCH_SIZE', 500) or 500)
    except (TypeError, ValueError):
        configured_batch_size = 500
    return max(1, min(configured_batch_size, 1000))


def _get_email_validation_provider_mode():
    setting = PlatformSetting.objects.filter(key='email_validation_provider_mode').first()
    configured = str(setting.value if setting else getattr(settings, 'EMAIL_VALIDATION_PROVIDER_MODE', 'own_system') or '').strip().lower()
    if configured not in {'own_system', 'zerobounce'}:
        return 'own_system'
    return configured


def _get_email_validation_worker_count():
    try:
        configured = int(getattr(settings, 'EMAIL_VALIDATION_MAX_WORKERS', 64) or 64)
    except (TypeError, ValueError):
        configured = 64
    return max(8, min(configured, 256))


def _get_smtp_retry_attempts():
    try:
        configured = int(getattr(settings, 'EMAIL_VALIDATION_SMTP_RETRY_ATTEMPTS', 2) or 2)
    except (TypeError, ValueError):
        configured = 2
    return max(1, min(configured, 4))


def _get_smtp_retry_backoff_seconds():
    try:
        configured = float(getattr(settings, 'EMAIL_VALIDATION_SMTP_RETRY_BACKOFF_SECONDS', 0.8) or 0.8)
    except (TypeError, ValueError):
        configured = 0.8
    return max(0.2, min(configured, 5.0))


def _get_smtp_mail_from_pool():
    configured_pool = getattr(settings, 'EMAIL_VALIDATION_SMTP_MAIL_FROM_POOL', None)
    if isinstance(configured_pool, (list, tuple, set)):
        normalized_pool = [str(item or '').strip().lower() for item in configured_pool if str(item or '').strip()]
        if normalized_pool:
            return normalized_pool[:5]
    return ['validator@bhisha.com', 'postmaster@bhisha.com']


def _detect_popular_domain_typo(domain):
    normalized = str(domain or '').strip().lower()
    if not normalized:
        return ''

    direct = _POPULAR_DOMAIN_TYPOS.get(normalized)
    if direct:
        return direct

    close = difflib.get_close_matches(normalized, list(_POPULAR_MAIL_DOMAINS), n=1, cutoff=0.86)
    return close[0] if close else ''


def _extract_email_from_history_result_item(item):
    if not isinstance(item, dict):
        return ''
    if isinstance(item.get('bhisha_result'), dict):
        return str(item['bhisha_result'].get('email') or '').strip().lower()
    return str(item.get('email') or '').strip().lower()


def _extract_inbox_valid_from_history_result_item(item):
    if not isinstance(item, dict):
        return None
    if isinstance(item.get('bhisha_result'), dict):
        return bool(item['bhisha_result'].get('valid_inbox'))
    if 'valid_inbox' in item:
        return bool(item.get('valid_inbox'))
    valid_syntax = bool(item.get('validSyntax')) if 'validSyntax' in item else None
    valid_mailbox = bool(item.get('validMailbox')) if 'validMailbox' in item else None
    if valid_syntax is None or valid_mailbox is None:
        return None
    return bool(valid_syntax and valid_mailbox)


def _get_historical_verification_signal(email):
    normalized = str(email or '').strip().lower()
    if not normalized:
        return {
            'sample_size': 0,
            'valid_inbox_rate': None,
            'last_seen_at': '',
        }

    now_ts = time.time()
    with _HISTORY_SIGNAL_CACHE_LOCK:
        cached = _HISTORY_SIGNAL_CACHE.get(normalized)
        if cached and cached.get('expires_at', 0) > now_ts:
            return dict(cached.get('value') or {})

    sample_size = 0
    valid_count = 0
    last_seen_at = ''

    try:
        histories = EmailValidationHistory.objects.filter(
            status=EmailValidationHistory.STATUS_COMPLETED
        ).order_by('-created_at')[:250]

        for history in histories:
            summary = _get_history_summary(history)
            rows = summary.get('results') if isinstance(summary.get('results'), list) else []
            for item in rows:
                row_email = _extract_email_from_history_result_item(item)
                if row_email != normalized:
                    continue
                inbox_valid = _extract_inbox_valid_from_history_result_item(item)
                if inbox_valid is None:
                    continue
                sample_size += 1
                if inbox_valid:
                    valid_count += 1
                if not last_seen_at:
                    last_seen_at = str(getattr(history, 'created_at', '') or '')
                if sample_size >= 30:
                    break
            if sample_size >= 30:
                break
    except Exception:
        sample_size = 0
        valid_count = 0
        last_seen_at = ''

    valid_inbox_rate = None if sample_size == 0 else round(valid_count / sample_size, 4)
    signal = {
        'sample_size': sample_size,
        'valid_inbox_rate': valid_inbox_rate,
        'last_seen_at': last_seen_at,
    }

    with _HISTORY_SIGNAL_CACHE_LOCK:
        _HISTORY_SIGNAL_CACHE[normalized] = {
            'value': signal,
            'expires_at': now_ts + 300,
        }

    return signal


def _compute_domain_reputation(domain, *, disposable=False, typo_suggestion='', status_code='', history_signal=None):
    normalized = str(domain or '').strip().lower()
    code = str(status_code or '').strip().upper()
    history_signal = history_signal or {}

    score = 50
    reasons = []

    if normalized in _POPULAR_MAIL_DOMAINS:
        score += 20
        reasons.append('popular_provider')

    if disposable:
        score -= 45
        reasons.append('disposable_domain')

    if typo_suggestion:
        score -= 40
        reasons.append('likely_domain_typo')

    if code in {'DOMAIN_NOT_FOUND', 'NO_MX'}:
        score -= 35
        reasons.append(code.lower())
    elif code in {'DNS_LOOKUP_FAILED', 'DNS_UNAVAILABLE'}:
        score -= 20
        reasons.append(code.lower())

    valid_rate = history_signal.get('valid_inbox_rate')
    sample_size = int(history_signal.get('sample_size') or 0)
    if valid_rate is not None and sample_size >= 3:
        if valid_rate >= 0.7:
            score += 10
            reasons.append('historical_success')
        elif valid_rate <= 0.3:
            score -= 10
            reasons.append('historical_failures')

    score = max(0, min(score, 100))
    if score >= 70:
        tier = 'high'
    elif score >= 40:
        tier = 'medium'
    else:
        tier = 'low'

    return {
        'score': score,
        'tier': tier,
        'reasons': reasons,
    }


def _resolve_mx_hosts_with_error(domain):
    domain = str(domain or '').strip().lower()
    if not domain:
        return [], 'INVALID_DOMAIN'

    now_ts = time.time()
    with _MX_CACHE_LOCK:
        cached = _MX_CACHE.get(domain)
        if cached and cached['expires_at'] > now_ts:
            return list(cached.get('mx_hosts') or []), ''

    if dns is None:
        return [], 'DNS_UNAVAILABLE'

    try:
        records = dns.resolver.resolve(domain, 'MX')
        sorted_records = sorted(records, key=lambda record: int(getattr(record, 'preference', 0)))
        mx_hosts = [str(record.exchange).rstrip('.') for record in sorted_records if str(record.exchange).strip()]
        mx_error = ''
    except dns.resolver.NXDOMAIN:
        mx_hosts = []
        mx_error = 'DOMAIN_NOT_FOUND'
    except dns.resolver.NoAnswer:
        mx_hosts = []
        mx_error = 'NO_MX'
    except dns.resolver.NoNameservers:
        mx_hosts = []
        mx_error = 'DNS_NO_NAMESERVERS'
    except dns.exception.Timeout:
        mx_hosts = []
        mx_error = 'DNS_TIMEOUT'
    except Exception:
        mx_hosts = []
        mx_error = 'DNS_LOOKUP_FAILED'

    with _MX_CACHE_LOCK:
        _MX_CACHE[domain] = {
            'mx_hosts': mx_hosts,
            'expires_at': now_ts + 300,
        }

    return mx_hosts, mx_error


def _resolve_mx_hosts(domain):
    mx_hosts, _ = _resolve_mx_hosts_with_error(domain)
    return mx_hosts


def _domain_resolves(domain):
    domain = str(domain or '').strip().lower()
    if not domain:
        return False
    try:
        socket.getaddrinfo(domain, None)
        return True
    except Exception:
        return False


def _resolve_mx_host(domain):
    hosts = _resolve_mx_hosts(domain)
    return hosts[0] if hosts else ''


_EMAIL_REGEX = r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$'


def _build_validation_result(email, *, valid_syntax, valid_mailbox, catch_all=False, did_you_mean='', disposable=False, role_based=False, spam=False, risky=False, risk='low', provider_message_id='', status='Validation completed.', status_code='Success', classification='Deliverable', failure_reason=''):
    normalized_email = str(email or '').strip().lower()
    domain = normalized_email.split('@', 1)[1] if '@' in normalized_email else ''
    domain_related = bool(domain and domain not in _COMMON_FREE_EMAIL_DOMAINS)
    quality = 'deliverable' if (valid_syntax and valid_mailbox and not disposable and not role_based) else ('invalid' if not valid_syntax else 'risky')

    report_lines = [
        '#### Validation summary',
        f'Input data:**{normalized_email}**',
        f'Classification:{classification}',
        f'Status:{status}',
        f'Status code:{status_code}',
        '',
        '#### Validation report',
        f'Valid syntax: {str(bool(valid_syntax))}',
        f'Valid mailbox: {str(bool(valid_mailbox))}',
        f'Domain related mail: {str(domain_related)}',
        f'Catch-all domain: {str(bool(catch_all))}',
        f'Disposable address: {str(bool(disposable))}',
        f'Role-based address: {str(bool(role_based))}',
        f'Spam / do-not-mail: {str(bool(spam))}',
        f'Risk level: {risk}',
    ]

    return {
        'email': normalized_email,
        'domain': domain,
        'domainRelatedMail': domain_related,
        'validMailbox': bool(valid_mailbox),
        'validSyntax': bool(valid_syntax),
        'catchAll': bool(catch_all),
        'didYouMean': str(did_you_mean or normalized_email).strip(),
        'disposable': bool(disposable),
        'roleBased': bool(role_based),
        'spam': bool(spam),
        'risky': bool(risky),
        'risk': str(risk or 'unknown').strip().lower(),
        'providerMessageId': str(provider_message_id or '').strip(),
        'summary': f'{status} ({classification})',
        'report': '\n'.join(report_lines),
        'status': status,
        'statusCode': status_code,
        'classification': classification,
        'quality': quality,
        'failure_reason': failure_reason,
    }


def _validate_email_with_own_system_diagnostics(email):
    normalized = str(email or '').strip().lower()
    diagnostics = {
        'email': normalized,
        'provider': 'own_system',
        'syntax_valid': False,
        'mx_hosts': [],
        'mx_lookup_error': '',
        'domain_resolves': None,
        'domain_reputation': {},
        'historical_signal': {},
        'typo_suggestion': '',
        'heuristics': {},
        'retry_strategy': {
            'smtp_retry_attempts': _get_smtp_retry_attempts(),
            'mail_from_pool_size': len(_get_smtp_mail_from_pool()),
            'multi_ip_retry_supported': False,
            'multi_ip_note': 'Application retries across MX hosts and sender identities; multi-source-IP retry requires network-level egress infrastructure.',
        },
        'attempts': [],
        'final_status_code': '',
        'final_status': '',
    }

    if not re.fullmatch(_EMAIL_REGEX, normalized):
        result = _build_validation_result(
            normalized,
            valid_syntax=False,
            valid_mailbox=False,
            risky=True,
            risk='high',
            status='Invalid Email Format',
            status_code='INVALID_FORMAT',
            classification='Invalid',
            failure_reason='Invalid Email Format',
        )
        diagnostics['final_status_code'] = 'INVALID_FORMAT'
        diagnostics['final_status'] = 'Invalid Email Format'
        return result, diagnostics

    diagnostics['syntax_valid'] = True

    domain = normalized.split('@', 1)[1]
    typo_suggestion = _detect_popular_domain_typo(domain)
    if typo_suggestion and typo_suggestion != domain:
        diagnostics['typo_suggestion'] = typo_suggestion
        history_signal = _get_historical_verification_signal(normalized)
        reputation = _compute_domain_reputation(
            domain,
            typo_suggestion=typo_suggestion,
            status_code='INVALID_SYNTAX_DOMAIN_TYPO',
            history_signal=history_signal,
        )
        diagnostics['historical_signal'] = history_signal
        diagnostics['domain_reputation'] = reputation
        diagnostics['heuristics'] = {
            'confidence_score': max(0, reputation.get('score', 0) - 25),
            'decision_basis': ['syntax_regex', 'popular_domain_typo_detection'],
        }

        result = _build_validation_result(
            normalized,
            valid_syntax=False,
            valid_mailbox=False,
            did_you_mean=f"{normalized.split('@', 1)[0]}@{typo_suggestion}",
            risky=True,
            risk='high',
            status='Invalid domain spelling (popular provider typo detected)',
            status_code='INVALID_SYNTAX_DOMAIN_TYPO',
            classification='Invalid',
            failure_reason=f'Likely domain typo: {domain} -> {typo_suggestion}',
        )
        diagnostics['final_status_code'] = 'INVALID_SYNTAX_DOMAIN_TYPO'
        diagnostics['final_status'] = 'Invalid domain spelling (popular provider typo detected)'
        return result, diagnostics

    mx_hosts, mx_error = _resolve_mx_hosts_with_error(domain)
    diagnostics['mx_hosts'] = list(mx_hosts)
    diagnostics['mx_lookup_error'] = mx_error
    domain_resolves = _domain_resolves(domain)
    diagnostics['domain_resolves'] = domain_resolves

    if not mx_hosts:
        if mx_error == 'DOMAIN_NOT_FOUND' or not domain_resolves:
            status_text = 'Domain does not exist'
            status_code = 'DOMAIN_NOT_FOUND'
            failure_reason = 'Domain does not exist or cannot be resolved'
            risk_level = 'high'
        elif mx_error in {'DNS_UNAVAILABLE', 'DNS_NO_NAMESERVERS', 'DNS_TIMEOUT', 'DNS_LOOKUP_FAILED'}:
            status_text = 'Domain DNS lookup unavailable'
            status_code = 'DNS_LOOKUP_FAILED' if mx_error != 'DNS_UNAVAILABLE' else 'DNS_UNAVAILABLE'
            failure_reason = f'DNS lookup unavailable: {mx_error}'
            risk_level = 'medium'
        else:
            status_text = 'No MX Record Found'
            status_code = 'NO_MX'
            failure_reason = 'No MX Record Found'
            risk_level = 'high'

        result = _build_validation_result(
            normalized,
            valid_syntax=True,
            valid_mailbox=False,
            risky=True,
            risk=risk_level,
            status=status_text,
            status_code=status_code,
            classification='Invalid',
            failure_reason=failure_reason,
        )
        history_signal = _get_historical_verification_signal(normalized)
        diagnostics['historical_signal'] = history_signal
        diagnostics['domain_reputation'] = _compute_domain_reputation(
            domain,
            status_code=status_code,
            history_signal=history_signal,
        )
        diagnostics['heuristics'] = {
            'confidence_score': max(20, diagnostics['domain_reputation'].get('score', 0) - 10),
            'decision_basis': ['syntax_regex', 'dns_lookup', 'mx_lookup'],
        }
        diagnostics['final_status_code'] = status_code
        diagnostics['final_status'] = status_text
        return result, diagnostics

    last_failure_reason = ''
    smtp_retry_attempts = _get_smtp_retry_attempts()
    smtp_retry_backoff = _get_smtp_retry_backoff_seconds()
    mail_from_pool = _get_smtp_mail_from_pool()
    retryable_codes = {421, 450, 451, 452}
    greylist_markers = ('greylist', 'greylisting', 'try again later', 'temporarily deferred')

    for mx_host in mx_hosts:
        attempt = {
            'mx_host': mx_host,
            'connected': False,
            'starttls_used': False,
            'smtp_retries': [],
            'error': '',
        }

        successful = False
        for retry_index in range(smtp_retry_attempts):
            for mail_from in mail_from_pool:
                retry_event = {
                    'retry': retry_index + 1,
                    'mail_from': mail_from,
                    'connected': False,
                    'starttls_used': False,
                    'rcpt_code': None,
                    'rcpt_message': '',
                    'catch_all_probe_code': None,
                    'catch_all_probe_message': '',
                    'error': '',
                }
                attempt['smtp_retries'].append(retry_event)

                try:
                    with smtplib.SMTP(timeout=8) as server:
                        server.connect(mx_host, 25)
                        retry_event['connected'] = True
                        attempt['connected'] = True
                        server.ehlo_or_helo_if_needed()
                        if server.has_extn('starttls'):
                            server.starttls()
                            server.ehlo_or_helo_if_needed()
                            retry_event['starttls_used'] = True
                            attempt['starttls_used'] = True

                        server.mail(mail_from)
                        code, message = server.rcpt(normalized)
                        code = int(code or 0)
                        message_text = str(message or '').strip()
                        retry_event['rcpt_code'] = code
                        retry_event['rcpt_message'] = message_text

                        if code == 250:
                            probe_local_part = f'bhisha_probe_{secrets.token_hex(6)}'
                            probe_email = f'{probe_local_part}@{domain}'
                            probe_code, probe_message = server.rcpt(probe_email)
                            catch_all = int(probe_code or 0) == 250
                            retry_event['catch_all_probe_code'] = int(probe_code or 0)
                            retry_event['catch_all_probe_message'] = str(probe_message or '').strip()
                            diagnostics['attempts'].append(attempt)

                            history_signal = _get_historical_verification_signal(normalized)
                            diagnostics['historical_signal'] = history_signal
                            diagnostics['domain_reputation'] = _compute_domain_reputation(
                                domain,
                                status_code='CATCH_ALL_DOMAIN' if catch_all else 'SMTP_ACCEPTED',
                                history_signal=history_signal,
                            )
                            confidence_score = diagnostics['domain_reputation'].get('score', 50)
                            if catch_all:
                                confidence_score = max(45, confidence_score - 10)
                            diagnostics['heuristics'] = {
                                'confidence_score': confidence_score,
                                'decision_basis': ['syntax_regex', 'dns_lookup', 'mx_lookup', 'smtp_rcpt', 'catch_all_probe', 'historical_signal', 'domain_reputation'],
                            }

                            result = _build_validation_result(
                                normalized,
                                valid_syntax=True,
                                valid_mailbox=True,
                                catch_all=catch_all,
                                risky=catch_all,
                                risk='medium' if catch_all else 'low',
                                status='Catch-all domain detected' if catch_all else 'Valid Mailbox',
                                status_code='CATCH_ALL_DOMAIN' if catch_all else 'SMTP_ACCEPTED',
                                classification='Risky' if catch_all else 'Deliverable',
                                failure_reason='Catch-all domain accepted test recipient' if catch_all else '',
                            )
                            diagnostics['final_status_code'] = result.get('statusCode', '')
                            diagnostics['final_status'] = result.get('status', '')
                            return result, diagnostics

                        if code in {550, 551, 553}:
                            diagnostics['attempts'].append(attempt)
                            history_signal = _get_historical_verification_signal(normalized)
                            diagnostics['historical_signal'] = history_signal
                            diagnostics['domain_reputation'] = _compute_domain_reputation(
                                domain,
                                status_code='HARD_BOUNCE_MAILBOX_NOT_FOUND',
                                history_signal=history_signal,
                            )
                            diagnostics['heuristics'] = {
                                'confidence_score': min(95, diagnostics['domain_reputation'].get('score', 50) + 15),
                                'decision_basis': ['syntax_regex', 'dns_lookup', 'mx_lookup', 'smtp_rcpt_hard_bounce', 'historical_signal', 'domain_reputation'],
                            }

                            result = _build_validation_result(
                                normalized,
                                valid_syntax=True,
                                valid_mailbox=False,
                                risky=True,
                                risk='high',
                                status='Hard Bounce (Mailbox Not Found)',
                                status_code='HARD_BOUNCE_MAILBOX_NOT_FOUND',
                                classification='Invalid',
                                failure_reason=f'Hard bounce from {mx_host}: {message_text}',
                            )
                            diagnostics['final_status_code'] = result.get('statusCode', '')
                            diagnostics['final_status'] = result.get('status', '')
                            return result, diagnostics

                        lower_message = message_text.lower()
                        if code in retryable_codes:
                            if any(marker in lower_message for marker in greylist_markers):
                                if retry_index + 1 < smtp_retry_attempts:
                                    time.sleep(smtp_retry_backoff * (retry_index + 1))
                                    continue

                                diagnostics['attempts'].append(attempt)
                                history_signal = _get_historical_verification_signal(normalized)
                                diagnostics['historical_signal'] = history_signal
                                diagnostics['domain_reputation'] = _compute_domain_reputation(
                                    domain,
                                    status_code='GREYLISTED',
                                    history_signal=history_signal,
                                )
                                diagnostics['heuristics'] = {
                                    'confidence_score': max(40, diagnostics['domain_reputation'].get('score', 50) - 5),
                                    'decision_basis': ['syntax_regex', 'dns_lookup', 'mx_lookup', 'smtp_greylisting_signal', 'retry_logic'],
                                }

                                result = _build_validation_result(
                                    normalized,
                                    valid_syntax=True,
                                    valid_mailbox=False,
                                    risky=True,
                                    risk='medium',
                                    status='Greylisting detected (temporary deferral)',
                                    status_code='GREYLISTED',
                                    classification='Risky',
                                    failure_reason=f'Greylisting response from {mx_host}: {message_text}',
                                )
                                diagnostics['final_status_code'] = result.get('statusCode', '')
                                diagnostics['final_status'] = result.get('status', '')
                                return result, diagnostics

                            if retry_index + 1 < smtp_retry_attempts:
                                time.sleep(smtp_retry_backoff * (retry_index + 1))
                                continue

                            diagnostics['attempts'].append(attempt)
                            history_signal = _get_historical_verification_signal(normalized)
                            diagnostics['historical_signal'] = history_signal
                            diagnostics['domain_reputation'] = _compute_domain_reputation(
                                domain,
                                status_code='SMTP_TEMPORARY_FAILURE',
                                history_signal=history_signal,
                            )
                            diagnostics['heuristics'] = {
                                'confidence_score': max(35, diagnostics['domain_reputation'].get('score', 50) - 8),
                                'decision_basis': ['syntax_regex', 'dns_lookup', 'mx_lookup', 'smtp_rcpt_temporary_failure', 'retry_logic'],
                            }

                            result = _build_validation_result(
                                normalized,
                                valid_syntax=True,
                                valid_mailbox=False,
                                risky=True,
                                risk='medium',
                                status=f'Temporary SMTP failure ({code})',
                                status_code='SMTP_TEMPORARY_FAILURE',
                                classification='Risky',
                                failure_reason=f'Temporary SMTP response from {mx_host}: {message_text}',
                            )
                            diagnostics['final_status_code'] = result.get('statusCode', '')
                            diagnostics['final_status'] = result.get('status', '')
                            return result, diagnostics

                        last_failure_reason = f'SMTP response from {mx_host}: {code} {message_text}'.strip()
                except Exception as exc:
                    retry_event['error'] = str(exc)
                    last_failure_reason = f'SMTP error on {mx_host}: {exc}'
                    continue

            if successful:
                break

        diagnostics['attempts'].append(attempt)

    result = _build_validation_result(
        normalized,
        valid_syntax=True,
        valid_mailbox=False,
        risky=True,
        risk='high',
        status='SMTP validation failed across MX hosts',
        status_code='SMTP_CONNECTION_FAILED',
        classification='Risky',
        failure_reason=last_failure_reason or 'SMTP validation failed',
    )
    history_signal = _get_historical_verification_signal(normalized)
    diagnostics['historical_signal'] = history_signal
    diagnostics['domain_reputation'] = _compute_domain_reputation(
        domain,
        status_code='SMTP_CONNECTION_FAILED',
        history_signal=history_signal,
    )
    diagnostics['heuristics'] = {
        'confidence_score': max(30, diagnostics['domain_reputation'].get('score', 50) - 12),
        'decision_basis': ['syntax_regex', 'dns_lookup', 'mx_lookup', 'smtp_transport_failure', 'retry_logic'],
    }
    diagnostics['final_status_code'] = result.get('statusCode', '')
    diagnostics['final_status'] = result.get('status', '')
    return result, diagnostics
def _validate_email_with_own_system(email):
    result, _ = _validate_email_with_own_system_diagnostics(email)
    return result


def _validate_email_with_zerobounce(email):
    normalized = str(email or '').strip().lower()
    api_key = str(getattr(settings, 'ZEROBOUNCE_API_KEY', '') or '').strip()
    base_url = str(getattr(settings, 'ZEROBOUNCE_VALIDATE_URL', 'https://api.zerobounce.net/v2/validate') or '').strip()

    if not api_key:
        return _build_validation_result(
            normalized,
            valid_syntax=False,
            valid_mailbox=False,
            risky=True,
            risk='high',
            status='Validation provider is not configured.',
            status_code='PROVIDER_NOT_CONFIGURED',
            classification='Invalid',
            failure_reason='Validation provider is not configured.',
        )

    if not re.fullmatch(_EMAIL_REGEX, normalized):
        return _build_validation_result(
            normalized,
            valid_syntax=False,
            valid_mailbox=False,
            risky=True,
            risk='high',
            status='Invalid Email Format',
            status_code='INVALID_FORMAT',
            classification='Invalid',
            failure_reason='Invalid Email Format',
        )

    try:
        response = requests.get(
            base_url,
            params={'api_key': api_key, 'email': normalized, 'ip_address': ''},
            timeout=8,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        return _build_validation_result(
            normalized,
            valid_syntax=True,
            valid_mailbox=False,
            risky=True,
            risk='high',
            status=f'Validation request failed: {type(exc).__name__}',
            status_code='PROVIDER_ERROR',
            classification='Risky',
            failure_reason=f'Validation request failed: {type(exc).__name__}',
        )

    status_value = str(payload.get('status') or '').strip().lower()
    sub_status = str(payload.get('sub_status') or '').strip()
    did_you_mean = str(payload.get('did_you_mean') or '').strip()
    disposable = bool(payload.get('disposable'))
    role_based = bool(payload.get('role'))
    spam = status_value == 'do_not_mail'
    catch_all = status_value == 'catch-all'

    valid_syntax = status_value not in {'invalid'}
    valid_mailbox = status_value == 'valid'

    if status_value == 'valid':
        classification = 'Deliverable'
        risk = 'low'
        risky = False
    elif status_value in {'invalid', 'do_not_mail'}:
        classification = 'Invalid'
        risk = 'high'
        risky = True
    elif status_value in {'catch-all', 'unknown'}:
        classification = 'Risky'
        risk = 'medium'
        risky = True
    else:
        classification = 'Unknown'
        risk = 'unknown'
        risky = True

    status_text = f'Validation status: {status_value or "unknown"}'
    if sub_status:
        status_text = f'{status_text} ({sub_status})'

    return _build_validation_result(
        normalized,
        valid_syntax=valid_syntax,
        valid_mailbox=valid_mailbox,
        catch_all=catch_all,
        did_you_mean=did_you_mean,
        disposable=disposable,
        role_based=role_based,
        spam=spam,
        risky=risky,
        risk=risk,
        provider_message_id=str(payload.get('transaction_id') or ''),
        status=status_text,
        status_code=str(payload.get('error') or payload.get('status') or 'Success').upper(),
        classification=classification,
        failure_reason='' if classification == 'Deliverable' else status_text,
    )


def _validate_email_list_with_parallel_workers(unique_emails, validator_fn):
    workers = _get_email_validation_worker_count()
    result_map = {}

    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(validator_fn, candidate): candidate
            for candidate in unique_emails
        }
        for future in as_completed(future_map):
            candidate = future_map[future]
            try:
                result_map[candidate] = future.result()
            except Exception as exc:
                result_map[candidate] = _build_email_validation_error_result(candidate, str(exc))

    return [result_map[candidate] for candidate in unique_emails if candidate in result_map]


def _validate_email_list(unique_emails, provider_mode=None):
    mode = str(provider_mode or _get_email_validation_provider_mode() or 'own_system').strip().lower()

    if mode == 'zerobounce':
        return _validate_email_list_with_parallel_workers(unique_emails, _validate_email_with_zerobounce)

    return _validate_email_list_with_parallel_workers(unique_emails, _validate_email_with_own_system)


def _score_verifalia_entry(entry):
    if not isinstance(entry, dict):
        return -1

    score = 0
    lowered_keys = {str(key).strip().lower() for key in entry.keys()}
    for candidate in ['classification', 'statuscode', 'status_code', 'status', 'details', 'risk', 'quality']:
        if candidate in lowered_keys:
            score += 1

    details = entry.get('details')
    if isinstance(details, dict):
        score += len(details.keys())

    if lowered_keys == {'inputdata'}:
        score -= 10

    return score


def _extract_verifalia_entries(result_payload):
    def _collect_entries(node, bucket):
        if isinstance(node, dict):
            entries = node.get('entries')
            if isinstance(entries, list) and entries:
                bucket.extend([entry for entry in entries if isinstance(entry, dict)])
            elif isinstance(entries, dict):
                entries_data = entries.get('data')
                if isinstance(entries_data, list) and entries_data:
                    bucket.extend([entry for entry in entries_data if isinstance(entry, dict)])
            for value in node.values():
                _collect_entries(value, bucket)
        elif isinstance(node, list):
            for value in node:
                _collect_entries(value, bucket)

    entries = []
    _collect_entries(result_payload, entries)
    return entries


def _extract_verifalia_entry(result_payload):
    entries = _extract_verifalia_entries(result_payload)
    if not entries:
        return {}

    return max(entries, key=_score_verifalia_entry)


def _normalize_verifalia_status_text(value):
    if isinstance(value, dict):
        for candidate in ['description', 'message', 'text', 'name', 'status', 'value']:
            extracted = value.get(candidate)
            if extracted not in [None, '']:
                return str(extracted).strip()
        return ''
    return str(value or '').strip()


def _extract_verifalia_status_code(entry, payload):
    direct_code = _lookup_value_from_nested_dict(entry, ['statuscode', 'status_code'])
    if direct_code not in [None, '']:
        return str(direct_code).strip()

    direct_payload_code = _lookup_value_from_nested_dict(payload, ['statuscode', 'status_code'])
    if direct_payload_code not in [None, '']:
        return str(direct_payload_code).strip()

    status_node = _lookup_value_from_nested_dict(entry, ['status'])
    if isinstance(status_node, dict):
        nested_status_code = _lookup_value_from_nested_dict(status_node, ['code', 'statuscode', 'status_code'])
        if nested_status_code not in [None, '']:
            return str(nested_status_code).strip()

    payload_status_node = _lookup_value_from_nested_dict(payload, ['status'])
    if isinstance(payload_status_node, dict):
        nested_payload_status_code = _lookup_value_from_nested_dict(payload_status_node, ['code', 'statuscode', 'status_code'])
        if nested_payload_status_code not in [None, '']:
            return str(nested_payload_status_code).strip()

    return ''


def _lookup_value_from_nested_dict(payload, candidate_keys):
    if not isinstance(payload, (dict, list)):
        return None

    queue = [payload]
    lowered_candidates = [str(key or '').strip().lower() for key in candidate_keys if str(key or '').strip()]

    while queue:
        current = queue.pop(0)

        if isinstance(current, list):
            queue.extend(current)
            continue

        if not isinstance(current, dict):
            continue

        lowered_map = {str(k).strip().lower(): v for k, v in current.items()}
        for candidate in lowered_candidates:
            if candidate in lowered_map:
                value = lowered_map[candidate]
                if value not in [None, '']:
                    return value

        for value in current.values():
            if isinstance(value, dict):
                queue.append(value)
            elif isinstance(value, list):
                queue.extend(value)

    return None


def _to_bool_or_none(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in ['true', '1', 'yes', 'y']:
            return True
        if normalized in ['false', '0', 'no', 'n']:
            return False

    return None


def _extract_text_from_provider_field(value):
    if value in [None, '']:
        return ''

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, (int, float, bool)):
        return str(value).strip()

    if isinstance(value, dict):
        for candidate in ['summary', 'report', 'description', 'message', 'text', 'value', 'content']:
            if candidate in value and value[candidate] not in [None, '']:
                extracted = _extract_text_from_provider_field(value[candidate])
                if extracted:
                    return extracted

        for nested in value.values():
            extracted = _extract_text_from_provider_field(nested)
            if extracted:
                return extracted
        return ''

    if isinstance(value, list):
        for nested in value:
            extracted = _extract_text_from_provider_field(nested)
            if extracted:
                return extracted
        return ''

    return ''


def _infer_verifalia_classification(current_value, status_text='', summary_text='', report_text=''):
    normalized_current = str(current_value or '').strip()
    if normalized_current and normalized_current.lower() not in {'standard', 'unknown'}:
        return normalized_current

    combined_text = ' '.join(
        str(value or '').lower()
        for value in [status_text, summary_text, report_text, normalized_current]
        if str(value or '').strip()
    )

    if any(phrase in combined_text for phrase in [
        'high-risk email type',
        'well-known disposable email address provider',
        'disposable e-mail address provider',
        'domainiswellknowndea',
    ]):
        return 'Risky'

    if any(phrase in combined_text for phrase in [
        'invalid email address',
        'syntax error',
        'mailbox validation failed',
    ]):
        return 'Invalid'

    if any(phrase in combined_text for phrase in [
        'safe to send mail',
        'valid according to syntax rules',
        'can correctly receive messages sent to the email address being tested',
        'can correctly receive messages sent to the email address domain',
        'deliverable',
    ]):
        return 'Deliverable'

    return normalized_current or 'Unknown'


def _infer_verifalia_bool_from_text(text_blob, positive_phrases, negative_phrases):
    normalized_text = str(text_blob or '').lower()
    if any(phrase in normalized_text for phrase in negative_phrases):
        return False
    if any(phrase in normalized_text for phrase in positive_phrases):
        return True
    return None


def _normalize_email_validation_flags(email, entry, quality_info):
    normalized_email = str(email or '').strip().lower()
    domain = normalized_email.split('@', 1)[1] if '@' in normalized_email else ''
    provider_text = ' '.join(
        str(value or '').strip()
        for value in [
            quality_info.get('raw_summary'),
            quality_info.get('raw_report'),
            quality_info.get('status_text'),
            quality_info.get('classification'),
        ]
        if str(value or '').strip()
    )
    suggested_email = str(
        _lookup_value_from_nested_dict(entry, ['suggestedemailaddress', 'suggestion', 'didyoumean']) or ''
    ).strip()

    valid_syntax = _to_bool_or_none(
        _lookup_value_from_nested_dict(
            entry,
            ['isvalidsyntax', 'validsyntax', 'hassyntaxvalidity', 'issyntaxvalid', 'syntaxisvalid'],
        )
    )
    if valid_syntax is None:
        syntax_failure = _to_bool_or_none(
            _lookup_value_from_nested_dict(entry, ['issyntaxfailure', 'syntaxfailure', 'hassyntaxfailure'])
        )
        if syntax_failure is not None:
            valid_syntax = not syntax_failure

    if valid_syntax is None:
        valid_syntax = _infer_verifalia_bool_from_text(
            provider_text,
            ['valid according to syntax rules', 'syntactically valid', 'syntax validation'],
            ['invalid email address', 'syntax error', 'syntax validation failed'],
        )

    valid_mailbox = _to_bool_or_none(
        _lookup_value_from_nested_dict(
            entry,
            [
                'isvalidmailbox',
                'validmailbox',
                'ismailboxvalid',
                'hasvalidmailbox',
                'isdeliverable',
                'deliverable',
            ],
        )
    )
    if valid_mailbox is None:
        valid_mailbox = _infer_verifalia_bool_from_text(
            provider_text,
            ['can correctly receive messages sent to the email address being tested', 'mailbox validation'],
            ['mailbox validation failed', 'cannot correctly receive messages', 'does not accept messages'],
        )
    if valid_mailbox is None:
        if quality_info.get('quality') == 'deliverable':
            valid_mailbox = True
        elif quality_info.get('quality') == 'invalid':
            valid_mailbox = False

    catch_all = _to_bool_or_none(
        _lookup_value_from_nested_dict(entry, ['iscatchall', 'catchall', 'iscatchallmailbox', 'iscatchalladdress'])
    )
    if catch_all is None:
        catch_all = _infer_verifalia_bool_from_text(
            provider_text,
            ['catch-all mail exchanger validation', 'accept messages sent to nonexistent email addresses'],
            ['does not accept messages sent to nonexistent email addresses'],
        )

    disposable = _to_bool_or_none(
        _lookup_value_from_nested_dict(
            entry,
            [
                'isdisposable',
                'disposable',
                'isdisposableemailaddress',
                'isdisposableaddress',
                'isdea',
                'iswellknowndea',
            ],
        )
    )
    if disposable is None:
        disposable = _infer_verifalia_bool_from_text(
            provider_text,
            ['disposable email address', 'well-known disposable email address provider', 'dea provider', 'disposable provider'],
            ['not associated with a well-known disposable email provider', 'not a known disposable e-mail address provider'],
        )
    if disposable is None:
        status_code_hint = str(quality_info.get('status_code') or '').strip().lower()
        if 'dea' in status_code_hint or 'disposable' in provider_text.lower():
            disposable = True

    role_based = _to_bool_or_none(
        _lookup_value_from_nested_dict(entry, ['isrolebased', 'rolebased', 'isroleaccount', 'isrolerelated'])
    )
    if role_based is None:
        role_based = _infer_verifalia_bool_from_text(
            provider_text,
            ['recognized role account', 'role account validation'],
            ['not a recognized role account'],
        )

    risk_value = _lookup_value_from_nested_dict(entry, ['risk', 'risklevel', 'risklabel'])
    risk_score = _lookup_value_from_nested_dict(entry, ['riskscore', 'risk_score', 'score'])

    risk = str(risk_value or '').strip().lower() or 'unknown'
    if risk == 'unknown' and ('high-risk email type' in provider_text.lower() or disposable):
        risk = 'high'

    status_code_hint = str(quality_info.get('status_code') or '').strip().lower()
    classification_hint = str(quality_info.get('classification') or '').strip().lower()
    has_syntax_failure_text = any(
        phrase in provider_text.lower()
        for phrase in ['invalid email address', 'syntax error', 'syntax validation failed']
    )
    has_mailbox_failure_text = any(
        phrase in provider_text.lower()
        for phrase in ['mailbox validation failed', 'cannot correctly receive messages']
    )

    # If Verifalia says this is a successful deliverable result, prefer that signal
    # over sparse/ambiguous nested booleans extracted from the payload.
    if status_code_hint == 'success' and classification_hint in ['deliverable', 'safe']:
        if not has_syntax_failure_text:
            valid_syntax = True
        if not has_mailbox_failure_text:
            valid_mailbox = True

    derived_risky = risk in ['high', 'medium', 'risky']
    if classification_hint == 'risky' or 'high-risk email type' in provider_text.lower() or 'well-known disposable email address provider' in provider_text.lower():
        derived_risky = True
    elif classification_hint in ['deliverable', 'invalid'] or 'safe to send mail' in provider_text.lower() or 'valid according to syntax rules' in provider_text.lower():
        derived_risky = False

    return {
        'email': normalized_email,
        'domain': domain,
        'validSyntax': valid_syntax,
        'validMailbox': valid_mailbox,
        'catchAll': catch_all,
        'didYouMean': suggested_email,
        'disposable': disposable,
        'roleBased': role_based,
        'risky': derived_risky,
        'risk': risk,
        'riskScore': risk_score,
    }


def _normalize_optional_bool(value):
    parsed = _to_bool_or_none(value)
    if parsed is None:
        return False
    return bool(parsed)


def _to_client_validation_result(item):
    entered_email = str(item.get('email') or '').strip().lower()
    did_you_mean = str(item.get('didYouMean') or '').strip()
    risk = str(item.get('risk') or 'unknown').strip().lower() or 'unknown'
    classification = str(item.get('classification') or 'Unknown').strip() or 'Unknown'
    status_text = str(item.get('status') or 'Validation completed.').strip() or 'Validation completed.'
    status_code = str(item.get('statusCode') or item.get('status_code') or 'Success').strip() or 'Success'
    classification_lower = classification.lower()
    status_text_lower = status_text.lower()
    status_code_lower = status_code.lower()

    deliverable_signal = bool(
        classification_lower in {'deliverable', 'safe'}
        or ('safe to send mail' in status_text_lower and status_code_lower == 'success')
    )

    valid_syntax_raw = _to_bool_or_none(item.get('validSyntax'))
    valid_mailbox_raw = _to_bool_or_none(item.get('validMailbox'))
    catch_all_raw = _to_bool_or_none(item.get('catchAll'))
    disposable_raw = _to_bool_or_none(item.get('disposable'))
    role_based_raw = _to_bool_or_none(item.get('roleBased'))
    risky_raw = _to_bool_or_none(item.get('risky'))

    valid_syntax = bool(valid_syntax_raw) if valid_syntax_raw is not None else deliverable_signal
    valid_mailbox = bool(valid_mailbox_raw) if valid_mailbox_raw is not None else deliverable_signal
    catch_all = bool(catch_all_raw) if catch_all_raw is not None else False
    disposable = bool(disposable_raw) if disposable_raw is not None else False
    role_based = bool(role_based_raw) if role_based_raw is not None else False

    if risky_raw is not None:
        risky = bool(risky_raw)
    elif deliverable_signal:
        risky = False
    else:
        risky = _is_high_risk_value(risk)

    if deliverable_signal and risk in {'unknown', ''}:
        risk = 'low'

    normalized_item = {
        **item,
        'email': entered_email,
        'validMailbox': valid_mailbox,
        'validSyntax': valid_syntax,
        'catchAll': catch_all,
        'disposable': disposable,
        'roleBased': role_based,
        'risky': risky,
        'risk': risk,
        'classification': classification,
        'status': status_text,
        'statusCode': status_code,
    }

    bhisha_result = _build_bhisha_api_validation_result(normalized_item)

    return {
        'email': entered_email,
        'validMailbox': valid_mailbox,
        'validSyntax': valid_syntax,
        'catchAll': catch_all,
        'didYouMean': did_you_mean or entered_email,
        'disposable': disposable,
        'roleBased': role_based,
        'risky': risky,
        'risk': risk,
        'providerMessageId': str(item.get('providerMessageId') or '').strip(),
        'classification': classification,
        'status': status_text,
        'statusCode': status_code,
        'summary': str(item.get('summary') or '').strip(),
        'report': str(item.get('report') or '').strip(),
        'bhisha_result': bhisha_result,
    }


def _is_high_risk_value(value):
    risk = str(value or '').strip().lower()
    return risk in {'high', 'very_high', 'medium', 'risky', 'unknown'}


def _is_safe_client_validation_result(item):
    return bool(
        item.get('validMailbox')
        and item.get('validSyntax')
        and not item.get('disposable')
        and not item.get('roleBased')
        and not _is_high_risk_value(item.get('risk'))
    )


def _build_verifalia_style_report(normalized_flags, quality_info):
    email = normalized_flags['email']
    domain = normalized_flags['domain'] or ''
    local_part = email.split('@', 1)[0] if '@' in email else email
    classification = quality_info.get('classification') or 'Unknown'
    raw_summary = _extract_text_from_provider_field(quality_info.get('raw_summary'))
    raw_report = _extract_text_from_provider_field(quality_info.get('raw_report'))

    summary_lines = [
        '#### Validation summary',
        f'Input data:**{email}**',
        '',
        f'Classification:{classification}',
        '',
        '---',
        'Status:',
    ]

    if quality_info.get('quality') == 'deliverable':
        status_text = 'Valid email, with no high-risk factors detected: safe to send mail.'
        status_code = 'Success'
    elif quality_info.get('quality') == 'invalid':
        status_text = 'Invalid email address.'
        status_code = 'Success'
    elif quality_info.get('quality') == 'risky':
        status_text = 'High-risk email type: the email address is associated with a well-known disposable email address provider (DEA). We strongly recommend removing DEAs from your lists.'
        status_code = 'DomainIsWellKnownDea'
    else:
        status_text = 'Validation completed.'
        status_code = 'Success'

    summary_lines.extend([
        status_text,
        '',
        'Status code:',
        status_code,
    ])

    report_lines = [
        '#### Validation report',
        'Syntax validation',
        ' The address is valid according to syntax rules.',
        f'Address (without comments and folding white spaces){email}Local part{local_part}Domain part{domain}ASCII domain part*The domain part is not [internationalized](https://en.wikipedia.org/wiki/Internationalized_domain_name) and doesn\'t require ASCII conversion.*',
        '',
        'Role account validation',
        ' The email address is not a recognized role account; role accounts refer to email addresses that are associated with a specific function or role within an organization, rather than being tied to an individual person.',
        '',
        'Syntax validation, ISP-specific',
        f" Following the syntactic rules of the target mail exchanger(s) for the '{domain}' domain, the address is considered syntactically valid.",
        '',
        'Disposable email address (DEA) validation',
        ' The address is not associated with a well-known [disposable email](https://en.wikipedia.org/wiki/Disposable_e-mail_address) provider. A disposable email provider, often referred to as DEA provider, is a service that offers temporary and anonymous email addresses for short-term use.',
        '',
        'Free email provider check',
        f' This email address is associated with a well-known free email provider ({domain}).',
        '',
        'DNS records validation',
        f" The '{domain}' domain has valid DNS records.",
        '',
        'Honeypot detection',
        ' The email address is not a [honeypot](https://en.wikipedia.org/wiki/Honeypot_(computing)) (also known as spamtrap).',
        '',
        'Parked / inactive mail exchanger detection',
        ' The mail exchanger hosting the email address under test does not appear to be parked or inactive.',
        '',
        'Disposable email address (DEA) validation, second pass',
        ' The mail exchanger responsible for the email address is not a known [disposable e-mail address](https://en.wikipedia.org/wiki/Disposable_e-mail_address) (DEA) provider.',
        '',
        'SMTP server validation',
        f" The mail exchanger(s) of the '{domain}' domain can be successfully connected to using the SMTP protocol.",
        '',
        'Mailbox validation',
        ' The mail exchanger responsible for the email address domain can correctly receive messages sent to the email address being tested.',
        '',
        'Catch-all mail exchanger validation',
        ' The mail exchanger responsible for the email address domain does not accept messages sent to nonexistent email addresses..',
    ]

    generated_summary = '\n'.join(summary_lines)
    generated_report = '\n'.join(report_lines)

    return {
        'summary': raw_summary or generated_summary,
        'report': raw_report or generated_report,
        'status': status_text,
        'statusCode': status_code,
        'classification': classification,
    }


def _collect_emails_from_text_blob(text_blob):
    pieces = re.split(r'[\n,;\s\t]+', str(text_blob or ''))
    return [str(item).strip().lower() for item in pieces if str(item).strip()]


def _collect_unique_emails_from_input(raw_value):
    if isinstance(raw_value, list):
        candidates = [str(item or '').strip().lower() for item in raw_value]
    else:
        candidates = _collect_emails_from_text_blob(raw_value)

    unique_emails = []
    seen = set()
    for email in candidates:
        if not email or email in seen:
            continue
        try:
            validate_email(email)
        except ValidationError:
            continue
        unique_emails.append(email)
        seen.add(email)
    return unique_emails


def _extract_emails_from_uploaded_file(source_file):
    if not source_file:
        return []

    filename = str(getattr(source_file, 'name', '') or '').lower()
    extracted = []

    def _add_candidate(value):
        if value is None:
            return
        normalized = str(value).strip().lower()
        if normalized:
            extracted.append(normalized)

    if filename.endswith('.txt') or filename.endswith('.csv') or filename.endswith('.xlsv'):
        source_file.seek(0)
        text_wrapper = io.TextIOWrapper(source_file, encoding='utf-8', errors='ignore')
        try:
            for line in text_wrapper:
                extracted.extend(_collect_emails_from_text_blob(line))
        finally:
            try:
                text_wrapper.detach()
            except Exception:
                pass
            source_file.seek(0)
        return extracted

    if filename.endswith('.xls'):
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError('XLS file support requires xlrd package') from exc

        try:
            source_file.seek(0)
            workbook = xlrd.open_workbook(file_contents=source_file.read())
            for sheet in workbook.sheets():
                for row_index in range(sheet.nrows):
                    for col_index in range(sheet.ncols):
                        _add_candidate(sheet.cell_value(row_index, col_index))
            return extracted
        except Exception as exc:
            raise ValueError(f'Unable to parse XLS file: {exc}') from exc

    if filename.endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError as exc:
            raise ValueError('XLSX file support requires openpyxl package') from exc

        try:
            source_file.seek(0)
            workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
            worksheet = workbook.active
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    _add_candidate(cell)
            workbook.close()
            return extracted
        except Exception as exc:
            raise ValueError(f'Unable to parse XLSX file: {exc}') from exc

    raise ValueError('Unsupported file type. Allowed types: .txt, .csv, .xlsv, .xls, .xlsx')


def _submit_verifalia_validation_job(emails):
    username = str(getattr(settings, 'VERIFALIA_USERNAME', '') or '').strip()
    password = str(getattr(settings, 'VERIFALIA_PASSWORD', '') or '').strip()
    if not username or not password:
        raise ValueError('Verifalia credentials are not configured on the server')

    base_url = str(getattr(settings, 'VERIFALIA_API_BASE_URL', 'https://api.verifalia.com/v2.6') or '').strip().rstrip('/')
    timeout_seconds = int(getattr(settings, 'VERIFALIA_WAIT_TIMEOUT_SECONDS', 15) or 15)

    creation_payload = {
        'entries': [{'inputData': email} for email in emails],
    }

    try:
        create_response = requests.post(
            f'{base_url}/email-validations',
            json=creation_payload,
            auth=(username, password),
            timeout=12,
        )
    except requests.RequestException as exc:
        raise ValueError(f'Verifalia request failed: {type(exc).__name__}')

    if create_response.status_code not in [200, 201, 202]:
        raise ValueError(f'Verifalia create failed with status {create_response.status_code}')

    try:
        create_payload = create_response.json()
    except ValueError:
        raise ValueError('Verifalia returned invalid JSON during create')

    job_id = str(create_payload.get('id') or create_payload.get('overview', {}).get('id') or '').strip()
    status_text = str(create_payload.get('status') or create_payload.get('overview', {}).get('status') or '').strip()

    current_payload = create_payload
    start_time = time.time()
    while job_id and str(status_text).lower() in ['inprogress', 'waiting', 'queued'] and (time.time() - start_time) < timeout_seconds:
        try:
            fetch_response = requests.get(
                f'{base_url}/email-validations/{job_id}',
                auth=(username, password),
                timeout=10,
            )
            if fetch_response.status_code == 200:
                current_payload = fetch_response.json()
                status_text = str(current_payload.get('status') or current_payload.get('overview', {}).get('status') or '').strip()
        except requests.RequestException:
            break
        time.sleep(1)

    return job_id, current_payload, status_text


def _build_provider_failure_reason(provider_report):
    if provider_report['statusCode'] not in ['Success', ''] or provider_report['classification'].lower() in ['invalid', 'risky']:
        return provider_report['status'] or provider_report['classification']
    return ''


def _build_validation_result_from_verifalia(email, entry, payload, provider_message_id, provider_status_text=''):
    raw_summary = (
        _lookup_value_from_nested_dict(entry, ['summary', 'validationsummary', 'validation_summary', 'summarytext'])
        or _lookup_value_from_nested_dict(payload, ['summary', 'validationsummary', 'validation_summary', 'summarytext'])
    )
    raw_report = (
        _lookup_value_from_nested_dict(entry, ['report', 'validationreport', 'validation_report', 'reporttext', 'details'])
        or _lookup_value_from_nested_dict(payload, ['report', 'validationreport', 'validation_report', 'reporttext', 'details'])
    )
    resolved_status_text = _normalize_verifalia_status_text(
        _lookup_value_from_nested_dict(entry, ['status', 'statustext', 'status_text'])
    ) or str(provider_status_text or '').strip()
    classification = str(
        _lookup_value_from_nested_dict(entry, ['classification', 'verdict'])
        or _lookup_value_from_nested_dict(payload, ['classification', 'verdict'])
        or ''
    ).strip()
    classification = _infer_verifalia_classification(
        classification,
        resolved_status_text,
        _extract_text_from_provider_field(raw_summary),
        _extract_text_from_provider_field(raw_report),
    )
    resolved_status_code = _extract_verifalia_status_code(entry, payload)
    quality_info = _map_verifalia_quality(classification, resolved_status_text or provider_status_text)
    quality_info = {
        **quality_info,
        'classification': classification or 'Unknown',
        'status_text': resolved_status_text,
        'status_code': resolved_status_code,
        'raw_summary': raw_summary,
        'raw_report': raw_report,
    }
    normalized_flags = _normalize_email_validation_flags(email, entry, quality_info)
    provider_report = _build_verifalia_style_report(normalized_flags, quality_info)

    return {
        'email': normalized_flags['email'],
        'validMailbox': normalized_flags['validMailbox'],
        'validSyntax': normalized_flags['validSyntax'],
        'catchAll': normalized_flags['catchAll'],
        'didYouMean': normalized_flags['didYouMean'],
        'disposable': normalized_flags['disposable'],
        'roleBased': normalized_flags['roleBased'],
        'risky': normalized_flags['risky'],
        'risk': normalized_flags['risk'],
        'providerMessageId': provider_message_id,
        'summary': provider_report['summary'],
        'report': provider_report['report'],
        'status': provider_report['status'],
        'statusCode': provider_report['statusCode'],
        'classification': provider_report['classification'],
        'failure_reason': _build_provider_failure_reason(provider_report),
    }


def _build_email_validation_error_result(candidate, status_text, classification='Unknown', quality='unknown', status_code='Error', risk='unknown'):
    normalized_flags = {
        'email': candidate,
        'domain': candidate.split('@', 1)[1].lower() if '@' in candidate else '',
        'validSyntax': bool('@' in str(candidate or '')),
        'validMailbox': False,
        'catchAll': False,
        'didYouMean': '',
        'disposable': False,
        'roleBased': False,
        'risky': True,
        'risk': risk,
    }
    report = _build_verifalia_style_report(normalized_flags, {
        'classification': classification,
        'quality': quality,
        'status_text': status_text,
        'status_code': status_code,
    })
    return {
        **normalized_flags,
        'summary': report['summary'],
        'report': report['report'],
        'status': report['status'],
        'statusCode': report['statusCode'],
        'classification': report['classification'],
        'failure_reason': report['status'],
    }


def _validate_email_batch_with_verifalia(emails):
    if not emails:
        return []

    provider_message_id, current_payload, provider_status_text = _submit_verifalia_validation_job(emails)
    all_entries = _extract_verifalia_entries(current_payload)
    entry_map = {}

    for entry in all_entries:
        mapped_email = str(_lookup_value_from_nested_dict(entry, ['inputdata', 'email', 'address']) or '').strip().lower()
        if not mapped_email:
            continue

        existing = entry_map.get(mapped_email)
        candidate_score = _score_verifalia_entry(entry)
        if existing is None or candidate_score > existing[0]:
            entry_map[mapped_email] = (candidate_score, entry)

    results = []
    for index, email in enumerate(emails):
        resolved_entry = entry_map.get(email, (None, None))[1]
        if resolved_entry is None and index < len(all_entries):
            resolved_entry = all_entries[index]
        results.append(_build_validation_result_from_verifalia(email, resolved_entry or {}, current_payload, provider_message_id, provider_status_text))

    return results


def _validate_email_with_verifalia(email):
    return _validate_email_batch_with_verifalia([str(email or '').strip().lower()])[0]


def _get_verifalia_admin_credits():
    now = timezone.now()
    cached_value = _EMAIL_PROVIDER_CREDITS_CACHE.get('value')
    cache_expiry = _EMAIL_PROVIDER_CREDITS_CACHE.get('expires_at')
    if cache_expiry and cache_expiry > now and cached_value is not None:
        return cached_value

    api_key = str(getattr(settings, 'ZEROBOUNCE_API_KEY', '') or '').strip()
    if not api_key:
        return None

    credits_url = str(getattr(settings, 'ZEROBOUNCE_CREDITS_URL', 'https://api.zerobounce.net/v2/getcredits') or '').strip()

    try:
        response = requests.get(credits_url, params={'api_key': api_key}, timeout=10)
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return None

    credits_value = _extract_balance_from_data(payload)
    if credits_value is None:
        return None

    resolved_credits = round(max(0.0, float(credits_value)), 4)
    _EMAIL_PROVIDER_CREDITS_CACHE['value'] = resolved_credits
    _EMAIL_PROVIDER_CREDITS_CACHE['expires_at'] = now + timedelta(seconds=45)
    return resolved_credits


def _extract_balance_from_data(payload):
    candidate_keys = [
        'wallet_balance', 'balance', 'points', 'credits', 'available_balance',
        'available_credits', 'remaining_balance', 'remaining_credits', 'amount'
    ]

    if isinstance(payload, dict):
        lowered_map = {str(k).strip().lower(): v for k, v in payload.items()}
        for key in candidate_keys:
            if key in lowered_map:
                value = _extract_first_numeric_value(lowered_map[key])
                if value is not None:
                    return max(0.0, value)

        for value in payload.values():
            nested_value = _extract_balance_from_data(value)
            if nested_value is not None:
                return nested_value

    if isinstance(payload, list):
        for item in payload:
            nested_value = _extract_balance_from_data(item)
            if nested_value is not None:
                return nested_value

    return _extract_first_numeric_value(payload)


def _get_provider_wallet_balance():
    now = timezone.now()
    cached_value = _PROVIDER_BALANCE_CACHE.get('value')
    cache_expiry = _PROVIDER_BALANCE_CACHE.get('expires_at')
    if cache_expiry and cache_expiry > now and cached_value is not None:
        return cached_value

    provider_config = _get_admin_managed_sms_provider_config() or _get_sms_provider_config()
    if not provider_config:
        return None

    candidate_urls = _get_balance_candidate_urls(provider_config)
    if not candidate_urls:
        return None

    method = str(getattr(settings, 'SMS_PROVIDER_BALANCE_METHOD', 'GET') or 'GET').strip().upper()
    method = method if method in ['GET', 'POST'] else 'GET'

    request_params = {
        'user': provider_config.get('user', ''),
        'username': provider_config.get('user', ''),
        'pwd': provider_config.get('password', ''),
        'password': provider_config.get('password', ''),
    }

    for balance_url in candidate_urls:
        normalized_url = str(balance_url or '').strip().lower()
        try:
            if 'infobip.com' in normalized_url:
                response = requests.get(
                    balance_url,
                    auth=(provider_config.get('user', ''), provider_config.get('password', '')),
                    headers={'Accept': 'application/json'},
                    timeout=8,
                )
            elif method == 'POST':
                response = requests.post(balance_url, data=request_params, timeout=8)
            else:
                response = requests.get(balance_url, params=request_params, timeout=8)
        except requests.RequestException:
            continue

        if response.status_code != 200:
            continue

        parsed_payload = None
        try:
            parsed_payload = response.json()
        except ValueError:
            parsed_payload = response.text

        balance_value = _extract_balance_from_data(parsed_payload)
        if balance_value is None:
            continue

        resolved_balance = round(max(0.0, float(balance_value)), 2)
        _PROVIDER_BALANCE_CACHE['value'] = resolved_balance
        _PROVIDER_BALANCE_CACHE['expires_at'] = now + timedelta(seconds=45)
        return resolved_balance

    return None


def _get_users_for_notification_filter(audience_filter):
    audience_filter = (audience_filter or 'all_users').strip().lower()
    users = User.objects.all()

    if audience_filter == 'all_users':
        return users
    if audience_filter == 'verified_users':
        return users.filter(is_active=True)
    if audience_filter == 'not_verified_users':
        return users.filter(is_active=False)
    if audience_filter == 'new_joiners':
        return users.filter(date_joined__gte=timezone.now() - timedelta(days=7))
    if audience_filter == 'active_users':
        return users.filter(is_active=True)
    if audience_filter == 'inactive_users':
        return users.filter(is_active=False)

    free_trial_user_ids = set(
        SMSMessage.objects.filter(send_mode='free_trial').values_list('sender_id', flat=True)
    ) | set(
        FreeTrialVerifiedNumber.objects.filter(is_verified=True).values_list('owner_id', flat=True)
    )

    if audience_filter == 'free_trial_users':
        return users.filter(id__in=free_trial_user_ids)

    if audience_filter == 'non_free_trial_users':
        return users.exclude(id__in=free_trial_user_ids)

    return users


def _get_user_sms_usage_summary(user):
    wallet = _get_or_create_wallet(user)
    wallet_balance = Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001'))
    sms_cost = _get_sms_cost_per_request()
    total_limit = int(wallet_balance // sms_cost) if sms_cost > 0 else 0

    used_messages = (
        SMSMessage.objects.filter(sender=user)
        .exclude(status='failed')
        .exclude(Q(batch_reference__startswith='free-trial') & Q(recipient_user__isnull=False))
        .count()
    )

    available_messages = max(0, total_limit)
    used_percentage = 0
    available_percentage = 100 if total_limit > 0 else 0

    return {
        'total_limit': total_limit,
        'used_messages': used_messages,
        'available_messages': available_messages,
        'used_percentage': used_percentage,
        'available_percentage': available_percentage,
        'wallet_balance': float(wallet_balance),
    }


def _get_or_create_wallet(user):
    wallet, _ = UserWallet.objects.get_or_create(
        user=user,
        defaults={'balance': Decimal('0'), 'email_validation_balance': Decimal('0')},
    )
    return wallet


def _get_email_validation_wallet_balance(user):
    wallet = _get_or_create_wallet(user)
    return Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001'))


def _get_email_validation_cost_per_request():
    setting = PlatformSetting.objects.filter(key='email_validation_cost_per_request').first()
    if not setting:
        return Decimal('0')
    try:
        cost = Decimal(str(setting.value or '1'))
        return cost if cost > 0 else Decimal('0')
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _get_platform_setting_decimal(key, default=Decimal('0')):
    setting = PlatformSetting.objects.filter(key=key).first()
    if not setting:
        return Decimal(str(default))
    try:
        return Decimal(str(setting.value or default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(str(default))


def _get_platform_setting_text(key, default=''):
    setting = PlatformSetting.objects.filter(key=key).first()
    if not setting:
        return str(default or '').strip()
    return str(setting.value or default or '').strip()


def _get_recharge_charge_percentages():
    service_charge_percentage = _get_platform_setting_decimal('recharge_service_charge_percentage', Decimal('0'))
    tax_percentage = _get_platform_setting_decimal('recharge_tax_percentage', Decimal('0'))

    service_charge_percentage = max(Decimal('0'), service_charge_percentage)
    tax_percentage = max(Decimal('0'), tax_percentage)
    return service_charge_percentage.quantize(Decimal('0.01')), tax_percentage.quantize(Decimal('0.01'))


def _get_razorpay_config():
    key_id = str(getattr(settings, 'RAZORPAY_KEY_ID', '') or '').strip()
    key_secret = str(getattr(settings, 'RAZORPAY_KEY_SECRET', '') or '').strip()

    # Support common alternate env names to reduce deployment mistakes.
    if not key_id:
        key_id = str(getattr(settings, 'RAZORPAY_API_KEY', '') or '').strip()
    if not key_secret:
        key_secret = str(getattr(settings, 'RAZORPAY_API_SECRET', '') or '').strip()

    currency = str(getattr(settings, 'RAZORPAY_CURRENCY', 'INR') or 'INR').strip().upper() or 'INR'
    return {
        'key_id': key_id,
        'key_secret': key_secret,
        'currency': currency,
        'configured': bool(key_id and key_secret),
    }


def _is_razorpay_sdk_installed():
    try:
        import razorpay  # noqa: F401
        return True
    except Exception:
        return False


def _get_razorpay_client_or_none():
    config = _get_razorpay_config()
    if not config['configured']:
        return None

    try:
        import razorpay
    except Exception:
        return None

    try:
        return razorpay.Client(auth=(config['key_id'], config['key_secret']))
    except Exception:
        return None


def _calculate_recharge_breakdown(entered_amount, service_charge_percentage, tax_percentage):
    entered = Decimal(str(entered_amount or '0')).quantize(Decimal('0.01'))
    service_amount = ((entered * service_charge_percentage) / Decimal('100')).quantize(Decimal('0.01'))
    tax_amount = ((entered * tax_percentage) / Decimal('100')).quantize(Decimal('0.01'))
    total_amount = (entered + service_amount + tax_amount).quantize(Decimal('0.01'))

    return {
        'entered_amount': entered,
        'service_charge_percentage': service_charge_percentage,
        'tax_percentage': tax_percentage,
        'service_charge_amount': service_amount,
        'tax_amount': tax_amount,
        'total_amount': total_amount,
    }


def _yes_no(value):
    if value is None:
        return 'unknown'
    return 'yes' if bool(value) else 'no'


def _generate_unique_uuid(model_class, field_name, max_attempts=25):
    """Generate a collision-safe UUID string for unique message/request IDs."""
    attempts = 0
    generated_id = str(uuid.uuid4())

    while model_class.objects.filter(**{field_name: generated_id}).exists() and attempts < max_attempts:
        attempts += 1
        generated_id = str(uuid.uuid4())

    if model_class.objects.filter(**{field_name: generated_id}).exists():
        raise ValueError(f'Could not generate a unique UUID for {field_name}')

    return generated_id


def _resolve_sms_service_code(sms_message):
    sms_type = str(getattr(sms_message, 'sms_type', '') or '').strip().lower()
    if sms_type in {'whatsapp', 'wa'}:
        return 'WA'
    if sms_type in {'rcs'}:
        return 'RCS'
    return 'SMS'


def _assign_sms_request_id(sms_message, user_for_id=None):
    if not sms_message or sms_message.message_id:
        return sms_message.message_id

    generated_id = _generate_unique_uuid(SMSMessage, 'message_id')

    sms_message.message_id = generated_id
    sms_message.save(update_fields=['message_id', 'updated_at'])
    return generated_id


def _assign_email_validation_request_id(history):
    if not history or history.request_id:
        return history.request_id

    generated_id = _generate_unique_uuid(EmailValidationHistory, 'request_id')

    history.request_id = generated_id
    history.save(update_fields=['request_id'])
    return generated_id


def _get_sms_cost_per_request():
    setting = PlatformSetting.objects.filter(key='sms_cost_per_request').first()
    if not setting:
        return Decimal('1')

    try:
        cost = Decimal(str(setting.value or '1'))
        return cost if cost > 0 else Decimal('1')
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('1')


def _deduct_sms_credits(user, message_count):
    message_count = max(1, int(message_count or 1))
    cost_per_request = _get_sms_cost_per_request()
    total_cost = (cost_per_request * Decimal(message_count)).quantize(Decimal('0.0001'))

    wallet = _get_or_create_wallet(user)
    current_balance = Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001'))
    if current_balance < total_cost:
        raise ValueError('Insufficient messaging credits.')

    wallet.balance = (current_balance - total_cost).quantize(Decimal('0.0001'))
    wallet.email_validation_balance = wallet.balance
    wallet.save(update_fields=['balance', 'email_validation_balance', 'updated_at'])
    return total_cost, wallet.balance


def _build_simple_validation_result(item):
    syntax_ok = _normalize_optional_bool(item.get('validSyntax'))
    valid_mailbox = _normalize_optional_bool(item.get('validMailbox'))
    disposable = _normalize_optional_bool(item.get('disposable'))
    role_based = _normalize_optional_bool(item.get('roleBased'))
    catch_all = _normalize_optional_bool(item.get('catchAll'))
    domain_valid = bool(str(item.get('domain') or '').strip())
    risk_value = str(item.get('risk') or '').strip().lower()
    risk_high = bool(item.get('risky')) or risk_value in {'high', 'very_high', 'unknown'}
    has_unknown_flag = any(value is None for value in [syntax_ok, valid_mailbox, disposable, role_based])
    safe_to_send = None if has_unknown_flag else (syntax_ok and valid_mailbox and (not disposable) and (not role_based) and (not risk_high))

    return {
        'email': str(item.get('email') or '').strip().lower(),
        'valid': _yes_no(None if (syntax_ok is None or valid_mailbox is None) else (syntax_ok and valid_mailbox)),
        'syntax_error': _yes_no(None if syntax_ok is None else (not syntax_ok)),
        'safe_to_send': _yes_no(safe_to_send),
        'valid_mailbox': _yes_no(valid_mailbox),
        'catch_all': _yes_no(catch_all),
        'disposable': _yes_no(disposable),
        'role_based': _yes_no(role_based),
        'domain_valid': _yes_no(domain_valid),
        'risk_high': _yes_no(risk_high),
    }


def _deduct_email_validation_credits(user, email_count):
    email_count = max(0, int(email_count or 0))
    cost_per_request = _get_email_validation_cost_per_request()
    total_cost = (cost_per_request * Decimal(email_count)).quantize(Decimal('0.0001'))

    if total_cost <= 0:
        available_balance = _get_email_validation_wallet_balance(user)
        if available_balance is None:
            return Decimal('0.0000'), Decimal('0.0000')
        return Decimal('0.0000'), Decimal(str(available_balance)).quantize(Decimal('0.0001'))

    wallet = _get_or_create_wallet(user)
    wallet_balance = Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001'))

    if wallet_balance < total_cost:
        raise ValueError('Insufficient email validation credits.')

    wallet.balance = (wallet_balance - total_cost).quantize(Decimal('0.0001'))
    wallet.email_validation_balance = wallet.balance
    wallet.save(update_fields=['balance', 'email_validation_balance', 'updated_at'])
    return total_cost, wallet.balance


def _validate_email_list_with_verifalia(unique_emails):
    return _validate_email_list(unique_emails)


def _authenticate_api_key_request(request):
    api_key_value = str(request.headers.get('X-API-Key') or request.data.get('api_key') or '').strip()
    user_id = request.data.get('user_id')
    login_value = str(
        request.data.get('login')
        or request.data.get('username')
        or request.data.get('email')
        or ''
    ).strip()
    password = str(request.data.get('password') or '').strip()

    if not api_key_value or not password:
        raise ValueError('api_key, password, and login/email are required')

    user = None

    if login_value:
        if '@' in login_value:
            user = User.objects.filter(email__iexact=login_value).order_by('-is_active', '-id').first()
        else:
            user = (
                User.objects
                .filter(Q(username__iexact=login_value) | Q(email__iexact=login_value))
                .order_by('-is_active', '-id')
                .first()
            )

    if user is None and user_id not in [None, '']:
        try:
            user = User.objects.get(id=int(user_id))
        except (User.DoesNotExist, ValueError, TypeError):
            user = None

    if user is None:
        raise ValueError('Invalid login/email')

    if not user.check_password(password):
        raise ValueError('Invalid login/email or password')

    api_key = UserAPIKey.objects.filter(user=user, key=api_key_value, is_active=True).first()
    if not api_key:
        raise ValueError('Invalid or inactive API key')

    api_key.last_used_at = timezone.now()
    api_key.save(update_fields=['last_used_at'])
    return user, api_key


def _collect_validation_emails(request, enforce_request_limit=True):
    raw_email = str(request.data.get('email') or '').strip().lower()
    raw_emails = request.data.get('emails')
    source_file = request.FILES.get('source_file') if hasattr(request, 'FILES') else None

    collected_emails = []
    if raw_email:
        collected_emails.append(raw_email)

    if isinstance(raw_emails, list):
        collected_emails.extend([str(item).strip().lower() for item in raw_emails if str(item).strip()])
    elif isinstance(raw_emails, str):
        pieces = re.split(r'[\n,;\s]+', raw_emails)
        collected_emails.extend([str(item).strip().lower() for item in pieces if str(item).strip()])

    file_name = ''
    if source_file:
        file_name = str(getattr(source_file, 'name', '') or '')
        max_file_size_bytes = _get_email_validation_max_file_size_bytes()
        if int(getattr(source_file, 'size', 0) or 0) > max_file_size_bytes:
            raise ValueError(f'Uploaded file is too large. Max {max_file_size_bytes // (1024 * 1024)}MB allowed.')

        try:
            file_emails = _extract_emails_from_uploaded_file(source_file)
            collected_emails.extend(file_emails)
        except ValueError as exc:
            raise ValueError(str(exc))

    unique_emails = []
    seen = set()
    for item in collected_emails:
        if item and item not in seen:
            seen.add(item)
            unique_emails.append(item)

    if not unique_emails:
        raise ValueError('Provide at least one email to validate')

    if enforce_request_limit:
        max_request_count = _get_email_validation_max_request_count()
        if len(unique_emails) > max_request_count:
            raise ValueError(f'Maximum {max_request_count} emails are allowed per request')

    return unique_emails, file_name


def _collect_validation_emails_from_file(request, enforce_request_limit=True):
    unique_emails, file_name = _collect_validation_emails(request, enforce_request_limit=enforce_request_limit)
    return unique_emails, file_name


def _is_free_email_domain(email, disposable=False):
    domain = str(email or '').strip().lower().split('@', 1)[1] if '@' in str(email or '') else ''
    if not domain:
        return False
    return disposable or domain in _COMMON_FREE_EMAIL_DOMAINS


def _build_bhisha_risk_factors(item):
    factors = []
    if item.get('spam'):
        factors.append('Spam / Do Not Mail')
    if item.get('roleBased'):
        factors.append('Role Based Address')
    if item.get('catchAll'):
        factors.append('Catch All Mailbox')

    risk_value = str(item.get('risk') or '').strip().lower()
    if risk_value in {'high', 'very_high', 'medium'} and not item.get('disposable'):
        factors.append(f'Provider Risk: {risk_value}')

    return ', '.join(factors) if factors else 'None Detected'


def _build_bhisha_raw_status_details(item):
    if item.get('disposable'):
        return 'do_not_mail (disposable)'
    if item.get('roleBased'):
        return 'do_not_mail (role_based)'
    if item.get('catchAll'):
        return 'risky (catch_all)'
    if item.get('validSyntax') is False:
        return 'invalid (syntax)'

    status_code = str(item.get('statusCode') or '').strip().upper()
    status_map = {
        'SMTP_ACCEPTED': 'safe_to_mail',
        'HARD_BOUNCE_MAILBOX_NOT_FOUND': 'do_not_mail (hard_bounce)',
        'DOMAIN_NOT_FOUND': 'do_not_mail (domain_not_found)',
        'NO_MX': 'do_not_mail (no_mx)',
        'DNS_LOOKUP_FAILED': 'risky (dns_lookup_failed)',
        'DNS_UNAVAILABLE': 'risky (dns_unavailable)',
        'GREYLISTED': 'risky (greylisted)',
        'SMTP_TEMPORARY_FAILURE': 'risky (temporary_failure)',
        'SMTP_CONNECTION_FAILED': 'risky (smtp_unreachable)',
        'CATCH_ALL_DOMAIN': 'risky (catch_all)',
        'INVALID_FORMAT': 'invalid (syntax)',
        'INVALID_SYNTAX_DOMAIN_TYPO': 'invalid (domain_typo)',
    }
    if status_code in status_map:
        return status_map[status_code]

    if item.get('validMailbox') is False:
        return 'do_not_mail'

    if status_code and status_code not in {'SUCCESS'}:
        return status_code.lower()

    classification = str(item.get('classification') or '').strip().lower()
    if classification and classification not in {'deliverable', 'unknown'}:
        return classification

    return 'safe_to_mail'


def _build_bhisha_api_validation_result(item):
    valid_syntax = bool(item.get('validSyntax'))
    valid_inbox = bool(
        item.get('validMailbox')
        and valid_syntax
        and not item.get('disposable')
        and not item.get('roleBased')
    )
    disposable = bool(item.get('disposable'))
    is_free_domain = _is_free_email_domain(item.get('email'), disposable=disposable)
    status_code = str(item.get('statusCode') or '').strip().upper()
    domain_related_mail = bool(item.get('validSyntax')) and status_code not in {
        'INVALID_FORMAT',
        'INVALID_SYNTAX_DOMAIN_TYPO',
        'DOMAIN_NOT_FOUND',
        'NO_MX',
        'DNS_LOOKUP_FAILED',
        'DNS_UNAVAILABLE',
    }

    result = {
        'email': str(item.get('email') or '').strip().lower(),
        'valid_inbox': valid_inbox,
        'valid_syntax': valid_syntax,
        'domain_related_mail': domain_related_mail,
        'disposable': disposable,
        'role_based': bool(item.get('roleBased')),
        'spam': bool(item.get('spam')),
        'catch_all': bool(item.get('catchAll')),
        'risk_factors': _build_bhisha_risk_factors(item),
        'raw_status_details': _build_bhisha_raw_status_details(item),
        'is_free_domain': is_free_domain,
    }

    result['result_profile'] = _build_bhisha_result_profile(result)
    return result


def _build_bhisha_result_profile(result):
    if not isinstance(result, dict):
        return ''

    email = str(result.get('email') or '').strip().lower()
    valid_inbox = bool(result.get('valid_inbox'))
    valid_syntax = bool(result.get('valid_syntax'))
    domain_related_mail = bool(result.get('domain_related_mail'))
    disposable = bool(result.get('disposable'))
    role_based = bool(result.get('role_based'))
    catch_all = bool(result.get('catch_all'))
    risk_factors = str(result.get('risk_factors') or 'None Detected').strip() or 'None Detected'
    raw_status_details = str(result.get('raw_status_details') or '').strip() or 'safe_to_mail'
    is_free_domain = bool(result.get('is_free_domain'))

    return '\n'.join([
        f'Results Profile for: {email}',
        '----------------------------------------',
        f'Valid Inbox:    {str(valid_inbox)}',
        f'Valid Syntax:   {str(valid_syntax)}',
        f'Domain Related Mail: {str(domain_related_mail)}',
        f'Disposable:     {str(disposable)}',
        f'Role Based:     {str(role_based)}',
        f'Catch All:      {str(catch_all)}',
        f'Risk Factors:   {risk_factors}',
        '----------------------------------------',
        f'Raw Status Details:  {raw_status_details}',
        f'Is Free Domain?:     {str(is_free_domain)}',
    ])


def _extract_api_result_profiles(result_items):
    profiles = []

    for item in result_items or []:
        if not isinstance(item, dict):
            continue

        bhisha_result = item.get('bhisha_result') if isinstance(item.get('bhisha_result'), dict) else item
        profile = str(bhisha_result.get('result_profile') or '').strip()
        if not profile and isinstance(bhisha_result, dict):
            profile = _build_bhisha_result_profile(bhisha_result)

        if profile:
            profiles.append(profile)

    return profiles


def _build_concise_api_validation_response(result_items):
    normalized_results = []
    for item in result_items or []:
        if not isinstance(item, dict):
            continue
        bhisha_result = item.get('bhisha_result') if isinstance(item.get('bhisha_result'), dict) else item
        if isinstance(bhisha_result, dict):
            normalized_results.append({
                'email': str(bhisha_result.get('email') or '').strip().lower(),
                'valid_inbox': bool(bhisha_result.get('valid_inbox')),
                'valid_syntax': bool(bhisha_result.get('valid_syntax')),
                'domain_related_mail': bool(bhisha_result.get('domain_related_mail')),
                'disposable': bool(bhisha_result.get('disposable')),
                'role_based': bool(bhisha_result.get('role_based')),
                'catch_all': bool(bhisha_result.get('catch_all')),
                'risk_factors': str(bhisha_result.get('risk_factors') or 'None Detected'),
                'raw_status_details': str(bhisha_result.get('raw_status_details') or ''),
                'is_free_domain': bool(bhisha_result.get('is_free_domain')),
            })

    return {
        'count': len(normalized_results),
        'results': normalized_results,
    }


def _build_concise_api_status_response(history):
    processing_state = _get_history_processing_state(history)

    if processing_state in {'completed'}:
        summary = _get_history_summary(history)
        stored_results = summary.get('results') if isinstance(summary.get('results'), list) else []
        return _build_concise_api_validation_response(stored_results)

    if processing_state in {'failed', 'cancelled', 'stopped'}:
        summary = _get_history_summary(history)
        failure_reason = str(summary.get('failure_reason') or summary.get('error') or processing_state).strip()
        return {
            'request_id': history.request_id,
            'status': processing_state,
            'detail': failure_reason,
        }

    return {
        'request_id': history.request_id,
        'status': processing_state,
    }


def _get_history_summary(history):
    summary = getattr(history, 'results_summary', {}) or {}
    if not isinstance(summary, dict):
        summary = {}
    return summary


def _get_history_control_state(history):
    summary = _get_history_summary(history)
    return str(summary.get('control_state') or 'running').strip().lower() or 'running'


def _get_history_processing_state(history):
    status_value = str(getattr(history, 'status', '') or '').strip().lower()
    if status_value == EmailValidationHistory.STATUS_COMPLETED:
        return 'completed'
    if status_value == EmailValidationHistory.STATUS_FAILED:
        summary = _get_history_summary(history)
        failure_reason = str(summary.get('failure_reason') or summary.get('error') or '').strip().lower()
        if failure_reason in {'cancelled', 'stopped'}:
            return failure_reason
        return 'failed'

    summary = _get_history_summary(history)
    progress_state = str(summary.get('processing_state') or '').strip().lower()
    if progress_state in {'running', 'paused', 'cancelled', 'stopped', 'completed', 'failed'}:
        return progress_state
    return 'running'


def _is_worker_active(history_id):
    thread_active = False
    with _EMAIL_VALIDATION_WORKERS_LOCK:
        existing = _EMAIL_VALIDATION_WORKERS.get(int(history_id))
        if existing and existing.is_alive():
            thread_active = True
        elif existing:
            _EMAIL_VALIDATION_WORKERS.pop(int(history_id), None)

    if _is_celery_enabled():
        history = EmailValidationHistory.objects.filter(id=int(history_id)).first()
        if not history:
            return thread_active
        return _is_celery_task_active(history) or thread_active

    return thread_active


def _set_history_control_state(history, control_state):
    summary = _get_history_summary(history)
    summary['control_state'] = str(control_state or 'running').strip().lower() or 'running'
    if summary['control_state'] == 'paused':
        summary['processing_state'] = 'paused'
    elif history.status == EmailValidationHistory.STATUS_PENDING:
        summary['processing_state'] = 'running'
    history.results_summary = summary
    history.save(update_fields=['results_summary'])


def _set_history_failed_with_reason(history, reason):
    summary = _get_history_summary(history)
    safe_count = int(summary.get('safe_count') or 0)
    unsafe_count = int(summary.get('unsafe_count') or 0)
    results = summary.get('results') if isinstance(summary.get('results'), list) else []
    summary.update(
        {
            'error': str(reason or 'Validation failed.'),
            'failure_reason': str(reason or 'Validation failed.'),
            'safe_count': safe_count,
            'unsafe_count': unsafe_count,
            'results': results,
            'control_state': 'cancelled' if str(reason).lower() in {'cancelled', 'stopped'} else 'running',
            'processing_state': str(reason).lower() if str(reason).lower() in {'cancelled', 'stopped'} else 'failed',
            'progress_percent': int(summary.get('progress_percent') or 0),
            'processed_count': int(summary.get('processed_count') or 0),
            'total_count': int(summary.get('total_count') or history.email_count or 0),
            'eta_seconds': 0,
        }
    )
    history.status = EmailValidationHistory.STATUS_FAILED
    history.results_summary = summary
    history.completed_at = timezone.now()
    history.save(update_fields=['status', 'results_summary', 'completed_at'])


def _update_history_progress(history, processed_count, total_count, started_at, results, provider_message_ids):
    total = max(1, int(total_count or 1))
    processed = max(0, int(processed_count or 0))
    elapsed_seconds = max(0, int((timezone.now() - started_at).total_seconds())) if started_at else 0
    progress_percent = min(100, int((processed * 100) / total))
    eta_seconds = 0
    if processed > 0 and processed < total:
        rate = elapsed_seconds / float(processed)
        eta_seconds = max(0, int((total - processed) * rate))

    safe_count = sum(1 for item in results if _is_safe_client_validation_result(item))
    unsafe_count = len(results) - safe_count
    summary = _get_history_summary(history)
    summary.update(
        {
            'safe_count': safe_count,
            'unsafe_count': unsafe_count,
            'provider_message_id': provider_message_ids[0] if provider_message_ids else '',
            'provider_message_ids': provider_message_ids,
            'results': results,
            'processed_count': processed,
            'total_count': total,
            'progress_percent': progress_percent,
            'elapsed_seconds': elapsed_seconds,
            'eta_seconds': eta_seconds,
            'processing_state': 'running',
        }
    )
    if not summary.get('started_at'):
        summary['started_at'] = started_at.isoformat() if started_at else timezone.now().isoformat()
    history.results_summary = summary
    history.save(update_fields=['results_summary'])


def _start_email_validation_worker(history_id):
    history_key = int(history_id)

    if _is_celery_enabled():
        try:
            history = EmailValidationHistory.objects.get(id=history_key)
            summary = _get_history_summary(history)
            if _is_celery_task_active(history):
                return summary.get('celery_task_id')

            from accounts.tasks import process_email_validation_history_job

            async_result = process_email_validation_history_job.delay(history_key)
            summary['celery_task_id'] = str(async_result.id)
            summary['processing_state'] = 'running'
            summary['control_state'] = 'running'
            history.results_summary = summary
            history.save(update_fields=['results_summary'])
            return str(async_result.id)
        except Exception:
            # Fall back to thread mode if Celery is unavailable at runtime.
            pass

    with _EMAIL_VALIDATION_WORKERS_LOCK:
        existing = _EMAIL_VALIDATION_WORKERS.get(history_key)
        if existing and existing.is_alive():
            return existing

        worker = threading.Thread(target=_process_email_validation_history_job, args=(history_key,), daemon=True)
        _EMAIL_VALIDATION_WORKERS[history_key] = worker
        worker.start()
        return worker


def _process_email_validation_history_job(history_id):
    close_old_connections()
    try:
        history = EmailValidationHistory.objects.select_related('user', 'api_key').get(id=history_id)
        requested_emails = [str(item or '').strip().lower() for item in (history.emails_requested or []) if str(item or '').strip()]
        total_count = len(requested_emails)
        started_at = timezone.now()
        summary = _get_history_summary(history)
        summary.setdefault('started_at', started_at.isoformat())
        summary.setdefault('processed_count', 0)
        summary.setdefault('total_count', total_count)
        summary.setdefault('progress_percent', 0)
        summary['processing_state'] = 'running'
        summary['control_state'] = str(summary.get('control_state') or 'running').strip().lower() or 'running'
        history.results_summary = summary
        history.save(update_fields=['results_summary'])

        batch_size = _get_email_validation_batch_size()
        provider_mode = _get_email_validation_provider_mode()
        collected_results = []
        provider_request_ids = []
        processed_count = 0

        for start in range(0, total_count, batch_size):
            history.refresh_from_db(fields=['status', 'results_summary'])
            control_state = _get_history_control_state(history)

            if control_state in {'cancelled', 'stopped'}:
                _set_history_failed_with_reason(history, 'cancelled' if control_state == 'cancelled' else 'stopped')
                return

            while control_state == 'paused':
                summary = _get_history_summary(history)
                summary['processing_state'] = 'paused'
                history.results_summary = summary
                history.save(update_fields=['results_summary'])
                time.sleep(1)
                history.refresh_from_db(fields=['status', 'results_summary'])
                control_state = _get_history_control_state(history)
                if control_state in {'cancelled', 'stopped'}:
                    _set_history_failed_with_reason(history, 'cancelled' if control_state == 'cancelled' else 'stopped')
                    return

            current_batch = requested_emails[start:start + batch_size]
            batch_provider_candidates = []
            batch_results = []
            for candidate in current_batch:
                if '@' not in candidate:
                    batch_results.append(
                        _build_email_validation_error_result(
                            candidate,
                            'Invalid email address.',
                            classification='Invalid',
                            quality='invalid',
                            status_code='Success',
                            risk='none',
                        )
                    )
                else:
                    batch_provider_candidates.append(candidate)

            if batch_provider_candidates:
                batch_results.extend(_validate_email_list(batch_provider_candidates, provider_mode=provider_mode))

            result_map = {str(item.get('email') or '').strip().lower(): item for item in batch_results}
            for candidate in current_batch:
                normalized = str(candidate or '').strip().lower()
                item = result_map.get(normalized) or _build_email_validation_error_result(normalized, 'Validation result unavailable.')
                client_item = _to_client_validation_result(item)
                collected_results.append(client_item)
                provider_id = str(client_item.get('providerMessageId') or '').strip()
                if provider_id:
                    provider_request_ids.append(provider_id)

            processed_count = min(total_count, processed_count + len(current_batch))
            _update_history_progress(
                history,
                processed_count=processed_count,
                total_count=total_count,
                started_at=started_at,
                results=collected_results,
                provider_message_ids=provider_request_ids,
            )

        safe_count = sum(1 for item in collected_results if _is_safe_client_validation_result(item))
        unsafe_count = len(collected_results) - safe_count
        summary = _get_history_summary(history)
        summary.update(
            {
                'safe_count': safe_count,
                'unsafe_count': unsafe_count,
                'provider_message_id': provider_request_ids[0] if provider_request_ids else '',
                'provider_message_ids': provider_request_ids,
                'results': collected_results,
                'processed_count': total_count,
                'total_count': total_count,
                'progress_percent': 100,
                'elapsed_seconds': max(0, int((timezone.now() - started_at).total_seconds())),
                'eta_seconds': 0,
                'processing_state': 'completed',
                'control_state': 'running',
            }
        )
        history.status = EmailValidationHistory.STATUS_COMPLETED
        history.results_summary = summary
        history.completed_at = timezone.now()
        history.save(update_fields=['status', 'results_summary', 'completed_at'])
    except Exception as exc:
        try:
            history = EmailValidationHistory.objects.get(id=history_id)
            _set_history_failed_with_reason(history, str(exc))
        except Exception:
            pass
    finally:
        with _EMAIL_VALIDATION_WORKERS_LOCK:
            _EMAIL_VALIDATION_WORKERS.pop(int(history_id), None)
        close_old_connections()


def _employee_admin_otp_is_valid(employee, otp):
    candidate = str(otp or '').strip()
    if not candidate:
        return False
    if str(employee.admin_otp or '').strip() != candidate:
        return False
    created_at = getattr(employee, 'admin_otp_created', None)
    if not created_at:
        return False
    expiry_minutes = int(getattr(settings, 'OTP_EXPIRY_MINUTES', 10))
    return timezone.now() - created_at <= timedelta(minutes=expiry_minutes)

class SignupView(generics.CreateAPIView):
    serializer_class = SignupSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request, *args, **kwargs):
        first_name = (request.data.get('first_name') or '').strip()
        email = (request.data.get('email') or '').strip().lower()
        raw_phone_number = (request.data.get('phone_number') or '').strip()
        password = request.data.get('password') or ''

        if not first_name or not email or not raw_phone_number or not password:
            return Response({'detail': 'name, email, phone_number and password are required'}, status=400)

        phone_number = _normalize_phone_number(raw_phone_number)
        if not phone_number:
            return Response({'detail': 'Phone number must contain at least 10 digits'}, status=400)

        try:
            validate_password(password)
        except ValidationError as error:
            return Response({'password': list(error.messages)}, status=400)

        existing_user = _find_user_by_email(email)
        if existing_user and existing_user.is_active:
            return Response({'detail': 'User already exists with this email address.'}, status=400)

        if existing_user and not existing_user.is_active:
            existing_user.first_name = first_name
            existing_user.phone_number = phone_number
            existing_user.username = email
            existing_user.email = email
            existing_user.set_password(password)

            otp = generate_otp()
            existing_user.otp_code = otp
            existing_user.otp_created = timezone.now()
            existing_user.save()

            email_sent = send_otp_via_email(existing_user, otp)
            diagnostics = _otp_email_diagnostics(email_sent)
            return Response(
                {
                    'requires_otp': True,
                    'email_sent': email_sent,
                    'detail': 'Account exists but is not verified. A new OTP has been generated and sent.',
                    **diagnostics,
                },
                status=200,
                headers={'X-OTP-Email-Sent': 'true' if email_sent else 'false'},
            )

        serializer = self.get_serializer(data={
            'first_name': first_name,
            'username': email,
            'email': email,
            'phone_number': phone_number,
            'password': password,
        })
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        otp = generate_otp()
        user.otp_code = otp
        user.otp_created = timezone.now()

        admin_auto_login = _is_primary_admin_email(user.email)
        if admin_auto_login:
            user.is_staff = True
            user.is_superuser = True
            user.is_sms_enabled = True

        user.save()

        email_sent = send_otp_via_email(user, otp)
        diagnostics = _otp_email_diagnostics(email_sent)

        return Response({
            'requires_otp': True,
            'email_sent': email_sent,
            'detail': 'OTP generated. If email is not received, check SMTP credentials/server logs.',
            **diagnostics,
        }, status=201, headers={'X-OTP-Email-Sent': 'true' if email_sent else 'false'})

class OTPVerifyView(generics.GenericAPIView):
    serializer_class = OTPVerifySerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email'].strip().lower()
        otp = serializer.validated_data['otp']
        user = _find_user_by_email(email)
        if not user:
            return Response({'detail': 'Invalid email'}, status=400)
        if not otp_is_valid(user, otp):
            return Response({'detail': 'Invalid or expired OTP'}, status=400)
        user.is_active = True
        user.otp_code = ''
        user.save()
        _promote_primary_admin(user)
        # return JWT tokens
        refresh = RefreshToken.for_user(user)
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token)
        })

class LoginView(generics.GenericAPIView):
    serializer_class = LoginSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email'].strip().lower()
        password = serializer.validated_data['password']
        user = _find_user_by_email(email)
        if not user:
            return Response({'detail': 'Invalid credentials'}, status=401)

        _promote_primary_admin(user)

        if not user.check_password(password):
            return Response({'detail': 'Invalid credentials'}, status=401)

        if not user.is_active:
            return Response(
                {
                    'detail': 'Account not verified. Please verify OTP.',
                    'requires_otp_verification': True,
                    'email': user.email,
                },
                status=403,
            )

        refresh = RefreshToken.for_user(user)
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'is_admin': bool(user.is_staff or user.is_superuser or _has_primary_admin_access(user)),
            'is_primary_admin': _has_primary_admin_access(user),
        })

class ForgotPasswordView(generics.GenericAPIView):
    serializer_class = ForgotPasswordSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = _find_user_by_email(serializer.validated_data['email'])
        if not user:
            return Response(status=status.HTTP_204_NO_CONTENT)
        _promote_primary_admin(user)

        otp = generate_otp()
        user.otp_code = otp
        user.otp_created = timezone.now()
        user.save()
        email_sent = send_otp_via_email(user, otp)
        diagnostics = _otp_email_diagnostics(email_sent)
        headers = {'X-OTP-Email-Sent': 'true' if email_sent else 'false'}
        if not email_sent:
            return Response(
                {
                    'detail': 'OTP generated. Email sending failed; verify server EMAIL_* configuration and mail provider logs.',
                    'email_sent': False,
                    **diagnostics,
                },
                status=200,
                headers=headers,
            )
        return Response(
            {
                'detail': 'OTP sent successfully.',
                'email_sent': True,
                **diagnostics,
            },
            status=200,
            headers=headers,
        )


class ResendOTPView(generics.GenericAPIView):
    serializer_class = ForgotPasswordSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = _find_user_by_email(serializer.validated_data['email'])
        if not user:
            return Response(status=status.HTTP_204_NO_CONTENT)

        otp = generate_otp()
        user.otp_code = otp
        user.otp_created = timezone.now()
        user.save(update_fields=['otp_code', 'otp_created'])
        email_sent = send_otp_via_email(user, otp)
        diagnostics = _otp_email_diagnostics(email_sent)
        headers = {'X-OTP-Email-Sent': 'true' if email_sent else 'false'}

        return Response(
            {
                'detail': 'OTP sent successfully.' if email_sent else 'OTP generated but email sending failed.',
                'email_sent': email_sent,
                **diagnostics,
            },
            status=200,
            headers=headers,
        )

class ResetPasswordView(generics.GenericAPIView):
    serializer_class = ResetPasswordSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = _find_user_by_email(serializer.validated_data['email'])
        if not user:
            return Response({'detail': 'Invalid email'}, status=400)
        if not otp_is_valid(user, serializer.validated_data['otp']):
            return Response({'detail': 'Invalid or expired OTP'}, status=400)
        user.set_password(serializer.validated_data['new_password'])
        user.otp_code = ''
        user.is_active = True  # Activate user after successful password reset
        user.save()
        return Response(status=status.HTTP_200_OK)


class UserProfileView(generics.GenericAPIView):
    """Get authenticated user's profile"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        usage_summary = _get_user_sms_usage_summary(user)
        verified_numbers_count = 1 if _normalize_phone_number(user.phone_number) and not _has_primary_admin_access(user) else 0
        provider_config = _get_admin_managed_sms_provider_config() or {}
        resolved_free_trial_sender_id = ''
        if provider_config:
            try:
                resolved_free_trial_sender_id = _resolve_free_trial_sender_id(user, provider_config)
            except ValueError:
                resolved_free_trial_sender_id = ''
        data = {
            'id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'username': user.username,
            'email': user.email,
            'phone_number': user.phone_number,
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'is_primary_admin': _has_primary_admin_access(user),
            'is_sms_enabled': user.is_sms_enabled,
            'sender_id_type': user.sender_id_type,
            'sender_id': user.sender_id,
            'free_trial_sender_id': user.free_trial_sender_id,
            'date_joined': user.date_joined,
            'last_login': user.last_login,
            'sms_total_limit': usage_summary['total_limit'],
            'sms_used_messages': usage_summary['used_messages'],
            'sms_available_messages': usage_summary['available_messages'],
            'sms_used_percentage': usage_summary['used_percentage'],
            'sms_available_percentage': usage_summary['available_percentage'],
            'wallet_balance': usage_summary['wallet_balance'],
            'free_trial_limit': FREE_TRIAL_MESSAGE_LIMIT,
            'free_trial_verified_numbers_count': verified_numbers_count,
            'free_trial_service_sender_id': resolved_free_trial_sender_id,
            'role': 'primary_admin' if _has_primary_admin_access(user) else ('admin' if _has_admin_access(user) else ('employee' if _is_active_employee(user) else 'user')),
            'is_employee': _is_active_employee(user),
            'can_view_support_data': _has_support_read_access(user),
            'can_manage_support_data': _has_admin_access(user),
        }
        if _has_admin_access(user):
            provider_balance = _get_provider_wallet_balance()
            if provider_balance is not None:
                data['provider_message_balance'] = provider_balance
        return Response(data)

    def patch(self, request):
        user = request.user
        first_name = request.data.get('first_name', user.first_name)
        last_name = request.data.get('last_name', user.last_name)
        phone_number = request.data.get('phone_number', user.phone_number)
        incoming_type = request.data.get('sender_id_type', user.sender_id_type)
        incoming_sender_id = request.data.get('sender_id', user.sender_id or '')

        first_name = (first_name or '').strip()
        last_name = (last_name or '').strip()
        if not first_name:
            return Response({'detail': 'First name is required'}, status=status.HTTP_400_BAD_REQUEST)

        raw_phone = str(phone_number or '').strip()
        normalized_phone = _normalize_phone_number(raw_phone)
        if raw_phone and not normalized_phone:
            return Response({'detail': 'Phone number must contain at least 10 digits'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            normalized_type, normalized_sender_id = _normalize_sender_id(incoming_type, incoming_sender_id)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if normalized_sender_id and _sender_id_exists(normalized_sender_id, exclude_user_id=user.id):
            suggestions = _build_sender_id_suggestions(
                normalized_sender_id,
                normalized_type,
                exclude_user_id=user.id,
            )
            return Response(
                {
                    'detail': 'This sender ID is already used by another user.',
                    'suggestions': suggestions,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.first_name = first_name
        user.last_name = last_name
        user.phone_number = normalized_phone if raw_phone else None
        user.sender_id_type = normalized_type
        user.sender_id = normalized_sender_id or None

        try:
            user.save(update_fields=['first_name', 'last_name', 'phone_number', 'sender_id_type', 'sender_id'])
        except IntegrityError:
            suggestions = _build_sender_id_suggestions(
                normalized_sender_id,
                normalized_type,
                exclude_user_id=user.id,
            )
            return Response(
                {
                    'detail': 'This sender ID is already used by another user.',
                    'suggestions': suggestions,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return self.get(request)


class AdminUsersListView(generics.ListAPIView):
    """List all users - admin only"""
    permission_classes = [permissions.IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard
        
        users = User.objects.all().order_by('-date_joined')
        payload = []
        for user in users:
            wallet = UserWallet.objects.filter(user=user).first()
            payload.append(
                {
                    'id': user.id,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'username': user.username,
                    'email': user.email,
                    'phone_number': user.phone_number,
                    'is_active': user.is_active,
                    'is_staff': user.is_staff,
                    'is_superuser': user.is_superuser,
                    'is_sms_enabled': user.is_sms_enabled,
                    'sender_id_type': user.sender_id_type,
                    'sender_id': user.sender_id,
                    'free_trial_sender_id': user.free_trial_sender_id,
                    'date_joined': user.date_joined,
                    'last_login': user.last_login,
                    'wallet_balance': str(getattr(wallet, 'balance', Decimal('0'))),
                    'email_validation_balance': str(getattr(wallet, 'email_validation_balance', Decimal('0'))),
                    'api_key_count': UserAPIKey.objects.filter(user=user).count(),
                    'sms_message_count': SMSMessage.objects.filter(Q(sender=user) | Q(recipient_user=user)).count(),
                    'email_validation_count': EmailValidationHistory.objects.filter(user=user).count(),
                }
            )
        return Response(payload)


class AdminUserPermissionView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, user_id):
        guard = None if _has_admin_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=404)

        if 'first_name' in request.data:
            user.first_name = str(request.data.get('first_name') or '').strip()

        if 'last_name' in request.data:
            user.last_name = str(request.data.get('last_name') or '').strip()

        if 'phone_number' in request.data:
            raw_phone = str(request.data.get('phone_number') or '').strip()
            normalized_phone = _normalize_phone_number(raw_phone)
            if raw_phone and not normalized_phone:
                return Response({'detail': 'Phone number must contain at least 10 digits'}, status=status.HTTP_400_BAD_REQUEST)
            user.phone_number = normalized_phone if raw_phone else None

        for field in ['is_staff', 'is_superuser', 'is_active', 'is_sms_enabled']:
            if field in request.data:
                setattr(user, field, bool(request.data[field]))

        if 'sender_id_type' in request.data or 'sender_id' in request.data:
            incoming_type = request.data.get('sender_id_type', user.sender_id_type)
            incoming_sender_id = request.data.get('sender_id', user.sender_id or '')

            try:
                normalized_type, normalized_sender_id = _normalize_sender_id(incoming_type, incoming_sender_id)
            except ValueError as exc:
                return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

            if normalized_sender_id and _sender_id_exists(normalized_sender_id, exclude_user_id=user.id):
                suggestions = _build_sender_id_suggestions(
                    normalized_sender_id,
                    normalized_type,
                    exclude_user_id=user.id,
                )
                return Response(
                    {
                        'detail': 'This sender ID is already used by another user.',
                        'suggestions': suggestions,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            user.sender_id_type = normalized_type
            user.sender_id = normalized_sender_id or None
            if normalized_sender_id and not str(getattr(user, 'free_trial_sender_id', '') or '').strip():
                user.free_trial_sender_id = normalized_sender_id

        if 'free_trial_sender_id' in request.data:
            free_trial_sender_id = str(request.data.get('free_trial_sender_id') or '').strip()
            if free_trial_sender_id:
                provider_config = _get_admin_managed_sms_provider_config()
                if not provider_config:
                    return Response({'detail': 'Admin SMS credentials not configured'}, status=status.HTTP_400_BAD_REQUEST)

                allowed_sender_ids = [
                    str(item).strip()
                    for item in (provider_config.get('sender_ids') or [])
                    if str(item).strip()
                ]
                if free_trial_sender_id not in allowed_sender_ids:
                    return Response(
                        {'detail': 'Selected free trial sender ID is not available in admin credentials'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                user.free_trial_sender_id = free_trial_sender_id
            else:
                user.free_trial_sender_id = None

        try:
            user.save()
        except IntegrityError:
            suggestions = _build_sender_id_suggestions(
                request.data.get('sender_id', user.sender_id or ''),
                request.data.get('sender_id_type', user.sender_id_type),
                exclude_user_id=user.id,
            )
            return Response(
                {
                    'detail': 'This sender ID is already used by another user.',
                    'suggestions': suggestions,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'phone_number': user.phone_number,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'is_active': user.is_active,
            'is_sms_enabled': user.is_sms_enabled,
            'sender_id_type': user.sender_id_type,
            'sender_id': user.sender_id,
            'free_trial_sender_id': user.free_trial_sender_id,
        }, status=200)

    def delete(self, request, user_id):
        guard = _primary_admin_guard(request)
        if guard:
            return guard

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=404)

        if user.id == request.user.id:
            return Response({'detail': 'You cannot delete your own account'}, status=400)

        if _is_primary_admin_email(user.email or ''):
            return Response({'detail': 'Primary admin account cannot be deleted'}, status=400)

        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SMSSendView(generics.CreateAPIView):
    serializer_class = SMSSendSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def create(self, request, *args, **kwargs):
        guard = _primary_admin_guard(request)
        if guard:
            return guard

        self._process_due_scheduled_messages()

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        validated_data = serializer.validated_data
        transport = validated_data.get('transport') or 'api'
        display_sender_id = validated_data['display_sender_id']
        sms_type = validated_data['sms_type']
        message_content = validated_data['message_content']
        send_mode = validated_data.get('send_mode') or 'single'
        delivery_mode = validated_data.get('delivery_mode') or 'instant'
        destination_country = validated_data.get('destination_country') or 'OTHER'
        dlt_template_id = str(validated_data.get('dlt_template_id') or getattr(settings, 'SMS_DLT_TEMPLATE_ID', '') or '').strip()
        dlt_entity_id = str(validated_data.get('dlt_entity_id') or getattr(settings, 'SMS_DLT_ENTITY_ID', '') or '').strip()
        dlt_telemarketer_id = str(validated_data.get('dlt_telemarketer_id') or getattr(settings, 'SMS_DLT_TELEMARKETER_ID', '') or '').strip()

        usage_summary = _get_user_sms_usage_summary(request.user)
        wallet_balance = usage_summary.get('wallet_balance')
        if wallet_balance is not None:
            try:
                if float(wallet_balance) <= 0:
                    return Response(
                        {
                            'detail': 'Insufficient wallet balance. Please recharge credits before sending messages.',
                            'wallet_balance': wallet_balance,
                        },
                        status=status.HTTP_402_PAYMENT_REQUIRED,
                    )
            except (TypeError, ValueError):
                pass

        remaining_sms_balance = None

        schedule_at = None
        timezone_name = ''
        if delivery_mode == 'scheduled':
            try:
                schedule_at, timezone_name = self._build_schedule_datetime(validated_data)
            except ValueError as exc:
                return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        provider_config = None
        dispatch_config = {'transport': transport}
        if transport == 'api':
            provider_config = _get_sms_provider_config()
            if not provider_config:
                return Response({'detail': 'SMS credentials not configured'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            dispatch_config.update(provider_config)
        else:
            dispatch_config.update(self._build_smpp_config(validated_data))

        dispatch_config.update({
            'destination_country': destination_country,
            'dlt_template_id': dlt_template_id,
            'dlt_entity_id': dlt_entity_id,
            'dlt_telemarketer_id': dlt_telemarketer_id,
        })

        if send_mode == 'single':
            recipient_user_id = validated_data.get('recipient_user_id')
            recipient_number = _normalize_phone_number(validated_data.get('recipient_number'))
            if not recipient_number:
                return Response({'detail': 'Invalid phone number'}, status=status.HTTP_400_BAD_REQUEST)

            if destination_country == 'IN' and not _is_indian_number(recipient_number):
                return Response({'detail': 'Selected destination country is India, but recipient number is not a valid Indian number'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                _, remaining_sms_balance = _deduct_sms_credits(request.user, 1)
            except ValueError as exc:
                return Response({'detail': str(exc)}, status=status.HTTP_402_PAYMENT_REQUIRED)

            recipient_user = None
            try:
                if recipient_user_id:
                    recipient_user = User.objects.get(id=recipient_user_id)
                    if not recipient_user.is_sms_enabled:
                        return Response({'detail': 'Recipient user has SMS disabled'}, status=status.HTTP_400_BAD_REQUEST)
            except (TypeError, ValueError, User.DoesNotExist):
                recipient_user = None

            sms_msg = self._create_sms_record(
                request=request,
                recipient_number=recipient_number,
                recipient_user=recipient_user,
                display_sender_id=display_sender_id,
                message_content=message_content,
                sms_type=sms_type,
                send_mode=send_mode,
                schedule_at=schedule_at,
                timezone_name=timezone_name,
                batch_reference='',
                source_file_name='',
            )

            send_result = self._dispatch_or_schedule_message(sms_msg, dispatch_config)
            if transport == 'api':
                self._persist_sender_id(provider_config, display_sender_id)

            response_payload = SMSMessageSerializer(sms_msg).data
            response_payload['delivery_action'] = send_result
            response_payload['transport'] = transport
            response_payload['remaining_sms_credits'] = str(remaining_sms_balance) if remaining_sms_balance is not None else None
            response_payload['dlt'] = {
                'template_id': dlt_template_id,
                'entity_id': dlt_entity_id,
                'telemarketer_id': dlt_telemarketer_id,
            }
            if transport == 'smpp':
                response_payload['dlt']['smpp_profile'] = dispatch_config.get('profile') or 'standard'
            if send_result == 'scheduled':
                return Response(response_payload, status=status.HTTP_202_ACCEPTED)
            return Response(response_payload, status=status.HTTP_201_CREATED)

        source_file = validated_data.get('source_file')
        source_file_name = (source_file.name if source_file else '')
        batch_reference = uuid.uuid4().hex[:16]

        targets = []
        skipped_rows = 0
        if send_mode == 'file_numbers':
            try:
                recipient_numbers = self._extract_numbers_from_uploaded_file(source_file)
            except ValueError as exc:
                return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
            targets = [{'recipient_number': number, 'message_content': message_content} for number in recipient_numbers]

        elif send_mode == 'personalized_file':
            try:
                rows = self._extract_rows_from_excel(source_file)
            except ValueError as exc:
                return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
            for row_values in rows:
                recipient_number = _extract_phone_from_row(row_values)
                if not recipient_number:
                    skipped_rows += 1
                    continue

                rendered_message = _render_personalized_template(message_content, row_values)
                if not rendered_message:
                    skipped_rows += 1
                    continue

                try:
                    calculate_sms_segments(rendered_message, max_segments=int(getattr(settings, 'SMS_MAX_SEGMENTS', 10) or 10))
                except ValueError:
                    skipped_rows += 1
                    continue

                targets.append({'recipient_number': recipient_number, 'message_content': rendered_message})

        elif send_mode == 'group':
            try:
                group = SMSContactGroup.objects.get(id=validated_data.get('group_id'), owner=request.user)
            except SMSContactGroup.DoesNotExist:
                return Response({'detail': 'Selected group not found'}, status=status.HTTP_404_NOT_FOUND)

            contacts = group.contacts.all().values_list('phone_number', flat=True)
            targets = [{'recipient_number': _normalize_phone_number(number), 'message_content': message_content} for number in contacts]
            targets = [item for item in targets if item['recipient_number']]

        if not targets:
            if send_mode == 'personalized_file':
                return Response(
                    {
                        'detail': (
                            'No valid recipients found in uploaded personalized file. '
                            'Check that at least one column contains 10+ digit phone numbers '
                            'and rendered message length stays within configured SMS segment limits.'
                        ),
                        'skipped_rows': skipped_rows,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response({'detail': 'No valid recipients found for selected mode'}, status=status.HTTP_400_BAD_REQUEST)

        payable_target_count = 0
        for target in targets:
            if destination_country == 'IN' and not _is_indian_number(target['recipient_number']):
                continue
            payable_target_count += 1

        if payable_target_count <= 0:
            return Response({'detail': 'No valid recipients found for selected mode'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            _, remaining_sms_balance = _deduct_sms_credits(request.user, payable_target_count)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_402_PAYMENT_REQUIRED)

        sent_count = 0
        failed_count = 0
        scheduled_count = 0
        message_ids = []
        failure_reason_counts = {}
        failed_examples = []

        for target in targets:
            if destination_country == 'IN' and not _is_indian_number(target['recipient_number']):
                failed_count += 1
                reason = 'Recipient number is not valid for India destination selection'
                failure_reason_counts[reason] = failure_reason_counts.get(reason, 0) + 1
                if len(failed_examples) < 20:
                    failed_examples.append({
                        'recipient_number': target['recipient_number'],
                        'reason': reason,
                    })
                continue

            sms_msg = self._create_sms_record(
                request=request,
                recipient_number=target['recipient_number'],
                recipient_user=None,
                display_sender_id=display_sender_id,
                message_content=target['message_content'],
                sms_type=sms_type,
                send_mode=send_mode,
                schedule_at=schedule_at,
                timezone_name=timezone_name,
                batch_reference=batch_reference,
                source_file_name=source_file_name,
            )

            result = self._dispatch_or_schedule_message(sms_msg, dispatch_config)
            if result == 'scheduled':
                scheduled_count += 1
            elif result == 'sent':
                sent_count += 1
            else:
                failed_count += 1
                failure_reason = (sms_msg.failure_reason or 'Unknown failure').strip() or 'Unknown failure'
                failure_reason_counts[failure_reason] = failure_reason_counts.get(failure_reason, 0) + 1
                if len(failed_examples) < 20:
                    failed_examples.append({
                        'recipient_number': sms_msg.recipient_number,
                        'reason': failure_reason,
                    })

            if sms_msg.message_id:
                message_ids.append(sms_msg.message_id)

        if transport == 'api':
            self._persist_sender_id(provider_config, display_sender_id)

        return Response(
            {
                'detail': 'Bulk SMS processing complete',
                'transport': transport,
                'dlt': {
                    'template_id': dlt_template_id,
                    'entity_id': dlt_entity_id,
                    'telemarketer_id': dlt_telemarketer_id,
                    'smpp_profile': dispatch_config.get('profile') if transport == 'smpp' else '',
                },
                'send_mode': send_mode,
                'batch_reference': batch_reference,
                'total_targets': len(targets),
                'sent_count': sent_count,
                'scheduled_count': scheduled_count,
                'failed_count': failed_count,
                'skipped_rows': skipped_rows,
                'remaining_sms_credits': str(remaining_sms_balance) if remaining_sms_balance is not None else None,
                'message_ids': message_ids[:50],
                'failure_summary': [
                    {'reason': reason, 'count': count}
                    for reason, count in sorted(failure_reason_counts.items(), key=lambda item: item[1], reverse=True)
                ],
                'failed_examples': failed_examples,
            },
            status=status.HTTP_201_CREATED,
        )

    def _build_smpp_config(self, validated_data):
        return {
            'host': str(validated_data.get('smpp_host') or '').strip(),
            'port': int(validated_data.get('smpp_port') or 2775),
            'system_id': str(validated_data.get('smpp_system_id') or '').strip(),
            'password': str(validated_data.get('smpp_password') or '').strip(),
            'profile': str(validated_data.get('smpp_profile') or 'standard').strip(),
            'template_id': str(validated_data.get('smpp_template_id') or '').strip(),
            'source_addr_ton': int(validated_data.get('smpp_source_addr_ton') or 5),
            'source_addr_npi': int(validated_data.get('smpp_source_addr_npi') or 0),
            'dest_addr_ton': int(validated_data.get('smpp_dest_addr_ton') or 1),
            'dest_addr_npi': int(validated_data.get('smpp_dest_addr_npi') or 1),
            'data_coding': int(validated_data.get('smpp_data_coding') or 0),
            'registered_delivery': bool(validated_data.get('smpp_registered_delivery', True)),
            'destination_country': validated_data.get('destination_country') or 'OTHER',
            'dlt_template_id': str(validated_data.get('dlt_template_id') or '').strip(),
            'dlt_entity_id': str(validated_data.get('dlt_entity_id') or '').strip(),
            'dlt_telemarketer_id': str(validated_data.get('dlt_telemarketer_id') or '').strip(),
        }

    def _create_sms_record(
        self,
        request,
        recipient_number,
        recipient_user,
        display_sender_id,
        message_content,
        sms_type,
        send_mode,
        schedule_at,
        timezone_name,
        batch_reference,
        source_file_name,
    ):
        sms_record = SMSMessage.objects.create(
            sender=request.user,
            recipient_number=recipient_number,
            recipient_user=recipient_user,
            display_sender_id=display_sender_id,
            message_content=message_content,
            sms_type=sms_type,
            send_mode=send_mode,
            schedule_type='scheduled' if schedule_at else 'instant',
            scheduled_at=schedule_at,
            timezone_name=timezone_name,
            batch_reference=batch_reference,
            source_file_name=source_file_name,
            status='pending',
        )
        _assign_sms_request_id(sms_record, request.user)
        return sms_record

    def _dispatch_or_schedule_message(self, sms_msg, provider_config):
        if sms_msg.scheduled_at and sms_msg.scheduled_at > timezone.now():
            return 'scheduled'

        try:
            if provider_config.get('transport') == 'smpp':
                api_result = self._send_sms_via_smpp(
                    provider_config,
                    sms_msg.display_sender_id,
                    sms_msg.recipient_number,
                    sms_msg.message_content,
                )
            else:
                api_result = self._send_sms_via_api(
                    provider_config['user'],
                    provider_config['password'],
                    sms_msg.display_sender_id,
                    sms_msg.recipient_number,
                    sms_msg.message_content,
                    destination_country=provider_config.get('destination_country') or 'OTHER',
                    dlt_template_id=provider_config.get('dlt_template_id') or '',
                    dlt_entity_id=provider_config.get('dlt_entity_id') or '',
                    dlt_telemarketer_id=provider_config.get('dlt_telemarketer_id') or '',
                )
            sms_msg.provider_message_id = str(api_result.get('message_id') or '')
            sms_msg.status = api_result.get('status', 'sent')
            sms_msg.delivery_time = timezone.now() if sms_msg.status in ['sent', 'delivered'] else None
            sms_msg.failure_reason = ''
            sms_msg.save(update_fields=['provider_message_id', 'status', 'delivery_time', 'failure_reason', 'updated_at'])
            return 'sent'
        except Exception as exc:
            sms_msg.status = 'failed'
            sms_msg.failure_reason = str(exc)
            sms_msg.save(update_fields=['status', 'failure_reason', 'updated_at'])
            return 'failed'

    def _build_schedule_datetime(self, validated_data):
        timezone_name = (validated_data.get('timezone_name') or '').strip()
        start_date = validated_data.get('start_date')
        start_time = validated_data.get('start_time')

        try:
            selected_zone = ZoneInfo(timezone_name)
        except Exception:
            raise ValueError('Invalid timezone selected')

        combined = datetime.combine(start_date, start_time)
        localized = combined.replace(tzinfo=selected_zone)
        scheduled_utc = localized.astimezone(dt_timezone.utc)

        if scheduled_utc <= timezone.now():
            raise ValueError('Scheduled date/time must be in the future')

        return scheduled_utc, timezone_name

    def _extract_numbers_from_uploaded_file(self, source_file):
        if not source_file:
            raise ValueError('No file uploaded')

        filename = (source_file.name or '').lower()
        normalized_numbers = []
        seen = set()

        def _append_number(raw_value):
            normalized = _normalize_phone_number(raw_value)
            if normalized and normalized not in seen:
                seen.add(normalized)
                normalized_numbers.append(normalized)

        def _consume_text_line(text_line):
            text_line = str(text_line or '').strip()
            if not text_line:
                return

            for token in re.split(r'[\s,;|]+', text_line):
                _append_number(token)

            _append_number(text_line)

        if filename.endswith('.txt'):
            source_file.seek(0)
            raw_text = source_file.read().decode('utf-8', errors='ignore')
            for line in raw_text.splitlines():
                _consume_text_line(line)

            if not normalized_numbers:
                raise ValueError(
                    'No valid phone numbers found in TXT file. '
                    'Add one number per line or comma-separated numbers (minimum 10 digits each).'
                )
            return normalized_numbers

        if filename.endswith('.xls'):
            try:
                import xlrd
            except ImportError:
                raise ValueError('XLS upload requires xlrd package on server. Install xlrd or upload .xlsx/.txt file.')

            try:
                source_file.seek(0)
                workbook = xlrd.open_workbook(file_contents=source_file.read())
                for sheet in workbook.sheets():
                    for row_index in range(sheet.nrows):
                        for col_index in range(sheet.ncols):
                            _append_number(sheet.cell_value(row_index, col_index))
            except Exception as exc:
                raise ValueError(f'Unable to read XLS file: {exc}')

            if not normalized_numbers:
                raise ValueError(
                    'No valid phone numbers found in XLS file. '
                    'Ensure sheet contains mobile numbers with at least 10 digits.'
                )
            return normalized_numbers

        try:
            import openpyxl
            source_file.seek(0)
            workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
            worksheet = workbook.active
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    _append_number(cell)

            workbook.close()

            if not normalized_numbers:
                raise ValueError(
                    'No valid phone numbers found in Excel file. '
                    'Ensure at least one column contains mobile numbers with minimum 10 digits.'
                )
            return normalized_numbers
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError(f'Unable to read uploaded file: {exc}')

    def _extract_rows_from_excel(self, source_file):
        if not source_file:
            raise ValueError('No file uploaded')

        filename = (source_file.name or '').lower()
        rows = []

        if filename.endswith('.xls'):
            try:
                import xlrd
            except ImportError:
                raise ValueError('XLS upload requires xlrd package on server. Install xlrd or upload .xlsx file.')

            try:
                source_file.seek(0)
                workbook = xlrd.open_workbook(file_contents=source_file.read())
                for sheet in workbook.sheets():
                    for row_index in range(sheet.nrows):
                        values = [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
                        if any(str(value).strip() for value in values):
                            rows.append(values)
            except Exception as exc:
                raise ValueError(f'Unable to read XLS file: {exc}')

            if not rows:
                raise ValueError('Uploaded XLS file is empty or has no readable rows.')
            return rows

        try:
            import openpyxl
            source_file.seek(0)
            workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
            worksheet = workbook.active
            for row in worksheet.iter_rows(values_only=True):
                values = [item if item is not None else '' for item in row]
                if any(str(value).strip() for value in values):
                    rows.append(values)

            workbook.close()
        except Exception as exc:
            raise ValueError(f'Unable to read Excel file: {exc}')

        if not rows:
            raise ValueError('Uploaded Excel file is empty or has no readable rows.')
        return rows

    def _persist_sender_id(self, provider_config, display_sender_id):
        normalized_sender = (display_sender_id or '').strip()
        if not normalized_sender:
            return

        cred = provider_config.get('credential')
        if cred:
            existing_sender_ids = [str(item).strip() for item in (cred.sender_ids or []) if str(item).strip()]
            if normalized_sender not in existing_sender_ids:
                cred.sender_ids = [*existing_sender_ids, normalized_sender]
                cred.save(update_fields=['sender_ids', 'updated_at'])
            return

        seeded_sender_ids = provider_config.get('sender_ids', [])
        combined_sender_ids = [*seeded_sender_ids]
        if normalized_sender not in combined_sender_ids:
            combined_sender_ids.append(normalized_sender)
        SMSCredential.objects.create(
            user=provider_config['user'],
            password=provider_config['password'],
            sender_ids=combined_sender_ids,
            is_active=True,
        )

    def _process_due_scheduled_messages(self):
        provider_config = _get_sms_provider_config()
        if not provider_config:
            return

        due_messages = SMSMessage.objects.filter(
            status='pending',
            schedule_type='scheduled',
            scheduled_at__isnull=False,
            scheduled_at__lte=timezone.now(),
        ).order_by('scheduled_at')[:100]

        for sms_msg in due_messages:
            self._dispatch_or_schedule_message(sms_msg, provider_config)

    def _send_sms_via_api(
        self,
        user,
        password,
        sender_id,
        number,
        message,
        destination_country='OTHER',
        dlt_template_id='',
        dlt_entity_id='',
        dlt_telemarketer_id='',
    ):
        primary_url = getattr(settings, 'SMS_PROVIDER_URL', 'https://mshastra.com/bsms/buser/send_sms_center.aspx')
        json_fallback_url = getattr(settings, 'SMS_PROVIDER_JSON_URL', 'https://mshastra.com/sendsms_api_json.aspx')
        normalized_number = _normalize_phone_number(number)
        candidate_numbers = []
        if normalized_number:
            candidate_numbers.append(normalized_number)
            if len(normalized_number) == 10:
                candidate_numbers.append(f'91{normalized_number}')
            elif len(normalized_number) == 11 and normalized_number.startswith('0'):
                candidate_numbers.append(f'91{normalized_number[1:]}')
        else:
            raw_number = str(number or '').strip()
            if raw_number:
                candidate_numbers.append(raw_number)

        candidate_numbers = list(dict.fromkeys([item for item in candidate_numbers if item]))
        payload_variants = []
        for candidate_number in candidate_numbers:
            payload_item = {
                "user": user,
                "pwd": password,
                "number": candidate_number,
                "msg": message,
                "sender": sender_id,
                "language": "English"
            }

            if str(destination_country or '').upper() == 'IN':
                if dlt_template_id:
                    payload_item['templateid'] = dlt_template_id
                if dlt_entity_id:
                    payload_item['entityid'] = dlt_entity_id
                if dlt_telemarketer_id:
                    payload_item['telemarketerid'] = dlt_telemarketer_id

            payload_variants.append(payload_item)
            payload_variants.append([payload_item])

        if not payload_variants:
            raise Exception('Invalid recipient number')
        headers = {"Content-Type": "application/json"}

        def _post_with_retry(url, payload, max_attempts=3):
            last_exception = None
            for attempt in range(1, max_attempts + 1):
                try:
                    response = requests.post(url, json=payload, headers=headers, timeout=20)
                    if response.status_code >= 500 and attempt < max_attempts:
                        time.sleep(1.2)
                        continue
                    return response
                except requests.RequestException as exc:
                    last_exception = exc
                    if attempt < max_attempts:
                        time.sleep(1.2)
                        continue
                    raise Exception(f"SMS transport error: {type(exc).__name__}")
            if last_exception:
                raise Exception(f"SMS transport error: {type(last_exception).__name__}")
            raise Exception("SMS transport error")

        def _parse_response(resp):
            if resp.status_code != 200:
                error_text = (resp.text or '').strip().replace('\n', ' ')[:260]
                raise Exception(f"SMS API error {resp.status_code}: {error_text or 'No response body'}")

            try:
                result = resp.json()
            except ValueError:
                text = (resp.text or '').strip()
                lowered = text.lower()

                if not text:
                    raise Exception('Empty SMS provider response')

                if 'session time out' in lowered or 'login again' in lowered:
                    return {'retry_with_json_endpoint': True}

                if any(keyword in lowered for keyword in ['error', 'failed', 'invalid', 'unauthorized']):
                    raise Exception(f'SMS provider failure response: {text[:220]}')

                success_indicators = ['success', 'submitted', 'accepted', 'queued', 'sent']
                if any(keyword in lowered for keyword in success_indicators):
                    message_id_match = re.search(r'(msgid|message[_\s-]?id|id)\s*[:=]\s*([A-Za-z0-9_-]+)', text, flags=re.IGNORECASE)
                    message_id = message_id_match.group(2) if message_id_match else None
                    return {
                        'message_id': message_id or f'text-{timezone.now().timestamp()}',
                        'status': 'sent'
                    }

                raise Exception('SMS provider returned ambiguous response')

            if isinstance(result, list) and len(result) > 0 and isinstance(result[0], dict):
                item = result[0]
                msg_id = item.get('MsgId') or item.get('msgid') or item.get('id') or item.get('message_id')
                status_text = str(item.get('Status') or item.get('status') or '').lower()
                if any(keyword in status_text for keyword in ['fail', 'invalid', 'error']):
                    provider_reason = item.get('Reason') or item.get('reason') or item.get('Message') or item.get('message')
                    if provider_reason:
                        raise Exception(f'SMS provider failure status: {provider_reason}')
                    raise Exception(f'SMS provider returned failure status: {status_text or "failed"}')
                status_value = 'delivered' if 'deliver' in status_text else 'sent'
                return {'message_id': str(msg_id) if msg_id else f'api-{timezone.now().timestamp()}', 'status': status_value}

            if isinstance(result, dict):
                msg_id = result.get('MsgId') or result.get('msgid') or result.get('id') or result.get('message_id')
                status_text = str(result.get('Status') or result.get('status') or '').lower()
                if any(keyword in status_text for keyword in ['fail', 'invalid', 'error']):
                    provider_reason = result.get('Reason') or result.get('reason') or result.get('Message') or result.get('message')
                    if provider_reason:
                        raise Exception(f'SMS provider failure status: {provider_reason}')
                    raise Exception(f'SMS provider returned failure status: {status_text or "failed"}')
                status_value = 'delivered' if 'deliver' in status_text else 'sent'
                return {'message_id': str(msg_id) if msg_id else f'api-{timezone.now().timestamp()}', 'status': status_value}

            raise Exception('Unexpected SMS provider response format')

        def _send_and_parse(url):
            last_error = None
            for payload_variant in payload_variants:
                try:
                    resp = _post_with_retry(url, payload_variant)
                    return _parse_response(resp)
                except Exception as exc:
                    last_error = exc
            if last_error:
                raise last_error
            raise Exception('SMS provider call failed')

        parsed_primary = _send_and_parse(primary_url)

        if parsed_primary.get('retry_with_json_endpoint'):
            parsed_fallback = _send_and_parse(json_fallback_url)
            if parsed_fallback.get('retry_with_json_endpoint'):
                raise Exception('SMS provider endpoint requires interactive login')
            return parsed_fallback

        return parsed_primary

    def _send_sms_via_smpp(self, smpp_config, sender_id, number, message):
        try:
            import smpplib.client
        except ImportError as exc:
            raise Exception('SMPP support is not installed on the server') from exc

        normalized_number = _normalize_phone_number(number)
        if not normalized_number:
            raise Exception('Invalid recipient number')

        client = smpplib.client.Client(smpp_config['host'], smpp_config['port'])
        client.socket_timeout = 20

        try:
            client.connect()
            client.bind_transceiver(
                system_id=smpp_config['system_id'],
                password=smpp_config['password'],
            )

            send_kwargs = {
                'source_addr_ton': smpp_config['source_addr_ton'],
                'source_addr_npi': smpp_config['source_addr_npi'],
                'source_addr': sender_id,
                'dest_addr_ton': smpp_config['dest_addr_ton'],
                'dest_addr_npi': smpp_config['dest_addr_npi'],
                'destination_addr': normalized_number,
                'short_message': str(message or '').encode('utf-8'),
                'data_coding': smpp_config['data_coding'],
                'registered_delivery': smpp_config['registered_delivery'],
            }

            template_id = str(smpp_config.get('dlt_template_id') or smpp_config.get('template_id') or '').strip()
            entity_id = str(smpp_config.get('dlt_entity_id') or '').strip()
            telemarketer_id = str(smpp_config.get('dlt_telemarketer_id') or '').strip()
            if template_id or entity_id or telemarketer_id:
                optional_parameters = {}
                if entity_id:
                    optional_parameters[0x1400] = entity_id.encode()
                if template_id:
                    optional_parameters[0x1401] = template_id.encode()
                if telemarketer_id:
                    optional_parameters[0x1402] = telemarketer_id.encode()
                send_kwargs['optional_parameters'] = optional_parameters

            pdu = client.send_message(**send_kwargs)
            message_id = getattr(pdu, 'message_id', None) or getattr(pdu, 'sequence', None)

            return {
                'message_id': str(message_id) if message_id is not None else f'smpp-{timezone.now().timestamp()}',
                'status': 'sent',
            }
        except (socket.timeout, OSError) as exc:
            raise Exception(f'SMPP transport error: {type(exc).__name__}') from exc
        except Exception as exc:
            raise Exception(str(exc)) from exc
        finally:
            try:
                client.unbind()
            except Exception:
                pass
            try:
                client.disconnect()
            except Exception:
                pass


class SMSTimezoneListView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        query = (request.query_params.get('q') or '').strip().lower()
        now_utc = datetime.now(dt_timezone.utc)
        rows = []

        try:
            import pytz

            for country_code, zone_list in sorted(pytz.country_timezones.items(), key=lambda item: pytz.country_names.get(item[0], item[0])):
                country_name = pytz.country_names.get(country_code, country_code)

                if query and query not in country_name.lower() and not any(query in zone.lower() for zone in zone_list):
                    continue

                for timezone_name in zone_list:
                    try:
                        zone = ZoneInfo(timezone_name)
                        localized_now = now_utc.astimezone(zone)
                        offset_label, offset_compact, offset_minutes = _format_utc_offset(localized_now.utcoffset())
                    except Exception:
                        offset_label, offset_compact, offset_minutes = 'UTC+00:00', '+0.00', 0

                    city_label = timezone_name.split('/')[-1].replace('_', ' ')
                    rows.append(
                        {
                            'country_code': country_code,
                            'country_name': country_name,
                            'timezone_name': timezone_name,
                            'city_label': city_label,
                            'offset_label': offset_label,
                            'offset_compact': offset_compact,
                            'offset_minutes': offset_minutes,
                            'display_label': f'{city_label} ({offset_compact})',
                        }
                    )

            rows.sort(key=lambda item: (item['country_name'], item['offset_minutes'], item['timezone_name']))
            return Response({'count': len(rows), 'results': rows}, status=status.HTTP_200_OK)
        except Exception:
            fallback_zones = sorted(available_timezones())
            for timezone_name in fallback_zones:
                if query and query not in timezone_name.lower():
                    continue

                try:
                    zone = ZoneInfo(timezone_name)
                    localized_now = now_utc.astimezone(zone)
                    offset_label, offset_compact, offset_minutes = _format_utc_offset(localized_now.utcoffset())
                except Exception:
                    offset_label, offset_compact, offset_minutes = 'UTC+00:00', '+0.00', 0

                rows.append(
                    {
                        'country_code': 'ZZ',
                        'country_name': 'Other',
                        'timezone_name': timezone_name,
                        'city_label': timezone_name.replace('_', ' '),
                        'offset_label': offset_label,
                        'offset_compact': offset_compact,
                        'offset_minutes': offset_minutes,
                        'display_label': f'{timezone_name} ({offset_compact})',
                    }
                )

            rows.sort(key=lambda item: (item['offset_minutes'], item['timezone_name']))
            return Response({'count': len(rows), 'results': rows}, status=status.HTTP_200_OK)


class SMSUsageSummaryView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        summary = _get_user_sms_usage_summary(request.user)
        summary['free_trial_limit'] = FREE_TRIAL_MESSAGE_LIMIT
        summary['is_admin'] = _has_primary_admin_access(request.user)
        summary['verified_numbers_count'] = FreeTrialVerifiedNumber.objects.filter(owner=request.user, is_verified=True).count()
        return Response(summary, status=status.HTTP_200_OK)


class FreeTrialVerifiedNumbersView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        number = _normalize_phone_number(request.user.phone_number)
        numbers = [number] if number and not _has_primary_admin_access(request.user) else []
        return Response({'verified_numbers': numbers}, status=status.HTTP_200_OK)


class FreeTrialSendOTPView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if _has_primary_admin_access(request.user):
            return Response({'detail': 'Admin users already have full SMS access'}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                'detail': 'Free trial messages can only be sent to your signup mobile number. OTP verification is no longer required.',
                'otp_required': False,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )


class FreeTrialVerifyOTPView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        if request.user.is_authenticated and _has_primary_admin_access(request.user):
            return Response({'detail': 'Admin users do not require free trial OTP'}, status=status.HTTP_400_BAD_REQUEST)

        verified_numbers = []
        if request.user.is_authenticated:
            signup_number = _normalize_phone_number(request.user.phone_number)
            if signup_number:
                verified_numbers = [signup_number]

        return Response(
            {
                'detail': 'Free trial messages now go only to your signup mobile number. OTP verification is no longer required.',
                'verified': True,
                'verified_numbers': verified_numbers,
            },
            status=status.HTTP_200_OK,
        )


class FreeTrialSendSMSView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if _has_primary_admin_access(request.user):
            return Response({'detail': 'Admin users should use regular SMS sending'}, status=status.HTTP_400_BAD_REQUEST)

        usage_summary = _get_user_sms_usage_summary(request.user)
        if usage_summary['used_messages'] >= FREE_TRIAL_MESSAGE_LIMIT:
            return Response(
                {
                    'detail': 'You have used all 3 free trial messages. Please upgrade to continue.',
                    'used_messages': usage_summary['used_messages'],
                    'available_messages': usage_summary['available_messages'],
                    'free_trial_complete': True,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        recipient_number = _normalize_phone_number(request.user.phone_number)
        if not recipient_number:
            return Response({'detail': 'Add your mobile number in signup/profile before using free trial SMS'}, status=status.HTTP_400_BAD_REQUEST)

        message_content = str(request.data.get('message_content') or '').strip()
        if not message_content:
            return Response({'detail': 'Message content is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            calculate_sms_segments(message_content, max_segments=int(getattr(settings, 'SMS_MAX_SEGMENTS', 10) or 10))
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        provider_config = _get_admin_managed_sms_provider_config()
        if not provider_config:
            return Response({'detail': 'Free trial SMS service is temporarily unavailable. Please try again later.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        try:
            display_sender_id = _resolve_free_trial_sender_id(request.user, provider_config)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        mediator_user = _get_free_trial_mediator_user() or request.user

        sms_msg = SMSMessage.objects.create(
            sender=mediator_user,
            recipient_number=recipient_number,
            recipient_user=request.user,
            display_sender_id=display_sender_id,
            message_content=message_content,
            sms_type='transactional',
            send_mode='free_trial',
            schedule_type='instant',
            status='pending',
            batch_reference=f'free-trial-{request.user.id}',
            source_file_name='',
        )
        _assign_sms_request_id(sms_msg, request.user)

        try:
            api_result = SMSSendView()._send_sms_via_api(
                provider_config['user'],
                provider_config['password'],
                display_sender_id,
                recipient_number,
                message_content,
            )
            sms_msg.provider_message_id = str(api_result.get('message_id') or '')
            sms_msg.status = api_result.get('status', 'sent')
            sms_msg.delivery_time = timezone.now() if sms_msg.status in ['sent', 'delivered'] else None
            sms_msg.failure_reason = ''
            sms_msg.save(update_fields=['provider_message_id', 'status', 'delivery_time', 'failure_reason', 'updated_at'])
        except Exception as exc:
            sms_msg.status = 'failed'
            sms_msg.failure_reason = str(exc)
            sms_msg.save(update_fields=['status', 'failure_reason', 'updated_at'])
            return Response({'detail': f'Failed to send SMS: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

        updated_summary = _get_user_sms_usage_summary(request.user)
        return Response(
            {
                'detail': 'Free trial SMS sent successfully to your signup mobile number',
                'recipient_number': recipient_number,
                'display_sender_id': display_sender_id,
                'message_id': sms_msg.message_id,
                'status': sms_msg.status,
                'used_messages': updated_summary['used_messages'],
                'available_messages': updated_summary['available_messages'],
                'free_trial_complete': updated_summary['available_messages'] <= 0,
            },
            status=status.HTTP_201_CREATED,
        )


class SMSMessageListView(generics.ListAPIView):
    serializer_class = SMSMessageSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        try:
            SMSSendView()._process_due_scheduled_messages()
        except Exception:
            pass

        user = self.request.user
        if _has_support_read_access(user):
            queryset = SMSMessage.objects.all()
        else:
            queryset = SMSMessage.objects.filter(Q(sender=user) | Q(recipient_user=user))

        query = str(self.request.query_params.get('q') or '').strip()
        if query:
            queryset = queryset.filter(
                Q(message_id__icontains=query)
                | Q(provider_message_id__icontains=query)
                | Q(batch_reference__icontains=query)
                | Q(recipient_number__icontains=query)
                | Q(display_sender_id__icontains=query)
                | Q(message_content__icontains=query)
                | Q(status__icontains=query)
            )

        return queryset


class SMSMessageStatusView(generics.RetrieveAPIView):
    queryset = SMSMessage.objects.all()
    serializer_class = SMSMessageStatusSerializer
    permission_classes = [permissions.IsAuthenticated]


class SMSCredentialView(generics.GenericAPIView):
    serializer_class = SMSCredentialSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = None if _has_admin_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard
        cred = SMSCredential.objects.filter(is_active=True).order_by('-updated_at', '-id').first()
        if not cred:
            env_user = getattr(settings, 'SMS_PROVIDER_USER', '').strip()
            env_sender_ids = [str(item).strip() for item in getattr(settings, 'SMS_DEFAULT_SENDER_IDS', []) if str(item).strip()]
            env_free_trial_sender_id = str(getattr(settings, 'SMS_FREE_TRIAL_DEFAULT_SENDER_ID', '') or '').strip()
            has_env_provider = bool(env_user and getattr(settings, 'SMS_PROVIDER_PASSWORD', '').strip())
            return Response({
                'user': env_user,
                'password': '',
                'sender_ids': env_sender_ids,
                'free_trial_default_sender_id': env_free_trial_sender_id,
                'is_active': has_env_provider,
                'created_at': None,
                'updated_at': None,
            })
        return Response(self.get_serializer(cred).data)

    def patch(self, request):
        guard = None if _has_admin_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard
        cred = SMSCredential.objects.filter(is_active=True).order_by('-updated_at', '-id').first()
        if not cred:
            cred = SMSCredential.objects.create(
                user=request.data.get('user', ''),
                password=request.data.get('password', ''),
                sender_ids=request.data.get('sender_ids', []),
                free_trial_default_sender_id=request.data.get('free_trial_default_sender_id', ''),
                is_active=True,
            )
            SMSCredential.objects.filter(is_active=True).exclude(id=cred.id).update(is_active=False)
            return Response(self.get_serializer(cred).data, status=status.HTTP_201_CREATED)

        serializer = self.get_serializer(cred, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        SMSCredential.objects.filter(is_active=True).exclude(id=cred.id).update(is_active=False)
        return Response(serializer.data, status=status.HTTP_200_OK)


class SMSContactGroupView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        groups = SMSContactGroup.objects.filter(owner=request.user).annotate(member_count=Count('contacts'))
        serializer = SMSContactGroupSerializer(groups, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = SMSContactGroupCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        group_name = serializer.validated_data['name'].strip()
        members = serializer.validated_data['members']

        if not group_name:
            return Response({'detail': 'Group name is required'}, status=status.HTTP_400_BAD_REQUEST)

        group, created = SMSContactGroup.objects.get_or_create(owner=request.user, name=group_name)

        added = 0
        for raw_member in members:
            raw_text = str(raw_member or '').strip()
            if not raw_text:
                continue

            member_name = ''
            if ',' in raw_text:
                left, right = raw_text.rsplit(',', 1)
                normalized = _normalize_phone_number(right)
                if normalized:
                    member_name = left.strip()
                    phone_number = normalized
                else:
                    phone_number = _normalize_phone_number(raw_text)
            else:
                phone_number = _normalize_phone_number(raw_text)

            if not phone_number:
                continue

            _, was_created = SMSContact.objects.get_or_create(
                group=group,
                phone_number=phone_number,
                defaults={'name': member_name},
            )
            if was_created:
                added += 1

        refreshed = SMSContactGroup.objects.filter(id=group.id).annotate(member_count=Count('contacts')).first()
        output = SMSContactGroupSerializer(refreshed).data
        output['created'] = created
        output['added_members'] = added
        return Response(output, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class SMSShortURLView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        urls = SMSShortURL.objects.filter(owner=request.user, is_active=True)
        data = []
        for item in urls:
            serialized = SMSShortURLSerializer(item).data
            serialized['short_url'] = request.build_absolute_uri(f'/s/{item.short_code}/')
            data.append(serialized)
        return Response(data)

    def post(self, request):
        serializer = SMSShortURLSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        short_code = ''
        for _ in range(20):
            candidate = _generate_short_code(7)
            if not SMSShortURL.objects.filter(short_code=candidate).exists():
                short_code = candidate
                break

        if not short_code:
            return Response({'detail': 'Could not generate short URL code'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        short_url = SMSShortURL.objects.create(
            owner=request.user,
            link_name=serializer.validated_data['link_name'],
            short_code=short_code,
            redirect_url=serializer.validated_data['redirect_url'],
            is_active=True,
        )

        output = SMSShortURLSerializer(short_url).data
        output['short_url'] = request.build_absolute_uri(f'/s/{short_code}/')
        return Response(output, status=status.HTTP_201_CREATED)


class SMSShortURLDetailView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def delete(self, request, url_id):
        short_url = SMSShortURL.objects.filter(id=url_id, owner=request.user, is_active=True).first()
        if not short_url:
            return Response({'detail': 'Short URL not found'}, status=status.HTTP_404_NOT_FOUND)

        short_url.is_active = False
        short_url.save(update_fields=['is_active', 'updated_at'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class ShortURLRedirectView(View):
    def get(self, request, short_code, *args, **kwargs):
        short_url = SMSShortURL.objects.filter(short_code=short_code, is_active=True).first()
        if not short_url:
            raise Http404('Short URL not found')

        SMSShortURL.objects.filter(id=short_url.id).update(
            total_clicks=F('total_clicks') + 1,
            last_clicked_at=timezone.now(),
        )
        return HttpResponseRedirect(short_url.redirect_url)


class UserSMSEligibilityView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, user_id):
        guard = _primary_admin_guard(request)
        if guard:
            return guard

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        if 'is_sms_enabled' in request.data:
            user.is_sms_enabled = bool(request.data.get('is_sms_enabled'))
            user.save()

        return Response(UserSMSEligibilitySerializer(user).data)


class AdminUsersSMSListView(generics.ListAPIView):
    serializer_class = UserSMSEligibilitySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if not _has_support_read_access(self.request.user):
            return User.objects.none()
        return User.objects.all().order_by('-date_joined')


class AdminUsersExportView(generics.GenericAPIView):
    """Export all users to Excel - admin only"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import FileResponse
            from io import BytesIO
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"
            
            # Add header row with styling
            headers = ['ID', 'Username', 'Email', 'Phone Number', 'Status', 'Account Type', 'Joined Date']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add user data
            users = User.objects.all()
            for row_num, user in enumerate(users, 2):
                ws.cell(row=row_num, column=1).value = user.id
                ws.cell(row=row_num, column=2).value = user.username
                ws.cell(row=row_num, column=3).value = user.email
                ws.cell(row=row_num, column=4).value = user.phone_number or '-'
                ws.cell(row=row_num, column=5).value = 'Verified' if user.is_active else 'Not Verified'
                ws.cell(row=row_num, column=6).value = 'Admin' if user.is_staff else 'User'
                ws.cell(row=row_num, column=7).value = user.date_joined.strftime('%Y-%m-%d')
                
                # Center align status columns
                for col in [5, 6]:
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            
            # Save to BytesIO
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            
            # Return file
            response = FileResponse(
                stream,
                as_attachment=True,
                filename=f'users-export-{timezone.now().strftime("%Y%m%d")}.xlsx'
            )
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response
            
        except ImportError:
            return Response(
                {'detail': 'openpyxl package is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'detail': f'Error generating Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AdminNotificationPreviewView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = _primary_admin_guard(request)
        if guard:
            return guard

        audience_filter = request.query_params.get('audience_filter', 'all_users')
        users = _get_users_for_notification_filter(audience_filter).order_by('-date_joined')
        payload = NotificationRecipientPreviewSerializer(users[:200], many=True).data

        return Response({
            'audience_filter': audience_filter,
            'total_recipients': users.count(),
            'preview_recipients': payload,
        })


class AdminNotificationSendView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AdminNotificationSendSerializer

    def post(self, request):
        guard = _primary_admin_guard(request)
        if guard:
            return guard

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        content = serializer.validated_data['content'].strip()
        audience_filter = serializer.validated_data['audience_filter']
        target_users = list(_get_users_for_notification_filter(audience_filter))

        if not target_users:
            return Response({'detail': 'No users found for selected filter'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            notification = InternalNotification.objects.create(
                content=content,
                audience_filter=audience_filter,
                created_by=request.user,
                recipient_count=len(target_users),
            )

            InternalNotificationRecipient.objects.bulk_create([
                InternalNotificationRecipient(notification=notification, user=user)
                for user in target_users
            ])

        return Response(
            {
                'detail': 'Notification sent successfully',
                'notification': AdminNotificationHistorySerializer(notification).data,
            },
            status=status.HTTP_201_CREATED,
        )


class AdminNotificationHistoryView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AdminNotificationHistorySerializer

    def get_queryset(self):
        if not _has_support_read_access(self.request.user):
            return InternalNotification.objects.none()
        return InternalNotification.objects.select_related('created_by').all()


class UserNotificationListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = UserNotificationSerializer

    def get_queryset(self):
        return InternalNotificationRecipient.objects.filter(user=self.request.user).select_related('notification')


class UserNotificationReadView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, recipient_id):
        recipient = InternalNotificationRecipient.objects.filter(id=recipient_id, user=request.user).first()
        if not recipient:
            return Response({'detail': 'Notification not found'}, status=status.HTTP_404_NOT_FOUND)

        if not recipient.is_read:
            recipient.is_read = True
            recipient.read_at = timezone.now()
            recipient.save(update_fields=['is_read', 'read_at'])

        return Response(UserNotificationSerializer(recipient).data)



class ConfirmAdminPromotionView(generics.GenericAPIView):
    """Handle email confirmation link for admin promotion"""
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def get(self, request):
        token = request.query_params.get('token', '').strip()
        user_id = request.query_params.get('user_id')
        promotion_type = request.query_params.get('type', '').strip().lower()  # optional

        if not token or not user_id:
            return Response(
                {'detail': 'Invalid confirmation link', 'success': False},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(id=int(user_id))
        except (User.DoesNotExist, ValueError):
            return Response(
                {'detail': 'User not found', 'success': False},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if token matches and is not expired (24 hours)
        if user.admin_promotion_token != token:
            return Response(
                {'detail': 'Invalid or expired confirmation token', 'success': False},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not user.pending_admin_promotion:
            return Response(
                {'detail': 'No pending admin promotion for this user', 'success': False},
                status=status.HTTP_400_BAD_REQUEST
            )

        if user.admin_promotion_requested_at:
            expiry_time = user.admin_promotion_requested_at + timedelta(hours=24)
            if timezone.now() > expiry_time:
                user.pending_admin_promotion = False
                user.admin_promotion_token = None
                user.admin_promotion_requested_at = None
                user.save()
                return Response(
                    {'detail': 'Confirmation link has expired (24 hours)', 'success': False},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Infer type from token prefix when not explicitly provided.
        inferred_type = 'staff' if token.startswith('staff::') else 'full_admin'
        resolved_type = promotion_type if promotion_type in ['full_admin', 'staff'] else inferred_type

        # Confirm the promotion
        if resolved_type == 'full_admin':
            user.is_staff = True
            user.is_superuser = True
            promotion_label = 'FULL ADMIN'
        else:
            user.is_staff = True
            user.is_superuser = False
            promotion_label = 'STAFF'

        user.pending_admin_promotion = False
        user.admin_promotion_token = None
        user.admin_promotion_requested_at = None
        user.save()

        return Response(
            {
                'detail': f'✓ Congratulations! You have been successfully promoted to {promotion_label}.',
                'success': True,
                'promotion_type': resolved_type,
                'user_email': user.email,
                'next_step': 'You can now log in with your admin credentials at /admin/',
            },
            status=status.HTTP_200_OK
        )


class EmailValidationView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        source_file = request.FILES.get('source_file') if hasattr(request, 'FILES') else None
        provider_mode = _get_email_validation_provider_mode()
        available_validation_balance = _get_email_validation_wallet_balance(request.user)
        if available_validation_balance is None:
            return Response({'detail': 'Provider credits are unavailable for email validation.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        if available_validation_balance <= 0:
            return Response({'detail': 'No email validation credits available.'}, status=status.HTTP_402_PAYMENT_REQUIRED)

        try:
            unique_emails, file_name = _collect_validation_emails_from_file(request, enforce_request_limit=not bool(source_file))
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cost_deducted, remaining_balance = _deduct_email_validation_credits(request.user, len(unique_emails))
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_402_PAYMENT_REQUIRED)

        if source_file:
            history = EmailValidationHistory.objects.create(
                user=request.user,
                source='dashboard',
                status=EmailValidationHistory.STATUS_PENDING,
                email_count=len(unique_emails),
                emails_requested=unique_emails,
                results_summary={
                    'safe_count': 0,
                    'unsafe_count': 0,
                    'provider_message_id': '',
                    'provider_message_ids': [],
                    'results': [],
                    'queued': True,
                    'control_state': 'running',
                    'processing_state': 'running',
                    'processed_count': 0,
                    'total_count': len(unique_emails),
                    'progress_percent': 0,
                    'elapsed_seconds': 0,
                    'eta_seconds': 0,
                    'started_at': timezone.now().isoformat(),
                },
                cost_deducted=cost_deducted,
                file_name=file_name,
                completed_at=None,
            )
            _assign_email_validation_request_id(history)
            history_payload = EmailValidationHistorySerializer(history).data
            _start_email_validation_worker(history.id)
            return Response(
                {
                    'request_id': history.request_id,
                    'count': len(unique_emails),
                    'wallet_balance': str(remaining_balance),
                    'provider_mode': provider_mode,
                    'source_file_name': file_name,
                    'summary': {
                        'safe_to_send_yes': 0,
                        'safe_to_send_no': 0,
                    },
                    'results': [],
                    'simple_results': [],
                    'history': history_payload,
                    'dlr_report': {
                        'request_id': history.request_id,
                        'status': history.status,
                        'completed': False,
                        'delivery_time': None,
                        'failure_reason': '',
                    },
                },
                status=status.HTTP_202_ACCEPTED,
            )

        results = _validate_email_list(unique_emails, provider_mode=provider_mode)
        client_results = [_to_client_validation_result(item) for item in results]
        safe_count = sum(1 for item in client_results if _is_safe_client_validation_result(item))
        unsafe_count = len(client_results) - safe_count
        provider_request_ids = [
            str(item.get('providerMessageId') or '').strip()
            for item in client_results
            if str(item.get('providerMessageId') or '').strip()
        ]

        history = EmailValidationHistory.objects.create(
            user=request.user,
            source='dashboard',
            status=EmailValidationHistory.STATUS_COMPLETED,
            email_count=len(unique_emails),
            emails_requested=unique_emails,
            results_summary={
                'safe_count': safe_count,
                'unsafe_count': unsafe_count,
                'provider_message_id': provider_request_ids[0] if provider_request_ids else '',
                'provider_message_ids': provider_request_ids,
                'results': client_results,
            },
            cost_deducted=cost_deducted,
            file_name=file_name,
            completed_at=timezone.now(),
        )
        _assign_email_validation_request_id(history)
        history_payload = EmailValidationHistorySerializer(history).data

        simple_results = [_build_simple_validation_result(item) for item in client_results]

        return Response(
            {
                'request_id': history.request_id,
                'count': len(client_results),
                'wallet_balance': str(remaining_balance),
                'provider_mode': provider_mode,
                'source_file_name': file_name,
                'summary': {
                    'safe_to_send_yes': safe_count,
                    'safe_to_send_no': unsafe_count,
                },
                'simple_results': simple_results,
                'results': client_results,
                'history': history_payload,
                'dlr_report': {
                    'request_id': history.request_id,
                    'status': history.status,
                    'completed': history.status == EmailValidationHistory.STATUS_COMPLETED,
                    'delivery_time': history.completed_at,
                    'failure_reason': '',
                },
            },
            status=status.HTTP_200_OK,
        )


class UserWalletView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        wallet = _get_or_create_wallet(request.user)
        if Decimal(str(wallet.email_validation_balance or 0)).quantize(Decimal('0.0001')) != Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001')):
            wallet.email_validation_balance = wallet.balance
            wallet.save(update_fields=['email_validation_balance', 'updated_at'])
        payload = UserWalletSerializer(wallet).data
        payload['email_validation_provider_mode'] = _get_email_validation_provider_mode()
        if _has_admin_access(request.user):
            provider_email_balance = _get_verifalia_admin_credits()
            if provider_email_balance is not None:
                payload['provider_email_balance'] = str(provider_email_balance)
            provider_balance = _get_provider_wallet_balance()
            if provider_balance is not None:
                payload['provider_message_balance'] = str(provider_balance)
        return Response(payload)


class WalletRechargeConfigView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        service_charge_percentage, tax_percentage = _get_recharge_charge_percentages()
        razorpay_config = _get_razorpay_config()
        paypal_checkout_url = _get_platform_setting_text(
            'recharge_paypal_checkout_url',
            str(getattr(settings, 'PAYPAL_CHECKOUT_URL', '') or '').strip(),
        )
        upi_vpa = _get_platform_setting_text(
            'recharge_upi_vpa',
            str(getattr(settings, 'RAZORPAY_UPI_VPA', '') or '').strip(),
        )
        upi_payee_name = _get_platform_setting_text(
            'recharge_upi_payee_name',
            str(getattr(settings, 'COMPANY_NAME', 'Bhisha') or 'Bhisha').strip(),
        )
        return Response(
            {
                'service_charge_percentage': str(service_charge_percentage),
                'tax_percentage': str(tax_percentage),
                'currency': razorpay_config['currency'],
                'razorpay_key_id': razorpay_config['key_id'],
                'gateway_configured': razorpay_config['configured'],
                'sdk_installed': _is_razorpay_sdk_installed(),
                'paypal_checkout_url': paypal_checkout_url,
                'paypal_enabled': bool(paypal_checkout_url),
                'upi_vpa': upi_vpa,
                'upi_payee_name': upi_payee_name or 'Bhisha',
                'upi_enabled': bool(upi_vpa),
            }
        )


class WalletRechargeCreateOrderView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = WalletRechargeCreateOrderSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        selected_method = serializer.validated_data.get('payment_method', 'upi')

        razorpay_config = _get_razorpay_config()
        if not razorpay_config['configured']:
            return Response({'detail': 'Payment gateway is not configured. Set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET in backend environment.'}, status=status.HTTP_400_BAD_REQUEST)

        if not _is_razorpay_sdk_installed():
            return Response({'detail': 'Razorpay SDK is not installed on server. Install backend dependency: razorpay.'}, status=status.HTTP_400_BAD_REQUEST)

        razorpay_client = _get_razorpay_client_or_none()
        if not razorpay_client:
            return Response({'detail': 'Payment gateway client initialization failed.'}, status=status.HTTP_400_BAD_REQUEST)

        entered_amount = serializer.validated_data['amount']
        service_charge_percentage, tax_percentage = _get_recharge_charge_percentages()
        breakdown = _calculate_recharge_breakdown(entered_amount, service_charge_percentage, tax_percentage)

        total_amount_paise = int((breakdown['total_amount'] * Decimal('100')).quantize(Decimal('1')))
        if total_amount_paise <= 0:
            return Response({'detail': 'Recharge amount must be greater than zero.'}, status=status.HTTP_400_BAD_REQUEST)

        min_amount_paise = 100
        if total_amount_paise < min_amount_paise:
            return Response(
                {'detail': 'Minimum payable amount is INR 1.00 after charges.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        receipt_value = f"recharge-{request.user.id}-{uuid.uuid4().hex[:12]}"

        try:
            order = razorpay_client.order.create(
                {
                    'amount': total_amount_paise,
                    'currency': razorpay_config['currency'],
                    'receipt': receipt_value,
                    'payment_capture': 1,
                }
            )
        except Exception as exc:
            raw_message = str(exc) or type(exc).__name__
            return Response(
                {
                    'detail': f'Unable to create payment order: {raw_message}',
                    'gateway_error': raw_message,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_order_id = str(order.get('id') or '').strip()
        if not created_order_id:
            return Response(
                {'detail': 'Payment gateway did not return a valid order id.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payment = WalletRechargePayment.objects.create(
            user=request.user,
            entered_amount=breakdown['entered_amount'],
            service_charge_percentage=breakdown['service_charge_percentage'],
            tax_percentage=breakdown['tax_percentage'],
            service_charge_amount=breakdown['service_charge_amount'],
            tax_amount=breakdown['tax_amount'],
            total_amount=breakdown['total_amount'],
            currency=razorpay_config['currency'],
            razorpay_order_id=created_order_id,
            status=WalletRechargePayment.STATUS_PENDING,
        )

        return Response(
            {
                'order': {
                    'id': payment.razorpay_order_id,
                    'amount': total_amount_paise,
                    'currency': payment.currency,
                    'receipt': receipt_value,
                },
                'payment_method': selected_method,
                'charges': {
                    'entered_amount': str(payment.entered_amount),
                    'service_charge_percentage': str(payment.service_charge_percentage),
                    'tax_percentage': str(payment.tax_percentage),
                    'service_charge_amount': str(payment.service_charge_amount),
                    'tax_amount': str(payment.tax_amount),
                    'total_amount': str(payment.total_amount),
                },
                'wallet_credit_amount': str(payment.entered_amount),
            },
            status=status.HTTP_201_CREATED,
        )


class WalletRechargeVerifyView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = WalletRechargeVerifySerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        razorpay_order_id = serializer.validated_data['razorpay_order_id']
        razorpay_payment_id = serializer.validated_data['razorpay_payment_id']
        razorpay_signature = str(serializer.validated_data.get('razorpay_signature') or '').strip()

        razorpay_config = _get_razorpay_config()
        if not razorpay_config['configured']:
            return Response({'detail': 'Payment gateway is not configured.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        payment = WalletRechargePayment.objects.filter(
            user=request.user,
            razorpay_order_id=razorpay_order_id,
        ).first()
        if not payment:
            return Response({'detail': 'Payment order not found.'}, status=status.HTTP_404_NOT_FOUND)

        is_verified = False
        verification_failure_reason = ''

        if razorpay_signature:
            generated_signature = hmac.new(
                razorpay_config['key_secret'].encode('utf-8'),
                f"{razorpay_order_id}|{razorpay_payment_id}".encode('utf-8'),
                hashlib.sha256,
            ).hexdigest()
            is_verified = hmac.compare_digest(generated_signature, razorpay_signature)
            if not is_verified:
                verification_failure_reason = 'Signature verification failed.'
        else:
            razorpay_client = _get_razorpay_client_or_none()
            if not razorpay_client:
                verification_failure_reason = 'Payment gateway verification client unavailable.'
            else:
                try:
                    fetched_payment = razorpay_client.payment.fetch(razorpay_payment_id)
                    fetched_order_id = str(fetched_payment.get('order_id') or '')
                    fetched_status = str(fetched_payment.get('status') or '').lower()
                    fetched_amount = int(fetched_payment.get('amount') or 0)
                    expected_amount = int((Decimal(str(payment.total_amount or 0)) * Decimal('100')).quantize(Decimal('1')))

                    order_matches = fetched_order_id == razorpay_order_id
                    status_ok = fetched_status in {'captured', 'authorized'}
                    amount_matches = fetched_amount == expected_amount
                    is_verified = bool(order_matches and status_ok and amount_matches)

                    if not is_verified:
                        verification_failure_reason = 'Payment verification failed against Razorpay order details.'
                except Exception:
                    verification_failure_reason = 'Unable to verify payment with Razorpay.'

        if not is_verified:
            payment.status = WalletRechargePayment.STATUS_FAILED
            payment.failure_reason = verification_failure_reason or 'Payment verification failed.'
            payment.razorpay_payment_id = razorpay_payment_id
            payment.razorpay_signature = razorpay_signature
            payment.save(update_fields=['status', 'failure_reason', 'razorpay_payment_id', 'razorpay_signature', 'updated_at'])
            return Response({'detail': payment.failure_reason}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            payment = WalletRechargePayment.objects.select_for_update().get(id=payment.id)
            if payment.status != WalletRechargePayment.STATUS_SUCCESSFUL:
                wallet = _get_or_create_wallet(request.user)
                current_balance = Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001'))
                credit_amount = Decimal(str(payment.entered_amount or 0)).quantize(Decimal('0.0001'))
                wallet.balance = (current_balance + credit_amount).quantize(Decimal('0.0001'))
                wallet.email_validation_balance = wallet.balance
                wallet.save(update_fields=['balance', 'email_validation_balance', 'updated_at'])

                payment.status = WalletRechargePayment.STATUS_SUCCESSFUL
                payment.failure_reason = ''
                payment.credited_amount = credit_amount
                payment.credited_at = timezone.now()

            payment.razorpay_payment_id = razorpay_payment_id
            payment.razorpay_signature = razorpay_signature
            payment.save(
                update_fields=[
                    'status',
                    'failure_reason',
                    'credited_amount',
                    'credited_at',
                    'razorpay_payment_id',
                    'razorpay_signature',
                    'updated_at',
                ]
            )

        updated_wallet = _get_or_create_wallet(request.user)
        return Response(
            {
                'detail': 'Payment verified and wallet credited successfully.',
                'wallet_balance': str(Decimal(str(updated_wallet.balance or 0)).quantize(Decimal('0.0001'))),
                'payment': WalletRechargePaymentSerializer(payment).data,
            }
        )


class WalletRechargePaymentListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = WalletRechargePaymentSerializer

    def get_queryset(self):
        queryset = WalletRechargePayment.objects.filter(user=self.request.user).order_by('-created_at')
        status_filter = str(self.request.query_params.get('status') or '').strip().lower()
        if status_filter in {
            WalletRechargePayment.STATUS_PENDING,
            WalletRechargePayment.STATUS_SUCCESSFUL,
            WalletRechargePayment.STATUS_FAILED,
        }:
            queryset = queryset.filter(status=status_filter)
        return queryset


class UserAPIKeyListCreateView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if _has_admin_access(request.user):
            keys = UserAPIKey.objects.select_related('user').annotate(usage_count=Count('validations')).order_by('-created_at')
            return Response(AdminUserAPIKeySerializer(keys, many=True).data)

        keys = UserAPIKey.objects.filter(user=request.user).order_by('-created_at')
        return Response(UserAPIKeySerializer(keys, many=True).data)

    def post(self, request):
        name = str(request.data.get('name') or '').strip()
        if not name:
            return Response({'detail': 'name is required'}, status=status.HTTP_400_BAD_REQUEST)

        api_key = UserAPIKey.objects.create(user=request.user, name=name, key=UserAPIKey.generate_key(), is_active=True)
        return Response(UserAPIKeySerializer(api_key).data, status=status.HTTP_201_CREATED)


class UserAPIKeyDetailView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, key_id):
        api_key = UserAPIKey.objects.filter(id=key_id, user=request.user).first()
        if not api_key:
            return Response({'detail': 'API key not found'}, status=status.HTTP_404_NOT_FOUND)

        if 'name' in request.data:
            api_key.name = str(request.data.get('name') or '').strip() or api_key.name
        if 'is_active' in request.data:
            api_key.is_active = bool(request.data.get('is_active'))
        api_key.save(update_fields=['name', 'is_active'])
        return Response(UserAPIKeySerializer(api_key).data)

    def delete(self, request, key_id):
        api_key = UserAPIKey.objects.filter(id=key_id, user=request.user).first()
        if not api_key:
            return Response({'detail': 'API key not found'}, status=status.HTTP_404_NOT_FOUND)
        api_key.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class APIEmailValidationView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        source_file = request.FILES.get('source_file') if hasattr(request, 'FILES') else None
        provider_mode = _get_email_validation_provider_mode()
        try:
            user, api_key = _authenticate_api_key_request(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_401_UNAUTHORIZED)

        available_validation_balance = _get_email_validation_wallet_balance(user)
        if available_validation_balance is None:
            return Response({'detail': 'Provider credits are unavailable for email validation.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        if available_validation_balance <= 0:
            return Response({'detail': 'No email validation credits available.'}, status=status.HTTP_402_PAYMENT_REQUIRED)

        try:
            unique_emails, file_name = _collect_validation_emails_from_file(request, enforce_request_limit=not bool(source_file))
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cost_deducted, remaining_balance = _deduct_email_validation_credits(user, len(unique_emails))
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_402_PAYMENT_REQUIRED)

        history = EmailValidationHistory.objects.create(
            user=user,
            api_key=api_key,
            source='api',
            status=EmailValidationHistory.STATUS_PENDING if source_file else EmailValidationHistory.STATUS_COMPLETED,
            email_count=len(unique_emails),
            emails_requested=unique_emails,
            results_summary={
                'safe_count': 0,
                'unsafe_count': 0,
                'provider_message_id': '',
                'provider_message_ids': [],
                'results': [],
                'queued': bool(source_file),
                'control_state': 'running',
                'processing_state': 'running' if source_file else 'completed',
                'processed_count': 0,
                'total_count': len(unique_emails),
                'progress_percent': 0 if source_file else 100,
                'elapsed_seconds': 0,
                'eta_seconds': 0,
                'started_at': timezone.now().isoformat(),
            },
            cost_deducted=cost_deducted,
            file_name=file_name,
            completed_at=None if source_file else timezone.now(),
        )
        _assign_email_validation_request_id(history)
        history_payload = EmailValidationHistorySerializer(history).data

        if source_file:
            _start_email_validation_worker(history.id)
            return Response(
                {
                    'request_id': history.request_id,
                    'status': 'pending',
                    'provider_mode': provider_mode,
                    'detail': 'File accepted and queued for background validation.',
                },
                status=status.HTTP_202_ACCEPTED,
            )

        results = _validate_email_list(unique_emails, provider_mode=provider_mode)
        client_results = [_to_client_validation_result(item) for item in results]
        safe_count = sum(1 for item in client_results if _is_safe_client_validation_result(item))
        unsafe_count = len(client_results) - safe_count
        provider_request_ids = [
            str(item.get('providerMessageId') or '').strip()
            for item in client_results
            if str(item.get('providerMessageId') or '').strip()
        ]

        history.status = EmailValidationHistory.STATUS_COMPLETED
        history.results_summary = {
            'safe_count': safe_count,
            'unsafe_count': unsafe_count,
            'provider_message_id': provider_request_ids[0] if provider_request_ids else '',
            'provider_message_ids': provider_request_ids,
            'results': client_results,
        }
        history.completed_at = timezone.now()
        history.save(update_fields=['status', 'results_summary', 'completed_at'])

        concise_response = _build_concise_api_validation_response(client_results)
        concise_response['provider_mode'] = provider_mode
        return Response(concise_response, status=status.HTTP_200_OK)


class ValidationHistoryListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EmailValidationHistorySerializer

    def get_queryset(self):
        source = str(self.request.query_params.get('source') or '').strip().lower()
        query = str(self.request.query_params.get('q') or '').strip()
        if _has_support_read_access(self.request.user):
            queryset = EmailValidationHistory.objects.all().select_related('api_key', 'user').order_by('-created_at')
        else:
            queryset = EmailValidationHistory.objects.filter(user=self.request.user).select_related('api_key', 'user').order_by('-created_at')
        if source in {'dashboard', 'api'}:
            queryset = queryset.filter(source=source)
        if query:
            filters = Q(request_id__icontains=query) | Q(status__icontains=query) | Q(file_name__icontains=query) | Q(user__email__icontains=query)
            if query.isdigit():
                filters |= Q(user__id=int(query))
            queryset = queryset.filter(filters)
        return queryset


class EmailValidationStatusView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, request_id):
        history = EmailValidationHistory.objects.select_related('api_key', 'user').filter(request_id=request_id).first()
        if not history:
            return Response({'detail': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)

        if not (_has_support_read_access(request.user) or history.user_id == request.user.id):
            return Response({'detail': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

        payload = EmailValidationHistorySerializer(history).data
        payload['worker_active'] = _is_worker_active(history.id)
        payload['processing_state'] = _get_history_processing_state(history)
        return Response(payload)


class EmailValidationControlView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, request_id):
        action = str(request.data.get('action') or '').strip().lower()
        if action not in {'start', 'pause', 'resume', 'stop', 'cancel'}:
            return Response({'detail': 'action must be one of start, pause, resume, stop, cancel'}, status=status.HTTP_400_BAD_REQUEST)

        history = EmailValidationHistory.objects.select_related('api_key', 'user').filter(request_id=request_id).first()
        if not history:
            return Response({'detail': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)

        if not (_has_support_read_access(request.user) or history.user_id == request.user.id):
            return Response({'detail': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

        if history.status in {EmailValidationHistory.STATUS_COMPLETED, EmailValidationHistory.STATUS_FAILED} and action in {'pause', 'resume', 'cancel', 'stop'}:
            return Response({'detail': f'Cannot {action} a finished request'}, status=status.HTTP_400_BAD_REQUEST)

        if action == 'pause':
            _set_history_control_state(history, 'paused')
        elif action in {'resume', 'start'}:
            _set_history_control_state(history, 'running')
            history.status = EmailValidationHistory.STATUS_PENDING
            history.completed_at = None
            history.save(update_fields=['status', 'completed_at'])
            _start_email_validation_worker(history.id)
        elif action in {'cancel', 'stop'}:
            _set_history_control_state(history, 'cancelled' if action == 'cancel' else 'stopped')
            _revoke_celery_task(history)

        history.refresh_from_db()
        payload = EmailValidationHistorySerializer(history).data
        payload['worker_active'] = _is_worker_active(history.id)
        payload['processing_state'] = _get_history_processing_state(history)
        payload['last_action'] = action
        return Response(payload)


class SendDeliverableEmailsView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        deliverable_emails = _collect_unique_emails_from_input(request.data.get('deliverable_emails'))
        if not deliverable_emails:
            return Response({'detail': 'No valid deliverable email addresses were provided.'}, status=status.HTTP_400_BAD_REQUEST)

        subject = str(request.data.get('subject') or '').strip()
        body = str(request.data.get('body') or '').strip()
        if not subject:
            return Response({'detail': 'subject is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not body:
            return Response({'detail': 'body is required'}, status=status.HTTP_400_BAD_REQUEST)

        smtp_host = str(request.data.get('smtp_host') or '').strip()
        smtp_username = str(request.data.get('smtp_username') or '').strip()
        smtp_password = str(request.data.get('smtp_password') or '').strip()
        smtp_provider = str(request.data.get('smtp_provider') or '').strip()
        from_email = str(request.data.get('from_email') or smtp_username or '').strip()

        try:
            smtp_port = int(request.data.get('smtp_port') or 0)
        except (TypeError, ValueError):
            smtp_port = 0

        use_tls = bool(request.data.get('smtp_use_tls', True))
        use_ssl = bool(request.data.get('smtp_use_ssl', False))

        if not smtp_host:
            return Response({'detail': 'smtp_host is required'}, status=status.HTTP_400_BAD_REQUEST)
        if smtp_port <= 0:
            return Response({'detail': 'smtp_port must be a valid positive integer'}, status=status.HTTP_400_BAD_REQUEST)
        if not smtp_username:
            return Response({'detail': 'smtp_username is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not smtp_password:
            return Response({'detail': 'smtp_password is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not from_email:
            return Response({'detail': 'from_email is required'}, status=status.HTTP_400_BAD_REQUEST)
        if use_tls and use_ssl:
            return Response({'detail': 'Choose either TLS or SSL, not both.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_email(from_email)
        except ValidationError:
            return Response({'detail': 'from_email is not a valid email address'}, status=status.HTTP_400_BAD_REQUEST)

        max_recipients = int(getattr(settings, 'EMAIL_MANUAL_SEND_MAX_RECIPIENTS', 1000) or 1000)
        if len(deliverable_emails) > max_recipients:
            return Response(
                {'detail': f'Maximum {max_recipients} recipients allowed per request.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        connection = get_connection(
            backend='django.core.mail.backends.smtp.EmailBackend',
            host=smtp_host,
            port=smtp_port,
            username=smtp_username,
            password=smtp_password,
            use_tls=use_tls,
            use_ssl=use_ssl,
            timeout=20,
            fail_silently=False,
        )

        try:
            connection.open()
        except Exception as exc:
            return Response(
                {'detail': f'Unable to connect or authenticate with SMTP server: {type(exc).__name__}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        sent_count = 0
        failed_recipients = []
        for recipient in deliverable_emails:
            try:
                message = EmailMessage(
                    subject=subject,
                    body=body,
                    from_email=from_email,
                    to=[recipient],
                    connection=connection,
                )
                message.send(fail_silently=False)
                sent_count += 1
            except Exception as exc:
                failed_recipients.append(
                    {
                        'email': recipient,
                        'error': type(exc).__name__,
                    }
                )

        try:
            connection.close()
        except Exception:
            pass

        return Response(
            {
                'smtp_provider': smtp_provider,
                'requested_count': len(deliverable_emails),
                'sent_count': sent_count,
                'failed_count': len(failed_recipients),
                'failed_recipients': failed_recipients,
            },
            status=status.HTTP_200_OK,
        )


class APIEmailValidationStatusView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        try:
            user, _api_key = _authenticate_api_key_request(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_401_UNAUTHORIZED)

        request_id = str(request.data.get('request_id') or '').strip()
        if not request_id:
            return Response({'detail': 'request_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        history = EmailValidationHistory.objects.select_related('api_key', 'user').filter(request_id=request_id, user=user).first()
        if not history:
            return Response({'detail': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)

        concise_response = _build_concise_api_status_response(history)
        return Response(concise_response)


class APIEmailValidationControlView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        try:
            user, _api_key = _authenticate_api_key_request(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_401_UNAUTHORIZED)

        request_id = str(request.data.get('request_id') or '').strip()
        action = str(request.data.get('action') or '').strip().lower()
        if not request_id:
            return Response({'detail': 'request_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        if action not in {'start', 'pause', 'resume', 'stop', 'cancel'}:
            return Response({'detail': 'action must be one of start, pause, resume, stop, cancel'}, status=status.HTTP_400_BAD_REQUEST)

        history = EmailValidationHistory.objects.select_related('api_key', 'user').filter(request_id=request_id, user=user).first()
        if not history:
            return Response({'detail': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)

        if history.status in {EmailValidationHistory.STATUS_COMPLETED, EmailValidationHistory.STATUS_FAILED} and action in {'pause', 'resume', 'cancel', 'stop'}:
            return Response({'detail': f'Cannot {action} a finished request'}, status=status.HTTP_400_BAD_REQUEST)

        if action == 'pause':
            _set_history_control_state(history, 'paused')
        elif action in {'resume', 'start'}:
            _set_history_control_state(history, 'running')
            history.status = EmailValidationHistory.STATUS_PENDING
            history.completed_at = None
            history.save(update_fields=['status', 'completed_at'])
            _start_email_validation_worker(history.id)
        elif action in {'cancel', 'stop'}:
            _set_history_control_state(history, 'cancelled' if action == 'cancel' else 'stopped')
            _revoke_celery_task(history)

        history.refresh_from_db()
        payload = EmailValidationHistorySerializer(history).data
        payload['worker_active'] = _is_worker_active(history.id)
        payload['processing_state'] = _get_history_processing_state(history)
        payload['last_action'] = action
        return Response(payload)


class AdminLatestValidationHistoryView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        latest_entries = []
        users = User.objects.all().only('id', 'email').order_by('-date_joined')
        for user in users:
            latest = EmailValidationHistory.objects.filter(user=user).order_by('-created_at').first()
            if latest:
                latest_entries.append({
                    'user_id': user.id,
                    'user_email': user.email,
                    'latest_history': EmailValidationHistorySerializer(latest).data,
                })
        return Response(latest_entries)


class AdminUserValidationHistoryView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EmailValidationHistorySerializer

    def get_queryset(self):
        if not _has_support_read_access(self.request.user):
            return EmailValidationHistory.objects.none()

        user_id = self.kwargs.get('user_id')
        queryset = EmailValidationHistory.objects.filter(user_id=user_id).select_related('api_key', 'user').order_by('-created_at')
        query = str(self.request.query_params.get('q') or '').strip()
        if query:
            filters = Q(request_id__icontains=query) | Q(status__icontains=query) | Q(file_name__icontains=query)
            if query.isdigit():
                filters |= Q(user__id=int(query))
            queryset = queryset.filter(filters)
        return queryset


class AdminOwnSystemValidationDiagnosticsView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        try:
            unique_emails, _ = _collect_validation_emails_from_file(request, enforce_request_limit=True)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        # Diagnostics can be expensive; keep this endpoint focused for investigations.
        max_diagnostics = 200
        if len(unique_emails) > max_diagnostics:
            return Response(
                {'detail': f'Maximum {max_diagnostics} emails are allowed per diagnostics request.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        diagnostics_rows = []
        for candidate in unique_emails:
            result, diagnostics = _validate_email_with_own_system_diagnostics(candidate)
            client_result = _to_client_validation_result(result)
            diagnostics_rows.append(
                {
                    'email': client_result.get('email'),
                    'result': client_result,
                    'diagnostics': diagnostics,
                }
            )

        return Response(
            {
                'provider_mode': 'own_system',
                'count': len(diagnostics_rows),
                'rows': diagnostics_rows,
            },
            status=status.HTTP_200_OK,
        )


class AdminCreditSettingsView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard
        setting, _ = PlatformSetting.objects.get_or_create(
            key='email_validation_cost_per_request',
            defaults={'value': '1', 'description': 'Wallet credits charged for each validated email'},
        )
        provider_setting, _ = PlatformSetting.objects.get_or_create(
            key='email_validation_provider_mode',
            defaults={'value': 'own_system', 'description': 'Provider used for email validation: own_system or zerobounce'},
        )
        return Response(
            {
                **PlatformSettingSerializer(setting).data,
                'provider_mode': str(provider_setting.value or 'own_system').strip().lower(),
                'provider_mode_description': provider_setting.description,
                'provider_mode_options': ['own_system', 'zerobounce'],
            }
        )

    def patch(self, request):
        guard = None if _has_admin_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard
        value = request.data.get('value')
        if value is None:
            return Response({'detail': 'value is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            numeric = Decimal(str(value))
            if numeric <= 0:
                raise ValueError
        except (InvalidOperation, ValueError, TypeError):
            return Response({'detail': 'value must be a positive number'}, status=status.HTTP_400_BAD_REQUEST)

        setting, _ = PlatformSetting.objects.get_or_create(
            key='email_validation_cost_per_request',
            defaults={'value': str(numeric), 'description': 'Wallet credits charged for each validated email'},
        )
        setting.value = str(numeric)
        if 'description' in request.data:
            setting.description = str(request.data.get('description') or '').strip()
        setting.save(update_fields=['value', 'description', 'updated_at'])

        provider_mode = request.data.get('provider_mode')
        if provider_mode is not None:
            normalized_mode = str(provider_mode or '').strip().lower()
            if normalized_mode not in {'own_system', 'zerobounce'}:
                return Response({'detail': 'provider_mode must be own_system or zerobounce'}, status=status.HTTP_400_BAD_REQUEST)
            provider_setting, _ = PlatformSetting.objects.get_or_create(
                key='email_validation_provider_mode',
                defaults={'value': normalized_mode, 'description': 'Provider used for email validation: own_system or zerobounce'},
            )
            provider_setting.value = normalized_mode
            if 'provider_mode_description' in request.data:
                provider_setting.description = str(request.data.get('provider_mode_description') or '').strip()
            provider_setting.save(update_fields=['value', 'description', 'updated_at'])

        provider_mode_setting = PlatformSetting.objects.filter(key='email_validation_provider_mode').first()
        return Response(
            {
                **PlatformSettingSerializer(setting).data,
                'provider_mode': str((provider_mode_setting.value if provider_mode_setting else 'own_system') or 'own_system').strip().lower(),
                'provider_mode_description': str(provider_mode_setting.description if provider_mode_setting else 'Provider used for email validation: own_system or zerobounce').strip(),
                'provider_mode_options': ['own_system', 'zerobounce'],
            }
        )


class AdminRechargeChargeSettingsView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        service_charge_percentage, tax_percentage = _get_recharge_charge_percentages()
        razorpay_config = _get_razorpay_config()
        paypal_checkout_url = _get_platform_setting_text(
            'recharge_paypal_checkout_url',
            str(getattr(settings, 'PAYPAL_CHECKOUT_URL', '') or '').strip(),
        )
        upi_vpa = _get_platform_setting_text(
            'recharge_upi_vpa',
            str(getattr(settings, 'RAZORPAY_UPI_VPA', '') or '').strip(),
        )
        upi_payee_name = _get_platform_setting_text(
            'recharge_upi_payee_name',
            str(getattr(settings, 'COMPANY_NAME', 'Bhisha') or 'Bhisha').strip(),
        )
        return Response(
            {
                'service_charge_percentage': str(service_charge_percentage),
                'tax_percentage': str(tax_percentage),
                'gateway_configured': bool(razorpay_config.get('configured')),
                'paypal_checkout_url': paypal_checkout_url,
                'paypal_enabled': bool(paypal_checkout_url),
                'upi_vpa': upi_vpa,
                'upi_payee_name': upi_payee_name or 'Bhisha',
                'upi_enabled': bool(upi_vpa),
                'updated_at': timezone.now(),
            }
        )

    def patch(self, request):
        guard = None if _has_admin_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        service_charge_percentage = request.data.get('service_charge_percentage')
        tax_percentage = request.data.get('tax_percentage')
        if service_charge_percentage is None or tax_percentage is None:
            return Response(
                {'detail': 'service_charge_percentage and tax_percentage are required'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            service_charge_decimal = Decimal(str(service_charge_percentage)).quantize(Decimal('0.01'))
            tax_decimal = Decimal(str(tax_percentage)).quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError, ValueError):
            return Response({'detail': 'Percentages must be valid numbers'}, status=status.HTTP_400_BAD_REQUEST)

        if service_charge_decimal < 0 or tax_decimal < 0:
            return Response({'detail': 'Percentages cannot be negative'}, status=status.HTTP_400_BAD_REQUEST)

        if service_charge_decimal > Decimal('100') or tax_decimal > Decimal('100'):
            return Response({'detail': 'Percentages cannot exceed 100'}, status=status.HTTP_400_BAD_REQUEST)

        service_setting, _ = PlatformSetting.objects.get_or_create(
            key='recharge_service_charge_percentage',
            defaults={'value': '0', 'description': 'Service charge percentage applied to recharge amount'},
        )
        tax_setting, _ = PlatformSetting.objects.get_or_create(
            key='recharge_tax_percentage',
            defaults={'value': '0', 'description': 'Tax percentage applied to recharge amount'},
        )

        service_setting.value = str(service_charge_decimal)
        tax_setting.value = str(tax_decimal)
        if 'service_charge_description' in request.data:
            service_setting.description = str(request.data.get('service_charge_description') or '').strip()
        if 'tax_description' in request.data:
            tax_setting.description = str(request.data.get('tax_description') or '').strip()

        service_setting.save(update_fields=['value', 'description', 'updated_at'])
        tax_setting.save(update_fields=['value', 'description', 'updated_at'])

        paypal_checkout_url = request.data.get('paypal_checkout_url')
        if paypal_checkout_url is not None:
            paypal_setting, _ = PlatformSetting.objects.get_or_create(
                key='recharge_paypal_checkout_url',
                defaults={'value': '', 'description': 'PayPal checkout URL for wallet recharge'},
            )
            paypal_setting.value = str(paypal_checkout_url or '').strip()
            paypal_setting.save(update_fields=['value', 'updated_at'])

        upi_vpa = request.data.get('upi_vpa')
        if upi_vpa is not None:
            upi_setting, _ = PlatformSetting.objects.get_or_create(
                key='recharge_upi_vpa',
                defaults={'value': '', 'description': 'UPI VPA for wallet recharge QR'},
            )
            upi_setting.value = str(upi_vpa or '').strip()
            upi_setting.save(update_fields=['value', 'updated_at'])

        upi_payee_name = request.data.get('upi_payee_name')
        if upi_payee_name is not None:
            payee_name_setting, _ = PlatformSetting.objects.get_or_create(
                key='recharge_upi_payee_name',
                defaults={'value': 'Bhisha', 'description': 'UPI payee name for wallet recharge QR'},
            )
            payee_name_setting.value = str(upi_payee_name or 'Bhisha').strip() or 'Bhisha'
            payee_name_setting.save(update_fields=['value', 'updated_at'])

        razorpay_config = _get_razorpay_config()
        updated_paypal_checkout_url = _get_platform_setting_text(
            'recharge_paypal_checkout_url',
            str(getattr(settings, 'PAYPAL_CHECKOUT_URL', '') or '').strip(),
        )
        updated_upi_vpa = _get_platform_setting_text(
            'recharge_upi_vpa',
            str(getattr(settings, 'RAZORPAY_UPI_VPA', '') or '').strip(),
        )
        updated_upi_payee_name = _get_platform_setting_text(
            'recharge_upi_payee_name',
            str(getattr(settings, 'COMPANY_NAME', 'Bhisha') or 'Bhisha').strip(),
        )

        return Response(
            {
                'service_charge_percentage': str(service_charge_decimal),
                'tax_percentage': str(tax_decimal),
                'gateway_configured': bool(razorpay_config.get('configured')),
                'paypal_checkout_url': updated_paypal_checkout_url,
                'paypal_enabled': bool(updated_paypal_checkout_url),
                'upi_vpa': updated_upi_vpa,
                'upi_payee_name': updated_upi_payee_name or 'Bhisha',
                'upi_enabled': bool(updated_upi_vpa),
                'updated_at': max(service_setting.updated_at, tax_setting.updated_at),
            }
        )


class AdminUserWalletCreditsView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, user_id):
        guard = None if _has_admin_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        wallet = _get_or_create_wallet(target_user)
        current_sms = Decimal(str(wallet.balance or 0)).quantize(Decimal('0.0001'))

        add_sms_raw = request.data.get('add_message_credits', request.data.get('message_credits_delta', '0'))
        add_email_raw = request.data.get('add_email_validation_credits', request.data.get('email_validation_credits_delta', '0'))

        try:
            add_sms = Decimal(str(add_sms_raw or '0')).quantize(Decimal('0.0001'))
            add_email = Decimal(str(add_email_raw or '0')).quantize(Decimal('0.0001'))
        except (InvalidOperation, TypeError, ValueError):
            return Response({'detail': 'Credit values must be valid numbers'}, status=status.HTTP_400_BAD_REQUEST)

        if add_sms < 0 or add_email < 0:
            return Response({'detail': 'Credit values must be zero or positive numbers'}, status=status.HTTP_400_BAD_REQUEST)

        total_add = (add_sms + add_email).quantize(Decimal('0.0001'))
        wallet.balance = max(Decimal('0.0000'), (current_sms + total_add).quantize(Decimal('0.0001')))
        wallet.email_validation_balance = wallet.balance
        wallet.save(update_fields=['balance', 'email_validation_balance', 'updated_at'])

        return Response(
            {
                'user_id': target_user.id,
                'user_email': target_user.email,
                'message_credits': str(wallet.balance),
                'email_validation_credits': str(wallet.balance),
                'added_message_credits': str(add_sms),
                'added_email_validation_credits': str(add_email),
                'total_credits_added': str(total_add),
            },
            status=status.HTTP_200_OK,
        )


class SenderIdRequestView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SenderIdRequestSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        return SenderIdRequest.objects.filter(user=self.request.user).order_by('-created_at')


class AdminSenderIdRequestListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SenderIdRequestAdminSerializer

    def get_queryset(self):
        if not _has_support_read_access(self.request.user):
            return SenderIdRequest.objects.none()

        queryset = SenderIdRequest.objects.select_related('user').all().order_by('-created_at')
        query = str(self.request.query_params.get('q') or '').strip()
        status_filter = str(self.request.query_params.get('status') or '').strip()
        country_filter = str(self.request.query_params.get('destination_country') or '').strip()
        use_case_filter = str(self.request.query_params.get('primary_use_case') or '').strip()
        industry_filter = str(self.request.query_params.get('industry_sector_type') or '').strip()

        if query:
            queryset = queryset.filter(
                Q(full_name__icontains=query)
                | Q(email__icontains=query)
                | Q(contact_number__icontains=query)
                | Q(required_sender_id__icontains=query)
                | Q(company_name__icontains=query)
                | Q(destination_country__icontains=query)
                | Q(message_content__icontains=query)
                | Q(user__email__icontains=query)
            )

        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if country_filter:
            queryset = queryset.filter(destination_country__iexact=country_filter)
        if use_case_filter:
            queryset = queryset.filter(primary_use_case=use_case_filter)
        if industry_filter:
            queryset = queryset.filter(industry_sector_type=industry_filter)

        return queryset


class AdminSenderIdRequestDetailView(generics.RetrieveUpdateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SenderIdRequestAdminSerializer

    def get_queryset(self):
        if not _has_support_read_access(self.request.user):
            return SenderIdRequest.objects.none()
        return SenderIdRequest.objects.select_related('user').all()

    def update(self, request, *args, **kwargs):
        if not _has_admin_access(request.user):
            return Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not _has_admin_access(request.user):
            return Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)


class RequestStatusSearchView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin or employee access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        query = str(request.query_params.get('q') or '').strip()
        if not query:
            return Response({'detail': 'q is required'}, status=status.HTTP_400_BAD_REQUEST)

        sms_queryset = SMSMessage.objects.filter(
            Q(message_id__icontains=query)
            | Q(provider_message_id__icontains=query)
            | Q(batch_reference__icontains=query)
        )
        if query.isdigit():
            sms_queryset = sms_queryset.filter(Q(sender__id=int(query)) | Q(recipient_user__id=int(query)))
        sms_queryset = sms_queryset.order_by('-created_at')[:100]

        validation_queryset = EmailValidationHistory.objects.filter(
            Q(request_id__icontains=query)
            | Q(file_name__icontains=query)
            | Q(user__email__icontains=query)
        )
        if query.isdigit():
            validation_queryset = validation_queryset.filter(Q(user__id=int(query)))
        validation_queryset = validation_queryset.select_related('api_key', 'user').order_by('-created_at')[:100]

        return Response(
            {
                'query': query,
                'sms': SMSMessageSerializer(sms_queryset, many=True).data,
                'email_validations': EmailValidationHistorySerializer(validation_queryset, many=True).data,
            },
            status=status.HTTP_200_OK,
        )


class AdminAllAPIKeysView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = None if _has_support_read_access(request.user) else Response({'detail': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)
        if guard:
            return guard

        keys = UserAPIKey.objects.select_related('user').order_by('-created_at')
        return Response(AdminUserAPIKeySerializer(keys, many=True).data)


class EmployeeSignupView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = EmployeeSignupSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email'].strip().lower()
        existing_user = _find_user_by_email(email)
        if existing_user and existing_user.is_active:
            return Response({'detail': 'Employee account already exists'}, status=status.HTTP_400_BAD_REQUEST)

        first_name = serializer.validated_data['first_name'].strip()
        phone_number = serializer.validated_data['phone_number'].strip()
        password = serializer.validated_data['password']
        department = serializer.validated_data.get('department', '').strip()

        with transaction.atomic():
            user = existing_user
            if not user:
                user = User.objects.create(
                    username=email,
                    email=email,
                    first_name=first_name,
                    phone_number=phone_number,
                    is_active=False,
                )
            user.first_name = first_name
            user.phone_number = phone_number
            user.is_active = False
            user.set_password(password)
            user.otp_code = generate_otp()
            user.otp_created = timezone.now()
            user.save()

            employee, _ = Employee.objects.get_or_create(user=user)
            employee.department = department
            employee.status = Employee.STATUS_PENDING
            employee.employee_otp_verified = False
            employee.admin_otp_verified = False
            employee.admin_otp = generate_otp()
            employee.admin_otp_created = timezone.now()
            employee.save()

        employee_sent = send_otp_via_email(user, user.otp_code)

        primary_admin_email = str(getattr(settings, 'PRIMARY_ADMIN_EMAIL', '') or '').strip()
        admin_sent = False
        if primary_admin_email:
            admin_sent = send_otp_via_email(primary_admin_email, employee.admin_otp)

        if not employee_sent:
            return Response({'detail': 'Could not send employee OTP. Please try again.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        if not admin_sent:
            return Response({'detail': 'Employee OTP sent, but admin OTP delivery failed. Please retry later.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        return Response(
            {
                'detail': 'Employee signup initiated. Verify both employee OTP and admin OTP.',
                'requires_employee_otp': True,
                'requires_admin_otp': True,
                'email': email,
            },
            status=status.HTTP_200_OK,
        )


class EmployeeVerifyDualOTPView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = EmployeeDualOTPVerifySerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email'].strip().lower()
        employee_otp = serializer.validated_data['employee_otp'].strip()
        admin_otp = serializer.validated_data['admin_otp'].strip()

        user = _find_user_by_email(email)
        if not user:
            return Response({'detail': 'Employee account not found'}, status=status.HTTP_404_NOT_FOUND)

        employee = Employee.objects.filter(user=user).first()
        if not employee:
            return Response({'detail': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)

        if not otp_is_valid(user, employee_otp):
            return Response({'detail': 'Invalid or expired employee OTP'}, status=status.HTTP_400_BAD_REQUEST)

        if not employee.admin_otp or not employee.admin_otp_created:
            return Response({'detail': 'Admin OTP is not available. Please restart employee signup.'}, status=status.HTTP_400_BAD_REQUEST)

        if str(employee.admin_otp).strip() != str(admin_otp).strip():
            return Response({'detail': 'Invalid or expired admin OTP'}, status=status.HTTP_400_BAD_REQUEST)

        if timezone.now() > employee.admin_otp_created + timedelta(minutes=10):
            return Response({'detail': 'Invalid or expired admin OTP'}, status=status.HTTP_400_BAD_REQUEST)

        user.is_active = True
        user.otp_code = None
        user.otp_created = None
        user.save(update_fields=['is_active', 'otp_code', 'otp_created'])

        employee.employee_otp_verified = True
        employee.admin_otp_verified = True
        employee.status = Employee.STATUS_ACTIVE
        employee.admin_otp = ''
        employee.admin_otp_created = None
        employee.save(update_fields=['employee_otp_verified', 'admin_otp_verified', 'status', 'admin_otp', 'admin_otp_created', 'updated_at'])

        return Response({'detail': 'Employee verification completed successfully'}, status=status.HTTP_200_OK)


class EmployeeLoginView(generics.GenericAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = EmployeeLoginSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email'].strip().lower()
        password = serializer.validated_data['password']

        user = _find_user_by_email(email)
        if not user or not user.check_password(password):
            return Response({'detail': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        employee = Employee.objects.filter(user=user, status=Employee.STATUS_ACTIVE).first()
        if not employee:
            return Response({'detail': 'Employee account is not active'}, status=status.HTTP_403_FORBIDDEN)

        refresh = RefreshToken.for_user(user)
        return Response(
            {
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'role': 'employee',
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'first_name': user.first_name,
                    'department': employee.department,
                    'is_employee': True,
                    'status': employee.status,
                },
            },
            status=status.HTTP_200_OK,
        )


class AdminEmployeeListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        guard = _primary_admin_guard(request)
        if guard:
            return guard

        employees = Employee.objects.select_related('user').order_by('-created_at')
        payload = []
        for item in employees:
            payload.append(
                {
                    'employee_id': item.id,
                    'user_id': item.user.id,
                    'email': item.user.email,
                    'name': item.user.first_name,
                    'department': item.department,
                    'status': item.status,
                    'admin_otp_verified': item.admin_otp_verified,
                    'employee_otp_verified': item.employee_otp_verified,
                    'created_at': item.created_at,
                }
            )
        return Response(payload, status=status.HTTP_200_OK)

